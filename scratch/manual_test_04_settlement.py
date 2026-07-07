"""
Manuel Test 04 — Settlement Engine: Test 2 (iSolar Curve) ve Test 3 (GAOSB)
dosyalarından 2026-07-06 mahsup hesabı + Excel raporu.

Başarı kriterleri:
  - 24 satır eşleşme (inner join)
  - Gündüz saatlerinde üretim > 0
  - Mahsup mantığı tutarlı (tüketim = mahsup + çekiş; üretim = mahsup + fazla satış)
  - Excel dosyası oluşmalı
"""
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.settlement.engine import SettlementEngine  # noqa: E402
from app.settlement.report_writer import SettlementReportWriter  # noqa: E402

# Girdiler: en yeni Curve dosyası (eski ikisi bug'lı koşulardan, 07.07 verili)
isolar_dir = ROOT / "outputs" / "manual_tests" / "02_isolar_curve"
gaosb_dir = ROOT / "outputs" / "manual_tests" / "03_gaosb"
isolar_file = max(isolar_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime)
gaosb_file = max(gaosb_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime)

output_dir = ROOT / "outputs" / "manual_tests" / "04_settlement"
output_dir.mkdir(parents=True, exist_ok=True)
output_path = output_dir / "mahsup_20260706.xlsx"

print("TEST 4: Settlement Engine + Report Writer")
print("=" * 50)
print(f"iSolar girdi: {isolar_file.name}")
print(f"GAOSB girdi:  {gaosb_file.name}")

engine = SettlementEngine()

print("\n1. Mahsup hesaplanıyor...")
settlements = engine.calculate(isolar_file=isolar_file, gaosb_file=gaosb_file)
n = len(settlements)
print(f"   Eşleşen satır sayısı: {n}")

toplam_uretim = sum(s.production_kwh for s in settlements)
toplam_tuketim = sum(s.consumption_kwh for s in settlements)
toplam_mahsup = sum(s.settled_kwh for s in settlements)
toplam_cekis = sum(s.grid_import_kwh for s in settlements)
toplam_satis = sum(s.grid_export_kwh for s in settlements)

print(f"   Toplam üretim:      {toplam_uretim:>12,.1f} kWh")
print(f"   Toplam tüketim:     {toplam_tuketim:>12,.1f} kWh")
print(f"   Toplam mahsup:      {toplam_mahsup:>12,.1f} kWh")
print(f"   Toplam çekiş:       {toplam_cekis:>12,.1f} kWh")
print(f"   Toplam fazla satış: {toplam_satis:>12,.1f} kWh")

# Gündüz üretimi (10:00-15:00 arası saatler)
gunduz = [s for s in settlements if 10 <= int(s.timestamp[11:13]) <= 15]
gunduz_uretim = sum(s.production_kwh for s in gunduz)
print(f"   Gündüz (10-15h) üretim: {gunduz_uretim:,.1f} kWh")
print("\n   Örnek satırlar (03h gece / 12h öğle):")
for s in settlements:
    if s.timestamp[11:13] in ("03", "12"):
        print(f"     {s.timestamp} | üretim={s.production_kwh:9.1f} | tüketim={s.consumption_kwh:9.1f} "
              f"| mahsup={s.settled_kwh:9.1f} | çekiş={s.grid_import_kwh:9.1f} | satış={s.grid_export_kwh:9.1f}")

# Satır bazında mahsup mantığı tutarlılığı
tutarli = all(
    abs(s.consumption_kwh - (s.settled_kwh + s.grid_import_kwh)) < 0.01
    and abs(s.production_kwh - (s.settled_kwh + s.grid_export_kwh)) < 0.01
    and s.settled_kwh <= min(s.production_kwh, s.consumption_kwh) + 0.01
    for s in settlements
)

print("\n2. Excel raporu yazılıyor...")
writer = SettlementReportWriter()
result_path = writer.write(settlements, output_path)
excel_ok = Path(result_path).exists() and Path(result_path).stat().st_size > 0
print(f"   Rapor: {result_path} ({Path(result_path).stat().st_size if excel_ok else 0} bayt)")

# v2: GES Kırılımı sayfalı rapor
print("\n2b. GES Kırılımı sayfalı rapor (v2) yazılıyor...")
output_path_v2 = output_dir / "mahsup_20260706_v2.xlsx"
isolar_df = engine.load_isolar_curve(isolar_file)
result_path_v2 = writer.write(settlements, output_path_v2, isolar_df=isolar_df)

import openpyxl
wb = openpyxl.load_workbook(result_path_v2)
sheets = wb.sheetnames
print(f"   Rapor: {result_path_v2.name} ({result_path_v2.stat().st_size} bayt)")
print(f"   Sayfalar: {sheets}")
two_sheets_ok = sheets == ["Mahsuplaşma Raporu", "GES Kırılımı"]
if "GES Kırılımı" in sheets:
    ws2 = wb["GES Kırılımı"]
    header_row = [c.value for c in ws2[1]]
    print(f"   GES Kırılımı başlıkları: {header_row}")
    print(f"   GES Kırılımı satır sayısı (başlık+veri+toplam): {ws2.max_row}")
    print("   İlk veri satırı:", [c.value for c in ws2[2]][:5], "...")
    print("   TOPLAM satırı:", [c.value for c in ws2[ws2.max_row]][:4], "...")
wb.close()

k1 = n == 24
k2 = gunduz_uretim > 0
k3 = tutarli
k4 = excel_ok
k5 = two_sheets_ok

print("\nBAŞARI KRİTERLERİ:")
print(f"  24 satır eşleşme ({n}):        " + ("GECTI [OK]" if k1 else "KALDI [FAIL]"))
print("  Gündüz üretim > 0:            " + ("GECTI [OK]" if k2 else "KALDI [FAIL]"))
print("  Mahsup mantığı tutarlı:       " + ("GECTI [OK]" if k3 else "KALDI [FAIL]"))
print("  Excel dosyası oluştu:         " + ("GECTI [OK]" if k4 else "KALDI [FAIL]"))
print("  v2: iki sayfa (GES Kırılımı): " + ("GECTI [OK]" if k5 else "KALDI [FAIL]"))
print("\nSONUÇ: TEST 4 " + ("GEÇTİ [OK]" if (k1 and k2 and k3 and k4 and k5) else "KALDI [FAIL]"))
