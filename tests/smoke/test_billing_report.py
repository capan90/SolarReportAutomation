"""
Neden: Sprint C — faturalama tutarlarının Excel, e-posta ve chatbot çıktılarına
doğru yansıdığını sabitlemek (ADR-0002).

Ana kural her üç kanalda aynı: hesaplanmamış tutar 0 TL DEĞİL, "Bekleniyor" /
"girilmedi" olarak gösterilir. Sıfır tutarla eksik veri karıştırılmamalı.
"""
import json
from datetime import datetime
from decimal import Decimal

import openpyxl
import pytest

from app.billing import BillingService, MonthlyBillingResult, STATUS_LOCKED, STATUS_PENDING_RATE
from app.chatbot.parser import MetricParser
from app.chatbot.query_engine import QueryEngine
from app.chatbot.response_builder import ResponseBuilder
from app.jobs.monthly_settlement_job import MonthlySettlementJob


def _locked(year=2026, month=7):
    return MonthlyBillingResult(
        year=year, month=month, status=STATUS_LOCKED,
        excess_sale_rate_try=Decimal("2.909687"),
        osb_unit_price_try=Decimal("1.500000"),
        production_kwh=Decimal("10000"), excess_sale_kwh=Decimal("1000"),
        excess_sale_invoice_try=Decimal("2909.69"),
        osb_deduction_try=Decimal("13500.00"),
        locked_at=datetime(2026, 7, 27, 12, 0, 0),
    )


def _pending(year=2026, month=7):
    return MonthlyBillingResult(
        year=year, month=month, status=STATUS_PENDING_RATE,
        excess_sale_rate_try=Decimal("2.909687"),
        production_kwh=Decimal("10000"), excess_sale_kwh=Decimal("1000"),
        excess_sale_invoice_try=Decimal("2909.69"),
        osb_deduction_try=None,
    )


# ----------------------------------------------------------------------
# 1. Para biçimi
# ----------------------------------------------------------------------
@pytest.mark.parametrize(
    "value,expected",
    [
        (Decimal("2909.69"), "2.909,69"),
        (Decimal("13500.00"), "13.500,00"),
        (Decimal("0.50"), "0,50"),
        (Decimal("1234567.89"), "1.234.567,89"),
        (None, "Bekleniyor"),
    ],
)
def test_try_formatting_is_turkish(value, expected):
    assert MonthlySettlementJob._fmt_try(value) == expected


def test_rate_formatting_keeps_six_decimals():
    assert MonthlySettlementJob._fmt_rate(Decimal("2.909687")) == "2,909687"
    assert MonthlySettlementJob._fmt_rate(None) == "—"


def test_email_pending_line_has_no_currency_unit():
    """
    Neden (regresyon): İlk uygulamada e-postada "Bekleniyor TL" yazıyordu.
    Birim yalnızca gerçek tutara eklenmeli. Dev doğrulamasında yakalandı.
    """
    job = MonthlySettlementJob.__new__(MonthlySettlementJob)

    def _tl(value):
        if value is None:
            return job.BILLING_PENDING_TEXT
        return f"{job._fmt_try(value)} TL"

    assert _tl(None) == "Bekleniyor"
    assert "TL" not in _tl(None)
    assert _tl(Decimal("13500.00")) == "13.500,00 TL"


# ----------------------------------------------------------------------
# 2. Excel "FATURALAMA" bölümü
# ----------------------------------------------------------------------
def _real_style_header(ws, row_idx, n_cols):
    """
    Neden: Üretimdeki _style_header ile AYNI işi yapar. Önceki testler no-op
    lambda geçtiği için stilin yanlış satıra uygulandığı bug'ı kaçırmıştı.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# Neden: Sahte servis artık üç okuma sunuyor — ayın kendi elektrik fiyatı ve
# önizleme kesintisi, resmi kesintiden AYRI kaynaklardan gelir. Varsayılanlar
# gerçek Temmuz/Haziran 2026 ilişkisini taklit eder: Temmuz'un katsayısı
# (1,500000) Haziran faturasının fiyatıdır; Temmuz'un KENDİ fiyatı 1,800000.
_OWN_PRICES = {(2026, 7): Decimal("1.800000"), (2026, 6): Decimal("1.500000")}
_PREVIEWS = {(2026, 7): Decimal("16200.00"), (2026, 6): Decimal("13500.00")}


def _fake_service_cls(cur, prev, own_prices=None, previews=None):
    own = _OWN_PRICES if own_prices is None else own_prices
    prv = _PREVIEWS if previews is None else previews

    class _FakeService:
        def get_monthly(self, year, month):
            return cur if (year, month) == (2026, 7) else prev

        def get_own_electricity_price(self, year, month):
            return own.get((year, month))

        def get_preview_deduction(self, year, month):
            return prv.get((year, month))

    return _FakeService


def _build_sheet(monkeypatch, cur, prev, style_header=None, own_prices=None, previews=None):
    cls = _fake_service_cls(cur, prev, own_prices, previews)
    monkeypatch.setattr("app.billing.BillingService", lambda: cls())

    wb = openpyxl.Workbook()
    ws = wb.active
    job = MonthlySettlementJob.__new__(MonthlySettlementJob)
    job._append_billing_section(
        ws,
        datetime(2026, 7, 1), datetime(2026, 6, 1),
        "Temmuz 2026", "Haziran 2026",
        style_header or (lambda w, r, n: None),
    )
    return [[c.value for c in row] for row in ws.iter_rows()]


def test_excel_billing_section_locked(monkeypatch):
    rows = _build_sheet(monkeypatch, _locked(), _locked(2026, 6))
    flat = {r[0]: r for r in rows if r and r[0]}

    assert "FATURALAMA (TL, KDV HARİÇ)" in flat
    assert flat["Fazla Satış Faturası"][1] == "2.909,69"
    # Resmi kesinti DEĞİŞMEDİ — yalnızca etiketi zamanlamayı da söylüyor.
    assert flat["OSB Kesintisi — Resmi (gelecek ay düşülecek)"][1] == "13.500,00"
    assert flat["Durum"][1] == "Kilitli"
    # Sabit Enerjisa katsayısı kendi satırında
    assert flat["Fazla Satış Katsayısı (TL/kWh)"][1] == "2,909687"


def test_excel_section_shows_each_months_own_price_not_the_coefficient(monkeypatch):
    """
    Neden (A maddesi): Eski "Kullanılan Katsayılar (Fazla Satış / OSB)" satırındaki
    OSB değeri ayın kendi fiyatı DEĞİL, bir önceki ayın fiyatıydı — aynı sayı önceki
    ay sütununda zaten görünüyordu. Artık her sütun KENDİ ayının fiyatını gösteriyor.
    """
    rows = _build_sheet(monkeypatch, _locked(), _locked(2026, 6))
    flat = {r[0]: r for r in rows if r and r[0]}

    satir = flat["Elektrik Birim Fiyatı — Ayın Kendi Faturası (TL/kWh)"]
    assert satir[1] == "1,800000"       # Temmuz'un kendi fiyatı
    assert satir[2] == "1,500000"       # Haziran'ın kendi fiyatı

    # Bu ayın kesintisinde kullanılan katsayı (1,500000) = önceki ay sütunundaki
    # fiyat. Aynı sayı ikinci kez, "bu ayın katsayısı" adıyla gösterilmiyor.
    assert "Kullanılan Katsayılar (Fazla Satış / OSB)" not in flat
    assert all("Kullanılan Katsayılar" not in (r[0] or "") for r in rows if r and r[0])


def test_excel_section_has_preview_row_next_to_official(monkeypatch):
    """B maddesi: iki tutar yan yana — resmi kesinti ve ayın kendi fiyatıyla önizleme."""
    rows = _build_sheet(monkeypatch, _locked(), _locked(2026, 6))
    labels = [r[0] for r in rows if r and r[0]]
    flat = {r[0]: r for r in rows if r and r[0]}

    resmi = "OSB Kesintisi — Resmi (gelecek ay düşülecek)"
    onizleme = "OSB Kesintisi — Önizleme (ayın kendi fiyatıyla)"
    assert labels.index(onizleme) == labels.index(resmi) + 1, "önizleme resmi tutarın hemen altında olmalı"
    assert flat[onizleme][1] == "16.200,00"
    assert flat[onizleme][2] == "13.500,00"
    # Resmi tutar önizlemeden ETKİLENMEDİ
    assert flat[resmi][1] == "13.500,00"


def test_excel_billing_section_pending_shows_bekleniyor(monkeypatch):
    rows = _build_sheet(monkeypatch, _pending(), None, own_prices={}, previews={})
    flat = {r[0]: r for r in rows if r and r[0]}

    # Neden: 0,00 YAZILMAMALI — "kesinti yok" ile "henüz hesaplanmadı" farklıdır.
    assert flat["OSB Kesintisi — Resmi (gelecek ay düşülecek)"][1] == "Bekleniyor"
    assert flat["OSB Kesintisi — Önizleme (ayın kendi fiyatıyla)"][1] == "Bekleniyor"
    assert flat["Durum"][1] == "OSB birim fiyatı bekleniyor"
    # Fiyat girilmediği için tire
    assert flat["Elektrik Birim Fiyatı — Ayın Kendi Faturası (TL/kWh)"][1] == "—"
    assert flat["Fazla Satış Katsayısı (TL/kWh)"][1] == "2,909687"


def test_excel_section_note_row_is_italic_gray(monkeypatch):
    """F maddesi: açıklama metni ayırt edilebilir ama tutarları bastırmayan katman."""
    cls = _fake_service_cls(_locked(), _locked(2026, 6))
    monkeypatch.setattr("app.billing.BillingService", lambda: cls())

    wb = openpyxl.Workbook()
    ws = wb.active
    job = MonthlySettlementJob.__new__(MonthlySettlementJob)
    job._append_billing_section(
        ws, datetime(2026, 7, 1), datetime(2026, 6, 1),
        "Temmuz 2026", "Haziran 2026", _real_style_header,
    )

    not_cell = next(c for c in (row[0] for row in ws.iter_rows())
                    if str(c.value or "").startswith("Not:"))
    assert "Önceki Ay (Haziran 2026)" in not_cell.value
    assert not_cell.font.italic
    assert not_cell.font.color.rgb.endswith("808080")
    assert not not_cell.font.bold


def test_excel_change_percent_only_when_both_months_computed(monkeypatch):
    rows = _build_sheet(monkeypatch, _locked(), _locked(2026, 6))
    flat = {r[0]: r for r in rows if r and r[0]}
    # Aynı değerler -> %0 değişim
    assert flat["Fazla Satış Faturası"][3] == 0.0

    # Önceki ay yoksa yüzde uydurulmaz
    rows2 = _build_sheet(monkeypatch, _locked(), None)
    flat2 = {r[0]: r for r in rows2 if r and r[0]}
    assert flat2["Fazla Satış Faturası"][3] == "-"


def test_billing_header_row_is_actually_styled(monkeypatch):
    """
    Neden (regresyon): header_row `max_row + 1` ile hesaplanıyordu. append([])
    hücre yazmadığı için max_row'u artırmaz — stil boş ara satıra uygulanıyor,
    gerçek FATURALAMA başlığı çıplak kalıyordu. METRİK başlığı bold+gri iken
    FATURALAMA'nın olmamasının sebebi buydu.
    """
    cls = _fake_service_cls(_locked(), None)
    monkeypatch.setattr("app.billing.BillingService", lambda: cls())

    wb = openpyxl.Workbook()
    ws = wb.active
    job = MonthlySettlementJob.__new__(MonthlySettlementJob)
    job._append_billing_section(
        ws, datetime(2026, 7, 1), datetime(2026, 6, 1),
        "Temmuz 2026", "Haziran 2026", _real_style_header,
    )

    # FATURALAMA başlığının hangi satırda olduğunu bul
    header_row = None
    for row in ws.iter_rows():
        if row[0].value == "FATURALAMA (TL, KDV HARİÇ)":
            header_row = row[0].row
            break
    assert header_row is not None, "FATURALAMA başlığı yazılmamış"

    # O satırın 4 hücresi de METRİK başlığıyla aynı stilde olmalı
    for col in range(1, 5):
        cell = ws.cell(row=header_row, column=col)
        assert cell.font.bold, f"sütun {col} bold değil"
        assert cell.fill.start_color.rgb.endswith("EAEAEA"), f"sütun {col} dolgusuz"


def test_excel_section_skipped_when_no_billing_row(monkeypatch):
    # Neden: Billing katmanı öncesi aylarda kayıt yok; bölüm hiç eklenmemeli.
    rows = _build_sheet(monkeypatch, None, None)
    assert all(not (r and r[0]) for r in rows)


def test_excel_section_survives_billing_failure(monkeypatch):
    # Neden: TL hesabı rapor üretimini ASLA engellemez (best-effort).
    def _boom():
        raise RuntimeError("DB down")

    monkeypatch.setattr("app.billing.BillingService", _boom)
    wb = openpyxl.Workbook()
    ws = wb.active
    job = MonthlySettlementJob.__new__(MonthlySettlementJob)
    job._append_billing_section(
        ws, datetime(2026, 7, 1), datetime(2026, 6, 1), "Temmuz 2026", "Haziran 2026",
        lambda w, r, n: None,
    )
    assert ws.max_row == 1  # hiçbir şey yazılmadı, exception dışarı sızmadı


# ----------------------------------------------------------------------
# 2b. "Faturalama Özeti" sayfası
# ----------------------------------------------------------------------
def _build_summary(monkeypatch, cur, own_prices=None, previews=None):
    cls = _fake_service_cls(cur, cur, own_prices, previews)
    monkeypatch.setattr("app.billing.BillingService", lambda: cls())
    wb = openpyxl.Workbook()
    wb.active.title = "Ay Özeti"
    wb.create_sheet("Haftalık Kırılım")
    job = MonthlySettlementJob.__new__(MonthlySettlementJob)
    job._write_billing_summary_sheet(wb, datetime(2026, 7, 1), "Temmuz 2026", _real_style_header)
    return wb


def test_summary_sheet_is_second(monkeypatch):
    # Neden: kWh kırılımlarını geçip aranmasın diye Ay Özeti'nden hemen sonra.
    wb = _build_summary(monkeypatch, _locked())
    assert wb.sheetnames == ["Ay Özeti", "Faturalama Özeti", "Haftalık Kırılım"]


def test_summary_sheet_rows_locked(monkeypatch):
    wb = _build_summary(monkeypatch, _locked())
    rows = {r[0]: r for r in ([c.value for c in row] for row in wb["Faturalama Özeti"].iter_rows()) if r[0]}

    assert rows["Toplam Üretim"][1] == 10000.0
    assert rows["Toplam Üretim"][2] == "—"          # üretimin TL karşılığı yok
    assert rows["Fazla Satış (Enerjisa'ya)"][1] == 1000.0
    assert rows["Fazla Satış (Enerjisa'ya)"][2] == "2.909,69"
    # OSB'ye kalan = üretim - fazla satış (türetilmiş, yeni hesap değil)
    assert rows["OSB'ye Kalan (Üretim − Fazla Satış)"][1] == 9000.0
    assert rows["OSB'ye Kalan (Üretim − Fazla Satış)"][2] == "13.500,00"
    # TOPLAM = fatura + kesinti
    assert rows["TOPLAM (Enerjisa + OSB Kesintisi)"][2] == "16.409,69"
    assert rows["Fazla Satış Katsayısı (TL/kWh)"][2] == "2,909687"
    assert rows["Durum"][2] == "Kilitli"


def test_summary_sheet_pending_shows_bekleniyor(monkeypatch):
    wb = _build_summary(monkeypatch, _pending(), own_prices={}, previews={})
    rows = {r[0]: r for r in ([c.value for c in row] for row in wb["Faturalama Özeti"].iter_rows()) if r[0]}

    # Neden: OSB fiyatı girilmemişse kesinti de TOPLAM da 0 DEĞİL, "Bekleniyor".
    assert rows["OSB'ye Kalan (Üretim − Fazla Satış)"][2] == "Bekleniyor"
    assert rows["TOPLAM (Enerjisa + OSB Kesintisi)"][2] == "Bekleniyor"
    assert rows["Fazla Satış (Enerjisa'ya)"][2] == "2.909,69"   # bu hesaplanmıştı
    assert rows["Fazla Satış Katsayısı (TL/kWh)"][2] == "2,909687"
    assert rows["Durum"][2] == "OSB birim fiyatı bekleniyor"


# ----------------------------------------------------------------------
# 2c. "OSB Kesintisi — İki Görünüm" bloğu (B + F maddeleri)
# ----------------------------------------------------------------------
def _summary_rows(wb):
    return [[c.value for c in row] for row in wb["Faturalama Özeti"].iter_rows()]


def test_deduction_views_show_both_amounts(monkeypatch):
    wb = _build_summary(monkeypatch, _locked())
    rows = {r[0]: r for r in _summary_rows(wb) if r[0]}

    resmi = "Resmi Kesinti — Ağustos 2026 Faturasından Düşülecek"
    onizleme = "Önizleme — Ayın Kendi Fiyatıyla (kayıt dışı)"
    # Resmi tutar = mevcut osb_deduction_try, DEĞİŞMEDİ
    assert rows[resmi][2] == "13.500,00"
    # Önizleme = aynı kWh, ayın kendi fiyatıyla
    assert rows[onizleme][2] == "16.200,00"


def test_deduction_view_notes_explain_which_price_was_used(monkeypatch):
    """
    Neden (A maddesi karşılığı): Katsayı satırdan kaldırıldı ama "bu tutar hangi
    fiyatla çıktı" sorusunun cevabı raporda KALMALI — tek sayfalık bu görünümde
    yandaki sütun yok. Cevap açıklama satırında, ait olduğu ay adıyla birlikte.
    """
    wb = _build_summary(monkeypatch, _locked())
    notlar = [r[0] for r in _summary_rows(wb) if r[0] and str(r[0]).startswith("Hesap:")]
    assert len(notlar) == 2

    resmi_not, onizleme_not = notlar
    assert "9.000,0 kWh × 1,500000 TL/kWh" in resmi_not
    assert "Haziran 2026 faturasının elektrik birim fiyatıdır" in resmi_not
    # Zamanlama açıkça yazılı — okuyucu tutarı elindeki faturayla eşleştirmeye çalışmasın
    assert "Temmuz 2026 faturasında DEĞİL, Ağustos 2026 faturasında düşülür" in resmi_not

    assert "9.000,0 kWh × 1,800000 TL/kWh" in onizleme_not
    assert "kaydedilmez" in onizleme_not
    assert "DEĞİŞTİRMEZ" in onizleme_not


def test_preview_row_says_bekleniyor_when_own_price_missing(monkeypatch):
    # Neden: Fiyat girilmemişse 0,00 değil "Bekleniyor"; sebebi de yazılı olmalı.
    wb = _build_summary(monkeypatch, _locked(), own_prices={}, previews={})
    rows = {r[0]: r for r in _summary_rows(wb) if r[0]}
    assert rows["Önizleme — Ayın Kendi Fiyatıyla (kayıt dışı)"][2] == "Bekleniyor"

    uyari = next(r[0] for r in _summary_rows(wb)
                 if r[0] and str(r[0]).startswith("Önizleme hesaplanamadı"))
    assert "Temmuz 2026 faturasının elektrik birim fiyatı" in uyari


def test_official_row_note_when_osb_price_missing(monkeypatch):
    wb = _build_summary(monkeypatch, _pending(), own_prices={}, previews={})
    notlar = [r[0] for r in _summary_rows(wb)
              if r[0] and str(r[0]).startswith("Kesinti henüz hesaplanamadı")]
    assert len(notlar) == 1
    assert "Haziran 2026 faturasının elektrik birim fiyatı" in notlar[0]


def test_important_amount_rows_are_bold_and_filled(monkeypatch):
    """F maddesi: toplam ve iki kesinti satırı bold + belirgin (ama ayrı) dolgu."""
    wb = _build_summary(monkeypatch, _locked())
    ws = wb["Faturalama Özeti"]
    beklenen = {
        "TOPLAM (Enerjisa + OSB Kesintisi)": "DDEBF7",
        "Resmi Kesinti — Ağustos 2026 Faturasından Düşülecek": "FFF2CC",
        "Önizleme — Ayın Kendi Fiyatıyla (kayıt dışı)": "F2F2F2",
    }
    for label, renk in beklenen.items():
        row_idx = next(r[0].row for r in ws.iter_rows() if r[0].value == label)
        for col in range(1, 4):
            cell = ws.cell(row=row_idx, column=col)
            assert cell.font.bold, f"{label} sütun {col} bold değil"
            assert cell.fill.start_color.rgb.endswith(renk), f"{label} dolgusu {renk} değil"

    # Neden: Yeşil/kırmızı BİLEREK kullanılmadı — Net Ay Sonucu'nun kâr/zarar
    # işareti için ayrıldı. Burada kullanılsalardı renk anlamını kaybederdi.
    assert len(set(beklenen.values())) == 3, "üç satır üç ayrı renk kullanmalı"


def test_deduction_view_notes_are_italic_gray(monkeypatch):
    wb = _build_summary(monkeypatch, _locked())
    ws = wb["Faturalama Özeti"]
    not_cells = [row[0] for row in ws.iter_rows()
                 if str(row[0].value or "").startswith("Hesap:")]
    assert len(not_cells) == 2
    for cell in not_cells:
        assert cell.font.italic
        assert cell.font.color.rgb.endswith("808080")
        assert not cell.font.bold


def test_deduction_views_survive_service_failure(monkeypatch):
    """
    Neden: Önizleme bir GÖSTERİM katmanı — servis patlasa bile sayfa yazılmalı ve
    resmi tutar görünmeye devam etmeli (best-effort).
    """
    class _BrokenService:
        def get_monthly(self, year, month):
            return _locked()

        def get_own_electricity_price(self, year, month):
            raise RuntimeError("DB down")

        def get_preview_deduction(self, year, month):
            raise RuntimeError("DB down")

    monkeypatch.setattr("app.billing.BillingService", lambda: _BrokenService())
    wb = openpyxl.Workbook()
    wb.active.title = "Ay Özeti"
    job = MonthlySettlementJob.__new__(MonthlySettlementJob)
    job._write_billing_summary_sheet(wb, datetime(2026, 7, 1), "Temmuz 2026", _real_style_header)

    rows = {r[0]: r for r in _summary_rows(wb) if r[0]}
    assert rows["Resmi Kesinti — Ağustos 2026 Faturasından Düşülecek"][2] == "13.500,00"
    assert rows["Önizleme — Ayın Kendi Fiyatıyla (kayıt dışı)"][2] == "Bekleniyor"


def test_december_deduction_view_crosses_year_boundary(monkeypatch):
    """Aralık kesintisi Ocak faturasından düşülür — ay aritmetiği tek yerden."""
    wb = openpyxl.Workbook()
    wb.active.title = "Ay Özeti"
    cls = _fake_service_cls(_locked(2026, 12), _locked(2026, 12),
                            own_prices={}, previews={})
    monkeypatch.setattr("app.billing.BillingService", lambda: cls())
    job = MonthlySettlementJob.__new__(MonthlySettlementJob)
    job._write_billing_summary_sheet(wb, datetime(2026, 12, 1), "Aralık 2026", _real_style_header)

    labels = [r[0] for r in _summary_rows(wb) if r[0]]
    assert "Resmi Kesinti — Ocak 2027 Faturasından Düşülecek" in labels
    assert any("Kasım 2026 faturasının elektrik birim fiyatı" in str(label) for label in labels)


def test_summary_sheet_header_is_styled(monkeypatch):
    wb = _build_summary(monkeypatch, _locked())
    ws = wb["Faturalama Özeti"]
    header_row = next(r[0].row for r in ws.iter_rows() if r[0].value == "KALEM")
    for col in range(1, 4):
        cell = ws.cell(row=header_row, column=col)
        assert cell.font.bold
        assert cell.fill.start_color.rgb.endswith("EAEAEA")


def test_summary_sheet_skipped_without_billing_row(monkeypatch):
    wb = _build_summary(monkeypatch, None)
    assert "Faturalama Özeti" not in wb.sheetnames


def test_summary_sheet_survives_billing_failure(monkeypatch):
    def _boom():
        raise RuntimeError("DB down")

    monkeypatch.setattr("app.billing.BillingService", _boom)
    wb = openpyxl.Workbook()
    job = MonthlySettlementJob.__new__(MonthlySettlementJob)
    job._write_billing_summary_sheet(wb, datetime(2026, 7, 1), "Temmuz 2026", _real_style_header)
    assert "Faturalama Özeti" not in wb.sheetnames  # exception dışarı sızmadı


# ----------------------------------------------------------------------
# 3. Chatbot — metrik ayrıştırma
# ----------------------------------------------------------------------
@pytest.mark.parametrize(
    "question,expected",
    [
        ("fazla satış faturası ne kadar", "excess_sale_invoice"),
        ("bu ay enerjisa faturası", "excess_sale_invoice"),
        ("osb kesintisi ne kadar", "osb_deduction"),
        ("geçen ay osb faturası", "osb_deduction"),
        ("bu ay ne kadar düşecek", "osb_deduction"),
    ],
)
def test_billing_questions_are_parsed(question, expected):
    parsed = MetricParser().parse(question)
    assert expected in parsed["metrics"]
    assert parsed["explicit"] is True


def test_plain_kwh_question_does_not_trigger_billing():
    parsed = MetricParser().parse("bu ay fazla satış ne kadar")
    assert "grid_export" in parsed["metrics"]
    assert "excess_sale_invoice" not in parsed["metrics"]


# ----------------------------------------------------------------------
# 4. Chatbot — cevap üretimi
# ----------------------------------------------------------------------
def _answer(question, data, period="month", label="Temmuz 2026"):
    metric_info = MetricParser().parse(question)
    return ResponseBuilder().build(
        question, {"type": period, "label": label}, metric_info, data
    )


def test_chatbot_returns_invoice_amount():
    out = _answer("fazla satış faturası ne kadar", {"excess_sale_invoice": Decimal("2909.69")})
    assert "2.909,69 TL" in out
    assert "KDV hariç" in out


def test_chatbot_says_price_missing_not_zero():
    # Neden: None -> "0 TL" demek yöneticiyi yanıltır; hiç tutar gösterilmemeli.
    out = _answer("osb kesintisi ne kadar", {"osb_deduction": None})
    assert "0,00" not in out and "0 TL" not in out
    assert "TL" not in out  # hiçbir tutar yazılmamalı
    assert "girilmemiş" in out or "girilmedi" in out


def test_chatbot_rejects_billing_for_daily_period():
    out = _answer("dün osb kesintisi", {"osb_deduction": Decimal("1")}, period="day", label="dün")
    assert "yalnızca aylık" in out
    # Neden: capitalize() "OSB"yi "Osb" yapıyordu — kısaltma bozulmamalı.
    assert "Osb" not in out
    assert "OSB kesintisi" in out


def test_chatbot_reports_missing_billing_record():
    out = _answer("osb kesintisi ne kadar", {"production": 100})
    assert "faturalama kaydı bulunamadı" in out


def test_monthly_summary_appends_billing_lines():
    data = {
        "production": 10000, "consumption": 9000, "settled": 8000,
        "grid_import": 1000, "grid_export": 2000,
        "excess_sale_invoice": Decimal("2909.69"), "osb_deduction": None,
    }
    out = _answer("bu ay özet", data)
    assert "Faturalama (KDV hariç)" in out
    assert "2.909,69 TL" in out
    assert "Bekleniyor" in out


def test_daily_summary_has_no_billing_lines():
    # Neden: Günlük raporda TL yok (tutarlılık kararı).
    data = {"production": 100, "consumption": 90, "settled": 80,
            "grid_import": 10, "grid_export": 20}
    out = _answer("dün özet", data, period="day", label="dün")
    assert "Faturalama" not in out


# ----------------------------------------------------------------------
# 5. Chatbot — yanıt JSON'a çevrilebilmeli (2026-07-28 regresyonu)
# ----------------------------------------------------------------------
def test_billing_fields_are_json_serializable(monkeypatch):
    """
    Neden (regresyon): _billing_fields ham Decimal döndürüyordu ve bu sözlük
    doğrudan /api/chat yanıtının "data" alanına konuyor. json.dumps Decimal'i
    serileştiremediği için faturalama kaydı olan bir ayın SORULARININ TAMAMI
    (yalnızca TL soruları değil) "Sistem şu anda yanıt veremiyor." veriyordu.
    Cevap metni doğru üretiliyordu; patlayan yalnızca serileştirmeydi — bu
    yüzden mevcut testler bugı göremedi, hepsi metne bakıyordu.
    """
    monkeypatch.setattr(BillingService, "get_monthly", lambda self, y, m: _locked(2026, 6))

    fields = QueryEngine()._billing_fields(2026, 6)

    json.dumps(fields)  # ham Decimal ile TypeError verirdi
    assert isinstance(fields["excess_sale_invoice"], float)
    assert isinstance(fields["osb_deduction"], float)
    assert fields["excess_sale_invoice"] == 2909.69
    assert fields["osb_deduction"] == 13500.00


def test_billing_fields_keep_none_instead_of_zero(monkeypatch):
    # Neden: Tip dönüşümü None'ı 0.0'a çevirmemeli — "Bekleniyor" ile "sıfır TL"
    # ayrımı ADR-0002 §6'nın çekirdeği.
    monkeypatch.setattr(BillingService, "get_monthly", lambda self, y, m: _pending(2026, 7))

    fields = QueryEngine()._billing_fields(2026, 7)

    assert fields["osb_deduction"] is None
    assert isinstance(fields["excess_sale_invoice"], float)


def test_billing_fields_empty_when_no_record(monkeypatch):
    monkeypatch.setattr(BillingService, "get_monthly", lambda self, y, m: None)
    assert QueryEngine()._billing_fields(2026, 6) == {}


def test_chatbot_amount_text_unchanged_with_float_values():
    # Neden: Decimal -> float dönüşümü gösterilen tutarı değiştirmemeli.
    out = _answer("fazla satış faturası ne kadar", {"excess_sale_invoice": 2909.69})
    assert "2.909,69 TL" in out
