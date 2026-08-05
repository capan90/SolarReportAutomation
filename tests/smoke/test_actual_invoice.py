"""
Neden: Gerçek OSB fatura tutarı (dış girdi) ve ondan türetilen Net Ay Sonucu.

    Net = Gerçek Fatura[M] − (Enerjisa[M] + Kesinti[M]) + Kesinti[M−1]

Son terim geri EKLENİR: bu ayın faturasında zaten önceki ayın kesintisi
düşülmüştür (bir ayın kesintisi bir sonraki ayın faturasından düşülür).

Bu testler beş garantiyi sabitler:
1. Formül ve İŞARET yönü — pozitif ÖDENECEK, negatif/sıfır ALACAK.
2. Ay kaydırması geriye doğru da doğru (Ocak→Aralık yıl dönümü dahil).
3. Eksik girdi 0 ÜRETMEZ — net None ve "neyin beklendiği" adıyla söylenir.
4. Net KAYDEDİLMEZ; okuma yazma yapmaz, her çağrıda türetilir.
5. Excel'de renk/etiket işaret yönüyle tutarlı ve KDV uyarısı yazılı.
"""
from datetime import datetime
from decimal import Decimal

import openpyxl
import pytest

from app.billing.models import BillingValidationError, STATUS_LOCKED
from app.billing.service import BillingService
from app.jobs.monthly_settlement_job import MonthlySettlementJob


class FakeRepo:
    """monthly_billing + monthly_actual_invoice davranışını taklit eder."""

    def __init__(self, months=None, invoices=None):
        self.months = months or {}
        self.invoices = invoices or {}
        self.writes = []

    def get_monthly(self, year, month):
        row = self.months.get((year, month))
        return dict(row) if row else None

    def get_actual_invoice(self, year, month):
        row = self.invoices.get((year, month))
        return dict(row) if row else None

    def list_actual_invoices(self, limit=24):
        rows = sorted(self.invoices.values(),
                      key=lambda r: (r["year"], r["month"]), reverse=True)
        return [dict(r) for r in rows][:limit]

    def upsert_actual_invoice(self, year, month, amount_try, entered_by, note=None):
        previous = self.invoices.get((year, month), {}).get("amount_try")
        row = {"year": year, "month": month, "amount_try": amount_try,
               "entered_by": entered_by, "entered_at": None, "updated_at": None,
               "note": note}
        self.invoices[(year, month)] = row
        self.writes.append((year, month, amount_try))
        out = dict(row)
        out["_previous_amount_try"] = previous
        return out


def ay(year, month, enerjisa="11509229.10", kesinti="5311711.99"):
    return {
        "year": year, "month": month, "status": STATUS_LOCKED,
        "excess_sale_invoice_try": Decimal(enerjisa) if enerjisa else None,
        "osb_deduction_try": Decimal(kesinti) if kesinti else None,
    }


def fatura(year, month, tutar):
    return {"year": year, "month": month, "amount_try": Decimal(tutar),
            "entered_by": "murat", "entered_at": None, "updated_at": None, "note": None}


def svc(months=None, invoices=None):
    s = BillingService()
    s.repo = FakeRepo(months=months, invoices=invoices)
    return s


def tam_ay(actual="8000000.00"):
    """Tüm bağımlılıkları dolu bir Temmuz 2026 kurgusu."""
    return svc(
        months={
            (2026, 7): ay(2026, 7),
            (2026, 6): ay(2026, 6, kesinti="2879183.97"),
        },
        invoices={(2026, 7): fatura(2026, 7, actual)},
    )


# ---------------------------------------------------------------- ay kaydırması
@pytest.mark.parametrize("simdi,onceki", [
    ((2026, 7), (2026, 6)),
    ((2026, 12), (2026, 11)),
    ((2026, 1), (2025, 12)),   # yıl dönümü
])
def test_onceki_ay_hesabi(simdi, onceki):
    assert BillingService.previous_month(*simdi) == onceki


def test_next_ve_previous_birbirinin_tersi():
    # Neden: İki aritmetik ayrı yazıldı; sapma sessizce yanlış ay okuturdu.
    for year in (2025, 2026):
        for month in range(1, 13):
            assert BillingService.previous_month(*BillingService.next_month(year, month)) == (year, month)


# ---------------------------------------------------------------------- formül
def test_net_sonuc_formulu():
    # 8.000.000,00 − (11.509.229,10 + 5.311.711,99) + 2.879.183,97
    net = tam_ay().get_net_result(2026, 7)
    assert net["net_try"] == Decimal("-5941757.12")
    assert net["missing"] == []


def test_pozitif_net_odenecek_demektir():
    net = tam_ay(actual="20000000.00").get_net_result(2026, 7)
    assert net["net_try"] == Decimal("6058242.88")
    assert net["net_try"] > 0


def test_onceki_ay_kesintisi_geri_eklenir():
    """
    Neden: Bu ayın faturasında ZATEN önceki ayın kesintisi düşülmüştür; geri
    eklenmezse aynı tutar iki kez düşülmüş olur ve net olduğundan düşük çıkar.
    """
    az = tam_ay()
    az.repo.months[(2026, 6)]["osb_deduction_try"] = Decimal("0.00")
    cok = tam_ay()

    fark = cok.get_net_result(2026, 7)["net_try"] - az.get_net_result(2026, 7)["net_try"]
    assert fark == Decimal("2879183.97")


def test_net_bilesenleri_de_dondurulur():
    # Neden: Rapor ve ekran net rakamın nereden geldiğini kalem kalem gösteriyor.
    net = tam_ay().get_net_result(2026, 7)
    assert net["actual_invoice_try"] == Decimal("8000000.00")
    assert net["excess_sale_invoice_try"] == Decimal("11509229.10")
    assert net["osb_deduction_try"] == Decimal("5311711.99")
    assert net["previous_osb_deduction_try"] == Decimal("2879183.97")
    assert (net["previous_year"], net["previous_month"]) == (2026, 6)


def test_ocak_neti_gecen_yilin_aralik_kesintisini_kullanir():
    s = svc(
        months={(2026, 1): ay(2026, 1), (2025, 12): ay(2025, 12, kesinti="1000000.00")},
        invoices={(2026, 1): fatura(2026, 1, "20000000.00")},
    )
    net = s.get_net_result(2026, 1)
    assert (net["previous_year"], net["previous_month"]) == (2025, 12)
    assert net["previous_osb_deduction_try"] == Decimal("1000000.00")
    assert net["net_try"] == Decimal("4179058.91")


# ------------------------------------------------------------- eksik bağımlılık
@pytest.mark.parametrize("eksik,beklenen", [
    ("fatura", "gerçek OSB fatura tutarı"),
    ("enerjisa", "fazla satış faturası"),
    ("kesinti", "bu ayın OSB kesintisi"),
    ("onceki", "2026-06 OSB kesintisi"),
])
def test_eksik_girdi_sifir_uretmez_ve_adiyla_soylenir(eksik, beklenen):
    s = tam_ay()
    if eksik == "fatura":
        s.repo.invoices.clear()
    elif eksik == "enerjisa":
        s.repo.months[(2026, 7)]["excess_sale_invoice_try"] = None
    elif eksik == "kesinti":
        s.repo.months[(2026, 7)]["osb_deduction_try"] = None
    else:
        s.repo.months[(2026, 6)]["osb_deduction_try"] = None

    net = s.get_net_result(2026, 7)
    assert net["net_try"] is None, "eksik girdide sayı UYDURULMAMALI"
    assert beklenen in net["missing"]


def test_hicbir_kayit_yoksa_tum_eksikler_listelenir():
    net = svc().get_net_result(2026, 7)
    assert net["net_try"] is None
    assert len(net["missing"]) == 4


def test_sifir_fatura_bekleniyor_ile_karismaz():
    """
    Neden: "0 TL fatura geldi" ile "fatura henüz girilmedi" farklı şeyler. Sıfır
    tutar geçerli bir girdidir ve net hesabı yapılmalıdır.
    """
    s = tam_ay(actual="0.00")
    net = s.get_net_result(2026, 7)
    assert net["missing"] == []
    assert net["net_try"] == Decimal("-13941757.12")


# ------------------------------------------------------------------------ yazma
def test_net_hesabi_kayit_yazmaz():
    """Net KAYDEDİLMEZ — türetilmiş değeri saklamak ikinci doğruluk kaynağı olurdu."""
    s = tam_ay()
    once = {k: dict(v) for k, v in s.repo.invoices.items()}
    s.get_net_result(2026, 7)
    assert s.repo.writes == []
    assert {k: dict(v) for k, v in s.repo.invoices.items()} == once


def test_duzeltme_eski_degeri_dondurur():
    """Denetim izi eski -> yeni gösterebilmeli; LOCKED yok, aynı ay yeniden yazılır."""
    s = svc()
    ilk = s.set_actual_invoice(2026, 7, "8000000.00", entered_by="murat")
    assert ilk["_previous_amount_try"] is None

    duzeltme = s.set_actual_invoice(2026, 7, "8500000.00", entered_by="murat")
    assert duzeltme["_previous_amount_try"] == Decimal("8000000.00")
    assert duzeltme["amount_try"] == Decimal("8500000.00")


def test_negatif_tutar_reddedilir():
    with pytest.raises(BillingValidationError):
        svc().set_actual_invoice(2026, 7, "-100", entered_by="murat")


def test_sifir_tutar_kabul_edilir():
    kayit = svc().set_actual_invoice(2026, 7, "0", entered_by="murat")
    assert kayit["amount_try"] == Decimal("0")


@pytest.mark.parametrize("ay_no", [0, 13, -1])
def test_gecersiz_ay_reddedilir(ay_no):
    with pytest.raises(BillingValidationError):
        svc().set_actual_invoice(2026, ay_no, "100", entered_by="murat")


def test_entered_by_zorunlu():
    with pytest.raises(BillingValidationError):
        svc().set_actual_invoice(2026, 7, "100", entered_by="   ")


def test_yazim_sonrasi_net_turetilebilir():
    s = svc(months={(2026, 7): ay(2026, 7), (2026, 6): ay(2026, 6, kesinti="2879183.97")})
    assert s.get_net_result(2026, 7)["net_try"] is None
    s.set_actual_invoice(2026, 7, "8000000.00", entered_by="murat")
    assert s.get_net_result(2026, 7)["net_try"] == Decimal("-5941757.12")


# ------------------------------------------------------------------------ Excel
def _real_style_header(ws, row_idx, n_cols):
    from openpyxl.styles import Alignment, Font, PatternFill
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _build(monkeypatch, net_sonuc):
    """Faturalama Özeti sayfasını üretir; yalnızca Net bloğu ilgilendiriyor."""
    from app.billing import MonthlyBillingResult

    cur = MonthlyBillingResult(
        year=2026, month=7, status=STATUS_LOCKED,
        excess_sale_rate_try=Decimal("2.909687"), osb_unit_price_try=Decimal("1.452381"),
        production_kwh=Decimal("7612731.200"), excess_sale_kwh=Decimal("3955487.000"),
        excess_sale_invoice_try=Decimal("11509229.10"),
        osb_deduction_try=Decimal("5311711.99"),
    )

    class _FakeService:
        def get_monthly(self, year, month):
            return cur

        def get_own_electricity_price(self, year, month):
            return Decimal("1.680000")

        def get_preview_deduction(self, year, month):
            return Decimal("6144170.26")

        def get_net_result(self, year, month):
            return net_sonuc

    monkeypatch.setattr("app.billing.BillingService", lambda: _FakeService())
    wb = openpyxl.Workbook()
    wb.active.title = "Ay Özeti"
    job = MonthlySettlementJob.__new__(MonthlySettlementJob)
    job._write_billing_summary_sheet(wb, datetime(2026, 7, 1), "Temmuz 2026", _real_style_header)
    return wb["Faturalama Özeti"]


def _net(net_try, missing=None):
    return {
        "year": 2026, "month": 7, "previous_year": 2026, "previous_month": 6,
        "actual_invoice_try": Decimal("8000000.00"),
        "excess_sale_invoice_try": Decimal("11509229.10"),
        "osb_deduction_try": Decimal("5311711.99"),
        "previous_osb_deduction_try": Decimal("2879183.97"),
        "net_try": Decimal(net_try) if net_try is not None else None,
        "missing": missing or [],
    }


def _satirlar(ws):
    return {r[0]: r for r in ([c.value for c in row] for row in ws.iter_rows()) if r[0]}


def test_excel_net_blogu_kalemleri_tek_tek_yazar(monkeypatch):
    ws = _build(monkeypatch, _net("-5941757.12"))
    rows = _satirlar(ws)

    assert rows["Gerçek OSB Fatura Tutarı (elle girilen)"][2] == "8.000.000,00"
    assert rows["(−) Fazla Satış Faturası (Enerjisa)"][2] == "11.509.229,10"
    assert rows["(−) OSB Kesintisi — Resmi (bu ay)"][2] == "5.311.711,99"
    assert rows["(+) Haziran 2026 OSB Kesintisi"][2] == "2.879.183,97"
    assert rows["NET AY SONUCU (ÖDENECEK)"][2] == "-5.941.757,12"


def test_excel_kdv_ve_kalem_uyarisi_yazili(monkeypatch):
    """
    Neden: Fatura genelde KDV DAHİL gelir; sistemdeki her tutar KDV hariç.
    Karıştırılırsa net sonuç sessizce ~%20 yanlış çıkar.
    """
    ws = _build(monkeypatch, _net("-5941757.12"))
    uyari = next(r[0] for r in _satirlar(ws).values()
                 if str(r[0]).startswith("Bu tutar KDV HARİÇ"))
    assert "yalnızca ELEKTRİK kalemi" in uyari


@pytest.mark.parametrize("net_try,renk", [
    ("6058242.88", "FFC7CE"),    # pozitif -> ödenecek -> kırmızımsı
    ("-5941757.12", "C6EFCE"),   # negatif -> alacak -> yeşilimsi
    ("0.00", "C6EFCE"),          # sıfır da alacak tarafı (pozitif DEĞİL)
])
def test_excel_net_satiri_isaret_yonune_gore_renklenir(monkeypatch, net_try, renk):
    ws = _build(monkeypatch, _net(net_try))
    row_idx = next(r[0].row for r in ws.iter_rows() if r[0].value == "NET AY SONUCU (ÖDENECEK)")
    for col in range(1, 4):
        cell = ws.cell(row=row_idx, column=col)
        assert cell.font.bold
        assert cell.fill.start_color.rgb.endswith(renk)


def test_excel_net_aciklamasi_rengi_tek_basina_birakmaz(monkeypatch):
    """Neden: Renk körlüğü ve siyah-beyaz çıktı — yön yazıyla da söylenmeli."""
    ws_odenecek = _build(monkeypatch, _net("6058242.88"))
    assert any("ödenecek tutar kaldı" in str(r[0]) for r in _satirlar(ws_odenecek).values())

    ws_alacak = _build(monkeypatch, _net("-5941757.12"))
    assert any("alacak tarafında" in str(r[0]) for r in _satirlar(ws_alacak).values())


def test_excel_net_hesaplanamadiginda_bekleniyor_ve_sebep(monkeypatch):
    ws = _build(monkeypatch, _net(None, missing=["gerçek OSB fatura tutarı"]))
    rows = _satirlar(ws)

    assert rows["NET AY SONUCU (ÖDENECEK)"][2] == "Bekleniyor"
    row_idx = next(r[0].row for r in ws.iter_rows() if r[0].value == "NET AY SONUCU (ÖDENECEK)")
    # Nötr dolgu — hesaplanamayan bir sonuç kâr ya da zarar gibi renklenmemeli
    assert ws.cell(row=row_idx, column=1).fill.start_color.rgb.endswith("EAEAEA")
    assert any("bekleyen girdiler: gerçek OSB fatura tutarı" in str(r[0])
               for r in rows.values())


def test_excel_net_blogu_servis_patlarsa_sayfayi_dusurmez(monkeypatch):
    class _BrokenNet:
        def get_monthly(self, year, month):
            from app.billing import MonthlyBillingResult
            return MonthlyBillingResult(year=2026, month=7, status=STATUS_LOCKED,
                                        osb_deduction_try=Decimal("5311711.99"))

        def get_own_electricity_price(self, year, month):
            return None

        def get_preview_deduction(self, year, month):
            return None

        def get_net_result(self, year, month):
            raise RuntimeError("DB down")

    monkeypatch.setattr("app.billing.BillingService", lambda: _BrokenNet())
    wb = openpyxl.Workbook()
    wb.active.title = "Ay Özeti"
    job = MonthlySettlementJob.__new__(MonthlySettlementJob)
    job._write_billing_summary_sheet(wb, datetime(2026, 7, 1), "Temmuz 2026", _real_style_header)

    rows = _satirlar(wb["Faturalama Özeti"])
    assert "NET AY SONUCU (ÖDENECEK)" not in rows       # blok atlandı
    assert rows["Durum"][2] == "Kilitli"                # sayfa yazılmaya devam etti
