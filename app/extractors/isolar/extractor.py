import re
from pathlib import Path
from typing import Optional
from playwright.sync_api import Page
from app.core.config import settings
from app.core.logger import setup_logger
from app.core.utils import with_retry
from app.core.exceptions import (
    IsolarError,
    IsolarAuthenticationError,
    IsolarCredentialsError,
    IsolarServerSelectionError,
    IsolarTimeoutError,
    IsolarUnexpectedPageError,
    IsolarNavigationTimeoutError,
    IsolarReportPageNotFoundError,
    IsolarMenuStructureChangedError,
    IsolarDownloadError,
    IsolarDownloadTimeoutError,
    IsolarExportButtonNotFoundError
)


logger = setup_logger("IsolarExtractor")

class IsolarExtractor:
    """
    Neden: İsOlar Cloud portalına özgü sayfa etkileşimlerini, login akışını,
    oturum doğrulamayı ve hata yönetimini sarmalayıp yönetmek.
    """
    def __init__(self, page: Page, run_id: Optional[str] = None):
        self.page = page
        self.run_id = run_id

    @with_retry(
        max_retries=3,
        backoff_factor=2.0,
        retryable_exceptions=(IsolarTimeoutError, IsolarUnexpectedPageError, IsolarNavigationTimeoutError),
        non_retryable_exceptions=(IsolarCredentialsError, IsolarServerSelectionError)
    )
    def login_and_verify(self) -> None:
        """
        Neden: Sprint 1 hedefi olan İsOlar portalına başarılı şekilde giriş yapma
        ve oturum açıldığını doğrulama akışını orkestre etmek.
        """
        logger.info("İsOlar oturum açma akışı başlatılıyor...")
        
        self._open_login_page()
        self._handle_cookies()
        self._login()
        self._verify_login()
        
        logger.info("İsOlar oturumu başarıyla doğrulandı.")

    def _open_login_page(self) -> None:
        """
        Neden: İsOlar giriş sayfasına gitmek ve sayfanın ağ hareketliliği bitene kadar yüklenmesini beklemek.
        """
        base_url = settings.base_url
        logger.info(f"Giriş sayfasına gidiliyor: {base_url}")
        try:
            self.page.goto(base_url, wait_until="networkidle", timeout=20000)
        except Exception as e:
            raise IsolarTimeoutError(f"İsOlar portal sayfasına bağlanılamadı (Zaman aşımı): {e}")

    def _handle_cookies(self) -> None:
        """
        Neden: Çerez onay modalı çıktığında 'Evet katılıyorum' butonuna tıklayarak
        otomasyonu engelleyebilecek çerez perdesini kapatmak.
        """
        try:
            # Kullanıcı dostu get_by_role seçimi ile 'Evet' veya 'Accept' içeren birincil butonu bul
            cookie_btn = self.page.get_by_role("button", name=re.compile("Evet|Agree|Accept", re.I))
            if cookie_btn.is_visible():
                logger.info("Çerez onay modalı bulundu, onaylanıyor...")
                cookie_btn.click()
                # Modalın kapanması için kısa bir süre bekle
                self.page.wait_for_timeout(500)
        except Exception as e:
            logger.debug(f"Çerez onay modalı adımı atlandı veya hata oluştu (kritik değil): {e}")

    def _login(self) -> None:
        """
        Neden: Kimlik bilgilerini doldurmak, sözleşmeleri kabul etmek,
        güvenlik kontrollerini (CAPTCHA) denetlemek ve giriş işlemini tetiklemek.
        """
        logger.info("Giriş bilgileri dolduruluyor...")
        
        # 1. Kullanıcı Adı ve Şifre Alanlarını Doldur
        # Kullanıcı dostu get_by_placeholder kullanarak 'Hesap/Account' ve 'Şifre/Password' alanlarını hedefle
        account_input = self.page.get_by_placeholder(re.compile("Hesap|Account", re.I))
        password_input = self.page.get_by_placeholder(re.compile("Şifre|Password", re.I))
        
        # Elementlerin görünür ve aktif olmasını bekle
        account_input.wait_for(state="visible", timeout=5000)
        password_input.wait_for(state="visible", timeout=5000)
        
        account_input.fill(settings.username)
        password_input.fill(settings.password)

        # 2. Güvenlik ve CAPTCHA Kontrolü
        # iSolarCloud bazen şüpheli isteklerde veya çoklu girişlerde doğrulama kodu (Lütfen Girin) ister.
        captcha_input = self.page.get_by_placeholder(re.compile("Lütfen Girin|Verification Code", re.I))
        if captcha_input.is_visible():
            raise IsolarAuthenticationError(
                "İsOlar portalı CAPTCHA / Doğrulama Kodu talep ediyor. Otomasyon bu adımı manuel müdahale olmadan geçemez."
            )

        # 3. Gizlilik Politikası / Kullanıcı Sözleşmesi Kutusu Kontrolü
        # Eğer sayfada yasal onay onay kutusu varsa bulup işaretle
        checkboxes = self.page.locator("input[type='checkbox']").all()
        for cb in checkboxes:
            label_text = cb.evaluate("el => { const p = el.closest('.el-checkbox') || el.parentElement; return p ? p.innerText : ''; }").lower()
            if any(word in label_text for word in ["sözleşme", "gizlilik", "politika", "agree", "policy", "terms", "service"]):
                logger.info(f"Gizlilik/Sözleşme onay kutusu işaretleniyor: '{label_text}'")
                cb.check(force=True)

        # 4. Giriş Butonunu Tetikle
        # Birden fazla 'Giriş' butonu (mobil/masaüstü için gizli/görünür) arasından sadece görünür olanı bulup tıkla
        login_buttons = self.page.get_by_role("button", name=re.compile("Giriş|Login", re.I)).all()
        clicked = False
        for btn in login_buttons:
            if btn.is_visible():
                logger.info("Görünür Giriş butonuna tıklanıyor...")
                btn.click()
                clicked = True
                break
                
        if not clicked:
            raise IsolarUnexpectedPageError("Giriş yapabilmek için görünür bir 'Giriş/Login' butonu bulunamadı.")

    def _verify_login(self) -> None:
        """
        Neden: Giriş sonrasında başarılı bir yönlendirme yapıldığını veya
        sayfada oluşan giriş hatalarını tespit etmek.
        """
        logger.info("Giriş başarısı doğrulanıyor...")
        self._wait_for_authenticated_state()
        self._validate_authenticated_ui()

    def _wait_for_authenticated_state(self) -> None:
        """
        Neden: Giriş formundan çıkılıp URL'in yönlenmesini veya
        giriş bileşenlerinin kaybolmasını asenkron olarak beklemek.
        """
        logger.info("Oturumun açılması bekleniyor (sayfa durum değişimi)...")
        try:
            # 1. URL'in login içermeyen bir adrese geçmesini bekle
            self.page.wait_for_function(
                "() => !window.location.href.toLowerCase().includes('login')",
                timeout=12000
            )
            # 2. Şifre alanının gizlenmesini bekle
            password_input = self.page.get_by_placeholder(re.compile("Şifre|Password", re.I))
            password_input.wait_for(state="hidden", timeout=5000)
            
            logger.info("Giriş formundan başarıyla çıkıldı.")
        except Exception as e:
            logger.debug(f"Oturum açma geçiş süreci beklerken zaman aşımı veya uyarı: {e}")

    def _validate_authenticated_ui(self) -> None:
        """
        Neden: Giriş yapıldıktan sonra arayüzün kimliği doğrulanmış durumunu
        (authenticated layout) teyit etmek.
        """
        # 1. Login formu görünmüyor olmalı
        password_input = self.page.get_by_placeholder(re.compile("Şifre|Password", re.I))
        if password_input.is_visible():
            logger.warning("Doğrulama Hatası: Giriş formu hala görünür durumda.")
            self._check_login_failures()

        # 2. URL login sayfasından çıkmış olmalı
        if "login" in self.page.url.lower():
            logger.warning("Doğrulama Hatası: URL hala giriş sayfasını işaret ediyor.")
            self._check_login_failures()

        # 3. Authenticated layout görünmeli (Sidebar, Logout, User Menu, Plant List, Dashboard Container vb.)
        authenticated_indicators = [
            ".el-aside",             # Element Plus Sidebar
            ".sidebar",              # Genel Sidebar
            ".plant-list",           # Tesis Listesi Container
            ".plant-item",           # Tekil Tesis Elemanı
            "[href*='plantList']",   # Tesis Listesi Linki
            "[class*='plant']",      # Tesis içeren sınıf
            "[class*='user-avatar']",# Kullanıcı Profili/Avatarı
            ".user-avatar",          # Kullanıcı Avatarı
            "text=Oturumu Kapat",    # Oturumu kapat butonu
            "text=Çıkış",            # Türkçe Çıkış metni
            "text=Logout",           # İngilizce Logout metni
            ".main-container"        # Ana kapsayıcı dizin
        ]

        ui_valid = False
        for selector in authenticated_indicators:
            try:
                if self.page.locator(selector).first.is_visible():
                    logger.info(f"Oturum açma doğrulanmış arayüz bileşeni ile teyit edildi: {selector}")
                    ui_valid = True
                    break
            except Exception:
                pass

        # Alternatif olarak URL post-login şablonu içeriyorsa fakat eleman tam render olmadıysa tolerans sağla
        if not ui_valid:
            if any(p in self.page.url.lower() for p in ["plantlist", "main", "dashboard", "station"]):
                logger.info(f"Oturum açma post-login URL yapısı ile doğrulandı: {self.page.url}")
                ui_valid = True

        if not ui_valid:
            logger.error("Kimliği doğrulanmış arayüz bileşenleri veya yönlendirilmiş URL tespit edilemedi.")
            self._check_login_failures()

        logger.info(f"Oturum başarıyla açıldı ve arayüz doğrulandı. Mevcut URL: {self.page.url}")


    def _check_login_failures(self) -> None:
        """
        Neden: Giriş başarısız olduğunda sayfadaki hata mesajlarını veya diyalogları
        okuyarak anlamlı ve spesifik bir exception fırlatmak.
        """
        # Element Plus hata kutuları, alert ve mesaj sınıflarını sorgula
        error_locators = [
            ".el-message-box__message",
            ".el-message",
            ".error-message",
            ".el-notification__content",
            "[role='alert']"
        ]
        
        visible_errors = []
        for selector in error_locators:
            elements = self.page.locator(selector).all()
            for el in elements:
                if el.is_visible():
                    text = el.inner_text().strip()
                    if text:
                        visible_errors.append(text)
                        
        if visible_errors:
            error_msg = " | ".join(visible_errors)
            logger.error(f"Giriş hatası tespit edildi: {error_msg}")
            
            # Hata metnine göre spesifik exception fırlat
            if any(w in error_msg.lower() for w in ["şifre", "parola", "şifre yanlış", "password", "incorrect", "wrong"]):
                raise IsolarCredentialsError(f"Geçersiz şifre veya kullanıcı adı: {error_msg}")
            elif any(w in error_msg.lower() for w in ["sunucu", "mevcut", "server", "not exist"]):
                raise IsolarServerSelectionError(f"Sunucu seçimi yanlış veya kullanıcı mevcut değil: {error_msg}")
            else:
                raise IsolarAuthenticationError(f"Oturum açma hatası: {error_msg}")

        # Eğer görünür hata yoksa ama URL hala login sayfasındaysa timeout fırlat
        if "login" in self.page.url.lower():
            raise IsolarTimeoutError("Giriş yapıldı fakat ana sayfaya yönlendirme zaman aşımına uğradı.")
        else:
            raise IsolarUnexpectedPageError(f"Beklenmeyen bir sayfaya yönlendirme yapıldı. URL: {self.page.url}")

    @with_retry(
        max_retries=3,
        backoff_factor=2.0,
        retryable_exceptions=(IsolarError,)
    )
    def navigate_to_daily_report(self) -> None:
        """
        Neden: Giriş sonrasında global Report menüsü ve Yield Report alt menüsü üzerinden
        günlük üretim raporunun yer aldığı sayfaya güvenli navigasyonu gerçekleştirmek.
        """
        logger.info("Günlük üretim raporu sayfasına navigasyon başlatılıyor...")
        
        # 1. Global 'Report' menüsüne tıkla
        try:
            report_menu = self.page.locator("text=Report").first
            # Menu yapısının değişip değişmediğini kontrol et
            report_menu.wait_for(state="visible", timeout=10000)
            logger.info("Global Report menüsü bulundu, tıklanıyor...")
            report_menu.click()
        except Exception as e:
            raise IsolarMenuStructureChangedError(f"Global 'Report' menüsü bulunamadı veya tıklanamadı: {e}")
            
        # 2. Report Overview sayfasının yüklenmesini ve URL değişimini bekle
        try:
            self.page.wait_for_url(re.compile(r"reportTransformation/overview"), timeout=10000)
            self.page.wait_for_load_state("networkidle")
            logger.info("Report Overview sayfası yüklendi.")
        except Exception as e:
            raise IsolarNavigationTimeoutError(f"Report Overview sayfasına yönlendirme zaman aşımına uğradı: {e}")
            
        # 3. 'Yield report' alt menüsüne/kartına tıkla
        try:
            yield_report_btn = self.page.locator("text=Yield report").first
            yield_report_btn.wait_for(state="visible", timeout=8000)
            logger.info("Yield Report alt menüsü bulundu, tıklanıyor...")
            yield_report_btn.click()
        except Exception as e:
            raise IsolarMenuStructureChangedError(f"Yield Report alt menü kartı bulunamadı veya tıklanamadı: {e}")
            
        # 4. Yield Report sayfasının yüklenmesini ve URL değişimini bekle
        try:
            self.page.wait_for_url(re.compile(r"reportTransformation/yieldReport"), timeout=12000)
            self.page.wait_for_load_state("networkidle")
            logger.info("Yield Report sayfası yüklendi.")
        except Exception as e:
            raise IsolarNavigationTimeoutError(f"Yield Report sayfasına yönlendirme zaman aşımına uğradı: {e}")
            
        # 5. Rapor ekranı başarı kriterlerini doğrula (Table, Export Button, Day/Month/Year filters)
        self._verify_report_page_elements()

    def _verify_report_page_elements(self) -> None:
        """
        Neden: Rapor ekranının eksiksiz yüklendiğini, tablo ve dışa aktarım
        butonlarının görüntülendiğini teyit etmek.
        """
        logger.info("Rapor sayfası elemanları doğrulanıyor...")
        
        # Başarı Kriterleri: 
        # - Veri tablosu (el-table__body)
        # - Export butonu (Export metni veya .export-record-btn/.controls-btn)
        # - Date Filter (Day/Month/Year veya Gün/Ay/Yıl metinleri)
        try:
            table = self.page.locator(".el-table__body").first
            table.wait_for(state="visible", timeout=8000)
            logger.info("Veri tablosunun görünürlüğü doğrulandı.")
        except Exception as e:
            raise IsolarReportPageNotFoundError(f"Rapor veri tablosu yüklenemedi veya görünür değil: {e}")
            
        try:
            export_btn = self.page.get_by_role("button", name=re.compile("Export", re.I)).first
            export_btn.wait_for(state="visible", timeout=5000)
            logger.info("Dışa aktarma (Export) butonunun görünürlüğü doğrulandı.")
        except Exception as e:
            raise IsolarReportPageNotFoundError(f"Dışa aktarma (Export) butonu bulunamadı veya görünür değil: {e}")
            
        try:
            day_filter = self.page.locator("text=Day").first
            day_filter.wait_for(state="visible", timeout=5000)
            logger.info("Tarih filtresinin (Day) görünürlüğü doğrulandı.")
        except Exception as e:
            raise IsolarReportPageNotFoundError(f"Zaman kırılım/tarih filtresi bulunamadı veya görünür değil: {e}")
            
        logger.info("Tüm rapor sayfası başarı kriterleri başarıyla karşılandı.")

    @with_retry(
        max_retries=3,
        backoff_factor=2.0,
        retryable_exceptions=(IsolarError,)
    )
    def download_daily_report(self) -> Path:
        """
        Neden: Rapor sayfasında gün filtresinin seçildiğinden emin olmak,
        dışa aktarım işlemini (Export as Excel) asenkron olarak tetiklemek ve
        kuyruktan başarıyla tamamlanan rapor dosyasını indirip geçici konuma kaydetmek.
        """
        logger.info("Günlük üretim raporu indirme işlemi başlatılıyor...")

        # 1. Günlük görünümün (Day) seçildiğinden emin ol / Gerekirse tıkla
        try:
            day_filter = self.page.locator("text=Day").first
            day_filter.wait_for(state="visible", timeout=5000)
            logger.info("Day filtresi seçiliyor...")
            day_filter.click()
            self.page.wait_for_load_state("networkidle")
        except Exception as e:
            logger.warning(f"Day filtresi seçilemedi (devam ediliyor): {e}")

        # 2. Tüm tesisleri seç (İlk checkbox - Select All)
        try:
            inners = self.page.locator(".el-checkbox__inner").all()
            if inners:
                logger.info("Tüm tesisler tabloda işaretleniyor...")
                inners[0].click()
                self.page.wait_for_timeout(1000)
        except Exception as e:
            logger.warning(f"Tablo checkbox işaretleme hatası: {e}")

        # 3. Export butonunu bul ve tıkla (Dropdown aç)
        export_btn = self.page.locator("button.controls-btn").first
        if not export_btn.is_visible():
            raise IsolarExportButtonNotFoundError("Excel dışa aktarım (Export) butonu bulunamadı.")

        logger.info("Export butonu tıklanıyor...")
        export_btn.click()
        self.page.wait_for_timeout(1000)

        # 4. Dropdown menüden "Export as Excel" seç
        try:
            excel_option = self.page.locator("text=Export as Excel").first
            excel_option.wait_for(state="visible", timeout=5000)
            logger.info("Dropdown menüden 'Export as Excel' seçiliyor...")
            excel_option.click()
        except Exception as e:
            raise IsolarMenuStructureChangedError(f"Dropdown üzerinde 'Export as Excel' seçeneği bulunamadı: {e}")

        # 5. "Start export" bildiriminin görünmesini bekle
        try:
            toast = self.page.locator(".el-message").first
            toast.wait_for(state="visible", timeout=8000)
            logger.info(f"Bildirim yakalandı: {toast.inner_text().strip()}")
        except Exception as e:
            logger.warning(f"Bildirim ekranında 'Start export' görülmedi (yine de devam ediliyor): {e}")

        # 6. 'Export records' (Dışa aktarım geçmişi) sayfasına git
        try:
            export_records_btn = self.page.locator("button.export-record-btn").first
            export_records_btn.wait_for(state="visible", timeout=5000)
            logger.info("Export Records sayfasına yönlendiriliyor...")
            export_records_btn.click()
            self.page.wait_for_url(re.compile(r"reportTransformation/reportExport"), timeout=10000)
            self.page.wait_for_load_state("networkidle")
        except Exception as e:
            raise IsolarNavigationTimeoutError(f"Export Records sayfasına yönlendirme hatası: {e}")

        # 7. Asenkron işlem durumunu sorgula (durum 'Export successful' olana kadar bekle)
        success = False
        max_attempts = 15
        poll_interval = 4000
        
        logger.info("Rapor hazırlama durumu sorgulanıyor (Polling)...")
        for attempt in range(max_attempts):
            try:
                row = self.page.locator(".el-table__row").first
                row.wait_for(state="visible", timeout=5000)
                status_text = row.locator("td").nth(3).inner_text().strip()
                logger.info(f"Sorgu {attempt + 1}/{max_attempts}: Durum = '{status_text}'")
                
                if "Export successful" in status_text:
                    success = True
                    break
            except Exception as e:
                logger.warning(f"Sorgulama esnasında satır/durum okunamadı: {e}")
                
            logger.info(f"Rapor henüz hazır değil, {poll_interval/1000}s sonra tekrar denenecek...")
            self.page.wait_for_timeout(poll_interval)
            
            # Sayfayı yenileyerek güncel verileri çek
            try:
                self.page.reload(wait_until="networkidle")
                self.page.wait_for_timeout(2000)
            except Exception as e:
                logger.warning(f"Sayfa yenilenirken hata: {e}")

        if not success:
            raise IsolarDownloadTimeoutError("Belirtilen süre içerisinde rapor üretimi tamamlanamadı (Zaman Aşımı).")

        # 8. Başarıyla üretilmiş olan ilk satırdaki dosyayı indir
        logger.info("Rapor hazır. İndirme (Download) başlatılıyor...")
        try:
            row = self.page.locator(".el-table__row").first
            download_btn = row.locator(".icon-G2_Download_241").first
            
            with self.page.expect_download(timeout=30000) as download_info:
                download_btn.click()
            download = download_info.value
        except Exception as e:
            raise IsolarDownloadError(f"İndirme butonu tıklanırken veya indirme esnasında hata: {e}")

        if not download:
            raise IsolarDownloadError("Dosya indirme işlemi tetiklendi fakat dosya nesnesi alınamadı.")

        # 9. Dosyayı geçici dizine kaydet
        temp_dir = settings.download_directory / "temp"
        temp_dir.mkdir(parents=True, exist_ok=True)
        temp_path = temp_dir / download.suggested_filename

        logger.info(f"Dosya geçici konuma kaydediliyor: {temp_path}")
        try:
            download.save_as(str(temp_path))
        except Exception as e:
            raise IsolarDownloadError(f"İndirilen geçici dosya diske kaydedilemedi: {e}")

        logger.info(f"Dosya indirme başarıyla tamamlandı: {temp_path.name}")
        return temp_path

    def navigate_to_curve_page(self) -> None:
        """
        Neden: Maintenance menüsü ve Curve alt menüsü üzerinden
        eğri sorgulama (Curve) sayfasına navigasyonu gerçekleştirir.
        """
        logger.info("Curve sayfasına navigasyon başlatılıyor...")
        
        # 1. Sol menüden Maintenance'a tıkla
        try:
            maintenance_menu = self.page.locator("text=Maintenance").first
            if not maintenance_menu.is_visible(timeout=3000):
                maintenance_menu = self.page.locator("text=Mainten").first
            
            maintenance_menu.wait_for(state="visible", timeout=10000)
            logger.info("Maintenance menüsü bulundu, tıklanıyor...")
            maintenance_menu.click()
        except Exception as e:
            raise IsolarMenuStructureChangedError(f"Maintenance menüsü bulunamadı veya tıklanamadı: {e}")

        # 2. Curve alt menüsüne tıkla
        try:
            curve_submenu = self.page.locator("text=Curve").first
            curve_submenu.wait_for(state="visible", timeout=10000)
            logger.info("Curve alt menüsü bulundu, tıklanıyor...")
            curve_submenu.click()
        except Exception as e:
            raise IsolarMenuStructureChangedError(f"Curve alt menüsü bulunamadı veya tıklanamadı: {e}")

        # 3. Yönlendirmeyi ve 'Plant comparison' sekmesinin yüklenmesini bekle
        try:
            plant_comparison_tab = self.page.locator("text=Plant comparison").first
            plant_comparison_tab.wait_for(state="visible", timeout=20000)
            self.page.wait_for_load_state("networkidle")
            logger.info(f"Curve (Plant comparison) sayfası başarıyla yüklendi. Mevcut URL: {self.page.url}")
            self._take_screenshot("01_curve_page_loaded")
        except Exception as e:
            self._take_screenshot("error_navigate_curve")
            raise IsolarNavigationTimeoutError(f"Curve sayfasına navigasyon zaman aşımına uğradı (Plant comparison sekmesi yüklenemedi): {e}")

    def _take_screenshot(self, name: str) -> None:
        try:
            screenshot_dir = Path("outputs/test_isolar_curve/screenshots")
            screenshot_dir.mkdir(parents=True, exist_ok=True)
            self.page.screenshot(path=str(screenshot_dir / f"{name}.png"))
            logger.info(f"Ekran görüntüsü kaydedildi: {name}.png")
        except Exception as e:
            logger.warning(f"Ekran görüntüsü kaydedilemedi: {e}")

    def download_hourly_curve_report(self, date_str: Optional[str] = None, mode: str = "day") -> Path:
        """
        Neden: Curve sayfasındaki filtreleri ayarlayıp üretim raporunu Excel olarak indirir.

        mode="day"  : Günlük görünüm, saatlik veri. date_str formatı YYYY-MM-DD (yoksa dün).
        mode="month": Aylık görünüm. date_str formatı YYYY-MM (yoksa geçen ay);
                      tarih navigasyonu ay bazında ok butonuyla yapılır.
        """
        if mode not in ("day", "month"):
            raise ValueError(f"Geçersiz mode: '{mode}' ('day' veya 'month' bekleniyor).")

        # 1. Tarih çözümleme
        import datetime
        if not date_str:
            if mode == "month":
                first_of_month = datetime.date.today().replace(day=1)
                prev_month = first_of_month - datetime.timedelta(days=1)
                date_str = prev_month.strftime("%Y-%m")
            else:
                yesterday = datetime.date.today() - datetime.timedelta(days=1)
                date_str = yesterday.strftime("%Y-%m-%d")

        logger.info(f"Curve raporu indirme işlemi başlatılıyor (mode={mode}, Tarih: {date_str})...")

        # 2. Plant comparison sekmesinin aktif olduğundan emin ol
        try:
            plant_comparison_tab = self.page.locator("text=Plant comparison").first
            plant_comparison_tab.click()
            self.page.wait_for_timeout(500)
        except Exception as e:
            logger.warning(f"Plant comparison sekmesine tıklanırken hata (devam ediliyor): {e}")

        # 3. Select plant dropdown'da tüm santralleri seç (GES-2...8)
        try:
            # Plant select dropdown genelde placeholder'ında 'plant' veya 'select' içeren ilk select'tir.
            # Yoksa ilk .el-select'i tıklarız.
            plant_dropdown = self.page.locator(".el-select").first
            plant_dropdown.click()
            self.page.wait_for_timeout(1000)

            plant_ids = ["GES_2", "GES_3", "GES-4", "GES_5", "GES-6", "GES_7", "GES-8"]
            for pid in plant_ids:
                try:
                    # Seçeneği has-text ile bulurken timeout'u düşürelim ki tüm akışı 30 saniye bloke etmesin
                    option = self.page.locator(f".el-select-dropdown__item:has-text('{pid}')").first
                    is_selected = option.evaluate("el => el.classList.contains('selected')", timeout=2000)
                    if not is_selected:
                        option.click()
                        self.page.wait_for_timeout(200)
                        logger.info(f"  Santral seçildi: {pid}")
                    else:
                        logger.info(f"  Santral zaten seçili: {pid}")
                except Exception as ex:
                    logger.warning(f"  Santral seçilemedi veya bulunamadı ({pid}): {ex}")
            
            self.page.keyboard.press("Escape")
            self.page.wait_for_timeout(500)
            self._take_screenshot("02_plants_selected")
        except Exception as e:
            logger.warning(f"Santral dropdown seçimi yapılamadı (devam ediliyor): {e}")

        # 4. Measuring point: "Plant daily yield" seç
        try:
            # Placeholder'a göre veya index'ine göre bul
            measuring_dropdown = self.page.locator(".el-select").nth(1)
            measuring_dropdown.click()
            self.page.wait_for_timeout(500)
            option = self.page.locator(".el-select-dropdown__item:has-text('Plant daily yield')").first
            option.click()
            self.page.wait_for_timeout(500)
            logger.info("Measuring point 'Plant daily yield' seçildi.")
        except Exception as e:
            logger.warning(f"Measuring point (Plant daily yield) seçilemedi: {e}")

        # 5. Dönem butonu aktif olmalı (day -> Day, month -> Month)
        period_label = "Month" if mode == "month" else "Day"
        try:
            period_btn = self.page.locator(f"text={period_label}").first
            period_btn.click()
            self.page.wait_for_timeout(500)
            logger.info(f"{period_label} butonu tıklandı.")
        except Exception as e:
            logger.warning(f"{period_label} butonu tıklanamadı: {e}")

        # 6. 60 min dropdown seçili olmalı
        # Neden: Month görünümünde varsayılan '1 day' aralığı günlük satır üretir;
        # saatlik (744 satır) veri için burada da '60 min' seçilmelidir.
        try:
            interval_dropdown = self.page.locator(".el-select").nth(2)
            interval_dropdown.click()
            self.page.wait_for_timeout(500)
            option = self.page.locator(".el-select-dropdown__item").filter(has_text=re.compile("60")).first
            option.click()
            self.page.wait_for_timeout(500)
            logger.info("Interval '60 min' seçildi.")
        except Exception as e:
            logger.warning(f"Interval (60 min) seçilemedi: {e}")

        # 7. Hedef tarihe navigasyon — ok butonuyla gün gün geri git.
        # Neden: Tarih alanına string yazmak çalışmıyor; Element Plus bileşeni
        # beklediği DD/MMM/YYYY dışındaki formatı sessizce bugüne geri alıyor.
        # Ok butonu format ve locale bağımsızdır. Doğrulama başarısızsa hata
        # fırlatılır — yanlış günün verisiyle sessizce devam edilmez.
        from datetime import datetime as _dt

        _AYLAR = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
                  "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}

        def _read_current_date() -> "_dt":
            raw_val = self.page.locator(".el-date-editor input").first.input_value().strip()
            for fmt in ("%d/%b/%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"):
                try:
                    return _dt.strptime(raw_val, fmt)
                except ValueError:
                    continue
            # Neden: %b locale'e bağlı olabilir; ay kısaltmasını manuel çözümle.
            parts = raw_val.replace("-", "/").replace(".", "/").split("/")
            if len(parts) == 3 and parts[1][:3].lower() in _AYLAR:
                return _dt(int(parts[2]), _AYLAR[parts[1][:3].lower()], int(parts[0]))
            raise ValueError(f"Tarih alanı değeri çözümlenemedi: '{raw_val}'")

        def _click_prev_candidates() -> bool:
            """Önceki gün oku için aday selector'ları dener; tıklama yaptıysa True."""
            # Neden: iSolarCloud'da ok ikonu ".tool-date" içinde
            # "iconfont icon-a-G2_Leftarrow_20" span'ıdır (DOM keşfiyle doğrulandı).
            for sel in (".tool-date .icon-a-G2_Leftarrow_20",
                        ".icon-a-G2_Leftarrow_20",
                        "[class*='Leftarrow']",
                        "[class*='arrow-left']",
                        "[aria-label*='prev' i]"):
                try:
                    btn = self.page.locator(sel).first
                    if btn.is_visible(timeout=500):
                        btn.click()
                        return True
                except Exception:
                    continue
            return False

        def _read_current_month() -> tuple:
            """
            Neden: Month görünümünde tarih alanı 'May/2026', '2026-05', '05/2026' gibi
            farklı biçimlerde gelebilir; yıl regex'le, ay ise ad veya sayı olarak çözülür.
            (yıl, ay) döner.
            """
            raw_val = self.page.locator(".el-date-editor input").first.input_value().strip()
            m_year = re.search(r"(\d{4})", raw_val)
            if not m_year:
                raise ValueError(f"Ay alanında yıl bulunamadı: '{raw_val}'")
            year = int(m_year.group(1))
            rest = raw_val.replace(m_year.group(1), "")
            m_name = re.search(r"([A-Za-z]{3,})", rest)
            if m_name and m_name.group(1)[:3].lower() in _AYLAR:
                return year, _AYLAR[m_name.group(1)[:3].lower()]
            m_num = re.search(r"(\d{1,2})", rest)
            if m_num and 1 <= int(m_num.group(1)) <= 12:
                return year, int(m_num.group(1))
            raise ValueError(f"Ay alanı değeri çözümlenemedi: '{raw_val}'")

        if mode == "month":
            # Ay bazlı navigasyon: mevcut aydan hedef aya ok butonuyla ay ay geri git.
            target_dt = _dt.strptime(date_str, "%Y-%m")
            cur_year, cur_month = _read_current_month()
            diff_months = (cur_year - target_dt.year) * 12 + (cur_month - target_dt.month)
            logger.info(
                f"Ay navigasyonu: mevcut={cur_year}-{cur_month:02d}, "
                f"hedef={target_dt.year}-{target_dt.month:02d}, fark={diff_months} ay"
            )

            if diff_months < 0:
                raise ValueError(
                    f"Hedef ay ({target_dt.year}-{target_dt.month:02d}) sayfadaki aydan "
                    f"({cur_year}-{cur_month:02d}) ileride; ileri navigasyon desteklenmiyor."
                )

            for i in range(diff_months):
                before = _read_current_month()
                if not _click_prev_candidates():
                    raise ValueError("Önceki ay (<) butonu sayfada bulunamadı.")
                self.page.wait_for_timeout(800)
                after = _read_current_month()
                expected = (before[0], before[1] - 1) if before[1] > 1 else (before[0] - 1, 12)
                if after != expected:
                    raise ValueError(
                        f"Önceki ay tıklaması beklenen etkiyi yapmadı: {before} -> {after}"
                    )
                logger.info(f"  Önceki ay tıklandı ({i + 1}/{diff_months}): {after[0]}-{after[1]:02d}")

            final_year, final_month = _read_current_month()
            if (final_year, final_month) != (target_dt.year, target_dt.month):
                raise ValueError(
                    f"Ay doğrulaması başarısız: alan {final_year}-{final_month:02d} gösteriyor, "
                    f"hedef {target_dt.year}-{target_dt.month:02d}."
                )
            logger.info(f"Ay alanı doğrulandı: {final_year}-{final_month:02d}")
            # Neden: Aylık saatlik seri (744 nokta) günlük veriden belirgin şekilde
            # daha ağır yüklenir; export'un boş dosya vermemesi için veri beklenir.
            self.page.wait_for_timeout(3000)
        else:
            target_dt = _dt.strptime(date_str, "%Y-%m-%d")
            current_dt = _read_current_date()
            diff_days = (current_dt.date() - target_dt.date()).days
            logger.info(f"Tarih navigasyonu: mevcut={current_dt.date()}, hedef={target_dt.date()}, fark={diff_days} gün")

            if diff_days < 0:
                raise ValueError(
                    f"Hedef tarih ({target_dt.date()}) sayfadaki tarihten ({current_dt.date()}) ileride; "
                    f"ileri navigasyon desteklenmiyor."
                )

            for i in range(diff_days):
                before = _read_current_date()
                if not _click_prev_candidates():
                    raise ValueError("Önceki gün (<) butonu sayfada bulunamadı.")
                self.page.wait_for_timeout(800)
                after = _read_current_date()
                if (before.date() - after.date()).days != 1:
                    raise ValueError(
                        f"Önceki gün tıklaması beklenen etkiyi yapmadı: {before.date()} -> {after.date()}"
                    )
                logger.info(f"  Önceki gün tıklandı ({i + 1}/{diff_days}): {after.date()}")

            final_dt = _read_current_date()
            if final_dt.date() != target_dt.date():
                raise ValueError(
                    f"Tarih doğrulaması başarısız: alan {final_dt.date()} gösteriyor, hedef {target_dt.date()}."
                )
            logger.info(f"Tarih alanı doğrulandı: {final_dt.date()}")
        self._take_screenshot("03_filters_configured")



        # 7.5. Liste (Tablo) görünümüne geç
        try:
            list_view_btn = self.page.locator(".icon-G2_List_24").first
            list_view_btn.click()
            self.page.wait_for_timeout(1000)
            logger.info("List/Table görünümüne geçildi (.icon-G2_List_24 tıklandı).")
            self._take_screenshot("03c_list_view_active")
        except Exception as e:
            logger.warning(f"List görünümüne geçilemedi: {e}")



        # 8-9. Excel export + geçici konuma kaydetme (iki kez kullanılır: hedef dönem
        # ve referans için bir önceki dönem), bu yüzden yerel fonksiyona alınmıştır.
        temp_dir = settings.download_directory / "temp"
        temp_dir.mkdir(parents=True, exist_ok=True)

        def _export_excel(name_suffix: str = "") -> Path:
            # Excel indirme ikonu genelde 'export' veya 'Excel' veya .icon-G2_Export vb.
            download_btn = None
            priority_selectors = [
                ".icon-G2_Export_24",
                ".icon-G2_Export",
                "button:has-text('Export')",
                "button:has-text('Excel')",
                ".icon-G2_Download",
                ".icon-G2_Download_241",
                ".icon-G2_Download_24"
            ]
            for sel in priority_selectors:
                try:
                    btn = self.page.locator(sel).first
                    if btn.is_visible(timeout=1000):
                        download_btn = btn
                        logger.info(f"Kullanılacak export/download butonu eşleşen selector: {sel}")
                        break
                except Exception:
                    continue

            if not download_btn:
                raise IsolarDownloadError("Curve download/export butonu bulunamadı.")

            logger.info("Download başlatılıyor...")
            try:
                # Önce download butonuna basıp dropdown menüsünün açılmasını tetikleyelim
                download_btn.click()
                self.page.wait_for_timeout(800)

                # "Export as Excel" seçeneğini bul ve tıkla
                excel_option = self.page.locator("text=Export as Excel").first
                excel_option.wait_for(state="visible", timeout=5000)

                with self.page.expect_download(timeout=30000) as download_info:
                    excel_option.click()
                download = download_info.value
            except Exception as e:
                try:
                    page_text = self.page.evaluate("el => document.body.innerText")
                    logger.error(f"Hata anında sayfadaki metinler:\n{page_text}")
                except Exception as ex:
                    logger.warning(f"Sayfa metinleri dump edilemedi: {ex}")
                self._take_screenshot("error_download_click")
                raise IsolarDownloadError(f"Curve indirme tetiklenirken hata: {e}")

            if not download:
                raise IsolarDownloadError("Curve dosya indirme işlemi tetiklendi fakat dosya nesnesi alınamadı.")

            filename = download.suggested_filename
            if name_suffix:
                p = Path(filename)
                filename = f"{p.stem}{name_suffix}{p.suffix}"
            dest = temp_dir / filename
            try:
                download.save_as(str(dest))
            except Exception as e:
                raise IsolarDownloadError(f"İndirilen geçici Curve dosyası diske kaydedilemedi: {e}")
            return dest

        temp_path = _export_excel()
        self._take_screenshot("04_download_completed")
        logger.info(f"Curve dosya indirme başarıyla tamamlandı: {temp_path.name}")

        # 10. Delta düzeltmesi için önceki dönemin (gün/ay) SON kümülatif satırını
        # referans olarak dosyanın başına ekle.
        # Neden: Kümülatif seri gün başında sıfırlanmaz; 00:00 satırının deltası
        # ancak bir önceki günün 23:00 değeri bilinirse hesaplanabilir. Aksi halde
        # günün ilk saat üretimi kaybolur. Bu adım best-effort'tur: başarısız
        # olursa mevcut davranışa (ilk delta = 0) düşülür, akış bozulmaz.
        try:
            self._prepend_prev_period_reference(temp_path, mode, _click_prev_candidates, _export_excel)
        except Exception as e:
            logger.warning(
                f"Önceki dönem referans satırı eklenemedi (günün ilk saati 0 sayılabilir): {e}"
            )

        return temp_path

    def _prepend_prev_period_reference(self, target_path: Path, mode: str, click_prev, export_excel) -> None:
        """
        Neden: Sayfada bir dönem (gün/ay) daha geriye gidip export alınır ve o
        dosyanın SON veri satırı (önceki dönemin son saati) hedef dosyanın ilk
        veri satırı olarak eklenir. SettlementEngine bu satırı yalnızca ilk
        delta hesabı için referans alır, çıktıya dahil etmez.
        """
        import openpyxl

        if not click_prev():
            raise ValueError("Önceki dönem (<) butonu sayfada bulunamadı.")
        # Neden: Aylık saatlik seri belirgin şekilde ağır yüklenir.
        self.page.wait_for_timeout(3000 if mode == "month" else 1500)
        logger.info("Referans için önceki dönem görünümüne geçildi, export alınıyor...")

        prev_path = None
        try:
            # Neden: '_prev' son eki — hedef dosyayla aynı isim üretilirse çakışmasın.
            prev_path = export_excel("_prev")

            wb_prev = openpyxl.load_workbook(prev_path, data_only=True)
            try:
                rows = list(wb_prev.active.iter_rows(values_only=True))
            finally:
                wb_prev.close()

            # Dosya düzeni: satır1 etiket (Curve_...), satır2 başlık, satır3+ veri.
            if len(rows) < 3:
                raise ValueError(f"Önceki dönem dosyasında veri satırı yok ({prev_path.name}).")
            prev_header = rows[1]
            last_row = next(
                (r for r in reversed(rows[2:]) if r and r[0] not in (None, "")),
                None,
            )
            if last_row is None:
                raise ValueError(f"Önceki dönem dosyasında dolu veri satırı bulunamadı ({prev_path.name}).")

            wb_t = openpyxl.load_workbook(target_path)
            try:
                ws_t = wb_t.active
                target_header = tuple(next(ws_t.iter_rows(min_row=2, max_row=2, values_only=True)))
                # Neden: Kolon sırası/santral seti farklıysa yanlış hizalanmış referans
                # tüm santral deltalarını bozar; eşleşmiyorsa eklenmez.
                if tuple(prev_header) != target_header:
                    raise ValueError(
                        "Önceki dönem dosyasının başlıkları hedef dosyayla eşleşmiyor; referans eklenmedi."
                    )
                ws_t.insert_rows(3)
                for col_idx, value in enumerate(last_row, start=1):
                    ws_t.cell(row=3, column=col_idx, value=value)
                wb_t.save(str(target_path))
            finally:
                wb_t.close()

            logger.info(
                f"Önceki dönemin son satırı ({last_row[0]}) referans olarak eklendi: {target_path.name}"
            )
        finally:
            # Neden: Referans dosyası tek kullanımlıktır; geçici dizini kirletmesin.
            if prev_path is not None:
                try:
                    prev_path.unlink()
                except Exception:
                    pass



