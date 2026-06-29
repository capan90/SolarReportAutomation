import os
import sys
import json
import re
from pathlib import Path
from datetime import datetime, date
from typing import List, Dict, Any, Tuple
from dataclasses import asdict
import openpyxl

from app.core.config import settings
from app.core.logger import setup_logger
from app.core.exceptions import IsolarError
from app.profiling.profile_models import (
    ColumnProfile,
    DatasetSummary,
    SheetProfile,
    WorkbookProfile
)

logger = setup_logger("ExcelProfiler")

class ProfileJsonEncoder(json.JSONEncoder):
    """
    Neden: Dataclass içindeki datetime, date veya JSON uyumlu olmayan diğer nesneleri
    hata almadan düzgünce serileştirmek.
    """
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        try:
            return super().default(obj)
        except TypeError:
            return str(obj)

class ExcelProfiler:
    """
    Neden: Ham veri dosyalarının yapısal durumunu çıkarmak, veri tiplerini tahmin etmek,
    boş/tekil değer oranlarını belirlemek ve olası yapısal anormallikleri (suspicious findings)
    veri tabanına veya dönüştürme katmanına gitmeden önce tespit etmek.
    """
    def __init__(self):
        self.profiles_dir = settings.download_directory.parent / "profiles"

    def ensure_profiles_directory(self) -> None:
        """
        Neden: Profil raporlarının kaydedileceği dizinin varlığından emin olmak.
        """
        if not self.profiles_dir.exists():
            logger.info(f"Profil çıktı dizini oluşturuluyor: {self.profiles_dir}")
            self.profiles_dir.mkdir(parents=True, exist_ok=True)

    def _detect_header_row(self, rows: List[List[Any]]) -> Tuple[int, List[Any], List[List[Any]]]:
        """
        Neden: Satırlar arasından başlık satırını tespit etmek, üstteki satırları
        metadata_rows olarak ayırmak ve başlık satırı indeksini dönmek.
        Girdi: Satır değerleri listesi
        Çıktı: (header_row_index (0-based), header_columns, metadata_rows)
        """
        # Güçlü başlık göstergeleri
        header_indicators = {
            "plant name", "plantname", "installed power", "installedpower", 
            "grid connection", "gridconnection", "yield today", "yieldtoday", 
            "monthly yield", "monthlyyield", "annual yield", "annualyield", 
            "total yield", "totalyield", "equivalent hours", "equivalenthours", 
            "station name", "stationname"
        }
        
        for idx, row in enumerate(rows):
            # Satır hücrelerini temizle
            row_clean = [str(v).strip() if v is not None else "" for v in row]
            row_lower = [v.lower() for v in row_clean]
            
            # Sinyal 1: Güçlü başlık kelimelerinin bulunması
            if any(ind in row_lower or any(ind in cell for cell in row_lower) for ind in header_indicators):
                logger.info(f"Header row detected at index {idx} using indicators.")
                return idx, row, rows[:idx]
                
            # Sinyal 2: Satırda birden fazla (örn: >= 3) boş olmayan metin hücresi olması
            # ve bu hücrelerin hiçbirinin sayısal olmaması (başlık satırlarında genelde sayı bulunmaz)
            non_empty = [v for v in row_clean if v != ""]
            if len(non_empty) >= 3:
                all_text = True
                for v in non_empty:
                    try:
                        float(v)
                        all_text = False
                        break
                    except ValueError:
                        pass
                if all_text:
                    logger.info(f"Header row detected at index {idx} using text density/no numbers.")
                    return idx, row, rows[:idx]
                    
        # Varsayılan: Bulunamazsa 0. satır header kabul edilir
        logger.info("No header row detected. Defaulting to index 0.")
        default_header = rows[0] if rows else []
        return 0, default_header, []

    def _detect_sheet_role(self, sheet_name: str, total_rows: int, has_data: bool) -> str:
        """
        Neden: Sayfanın ismine ve içeriğine bakarak rolünü tahmin etmek.
        """
        name_lower = sheet_name.lower()
        if "chart" in name_lower or "grafik" in name_lower:
            return "chart"
        if "metadata" in name_lower or "info" in name_lower:
            return "metadata"
        if "yield" in name_lower or "report" in name_lower or "data" in name_lower:
            return "data"
        if total_rows > 2 and has_data:
            return "data"
        return "unknown"

    def profile_file(self, file_path: Path) -> WorkbookProfile:
        """
        Neden: Belirtilen Excel dosyasının (workbook) analizini yapmak ve profil raporunu çıkarmak.
        """
        logger.info(f"Excel dosya profili çıkarılıyor: {file_path}")
        
        if not file_path.exists():
            raise FileNotFoundError(f"Profil analizi yapılacak dosya bulunamadı: {file_path}")

        # 1. Dosya Üst Bilgileri (Metadata)
        stat = file_path.stat()
        file_size_bytes = stat.st_size
        created_at = datetime.fromtimestamp(stat.st_ctime).isoformat()
        modified_at = datetime.fromtimestamp(stat.st_mtime).isoformat()

        # 2. Openpyxl ile Workbook Yükleme (Veriler ve Formüller İçin Ayrı Ayrı)
        try:
            wb_data = openpyxl.load_workbook(str(file_path), data_only=True, read_only=False)
            wb_formulas = openpyxl.load_workbook(str(file_path), data_only=False, read_only=False)
        except Exception as e:
            raise IsolarError(f"Excel dosyası yüklenirken hata oluştu: {e}")

        sheet_names = wb_data.sheetnames
        total_sheets = len(sheet_names)

        sheets_profile: List[SheetProfile] = []
        suspicious_findings: List[str] = []

        # 3. Şüpheli Bulgular: Gizli Sayfalar (Hidden Sheets)
        for name in sheet_names:
            state = wb_data[name].sheet_state
            if state in ["hidden", "veryHidden"]:
                suspicious_findings.append(f"Hidden sheet detected: '{name}' (state={state})")

        # 4. Şüpheli Bulgular: Formül Hücreleri (Formula Cells)
        formulas_detected = False
        for name in sheet_names:
            ws_f = wb_formulas[name]
            for row in ws_f.iter_rows(values_only=True):
                for val in row:
                    if isinstance(val, str) and val.startswith("="):
                        formulas_detected = True
                        suspicious_findings.append(f"Formula cells detected in sheet '{name}': e.g., '{val}'")
                        break
                if formulas_detected:
                    break

        # 5. Workbook Boş mu Kontrolü (Empty Workbook)
        is_workbook_empty = True

        # 6. Her Sayfa (Sheet) İçin Profil Analizi
        for name in sheet_names:
            ws_d = wb_data[name]
            
            # Sayfa boyutlarını al
            used_range = ws_d.dimensions
            rows = list(ws_d.iter_rows(values_only=True))
            
            # Boş sayfa kontrolü
            if not rows or len(rows) == 0:
                logger.warning(f"Sayfa '{name}' tamamen boş.")
                continue

            is_workbook_empty = False

            # Şüpheli Bulgular: Birleştirilmiş Hücreler (Merged Cells)
            if len(ws_d.merged_cells.ranges) > 0:
                suspicious_findings.append(
                    f"Unexpected merged cells in sheet '{name}': Count={len(ws_d.merged_cells.ranges)}"
                )

            # Header Normalization Adımı
            header_idx, header_raw, metadata_rows = self._detect_header_row(rows)
            header_row_index = header_idx + 1
            data_start_row_index = header_idx + 2

            header: List[str] = []
            
            # Şüpheli Bulgular: Boş Başlık Hücresi (Empty Header)
            for i, col_name in enumerate(header_raw):
                c_name = str(col_name).strip() if col_name is not None else ""
                if c_name == "":
                    suspicious_findings.append(f"Empty header detected in sheet '{name}' at column index {i+1}")
                    header.append(f"Column_{i+1}")
                else:
                    header.append(c_name)

            # Şüpheli Bulgular: Çiftlenen Başlık Adı (Duplicate Column Name)
            seen_headers = set()
            for h in header:
                if h in seen_headers:
                    suspicious_findings.append(f"Duplicate column name detected in sheet '{name}': '{h}'")
                seen_headers.add(h)

            # Veri Satırları (Header sonrasındaki satırlar)
            data_rows = rows[header_idx + 1:] if len(rows) > (header_idx + 1) else []
            total_rows = len(rows)
            total_columns = len(header)

            sheet_role = self._detect_sheet_role(name, total_rows, len(data_rows) > 0)

            columns_profile: List[ColumnProfile] = []
            completely_empty_columns: List[str] = []
            completely_empty_rows: List[int] = []
            
            # Veri yoğunluğu ve hafıza boyutu hesabı
            estimated_size_bytes = 0
            
            # 7. Kolon Bazlı Profil (Column Profiling)
            for col_idx in range(total_columns):
                col_name = header[col_idx]
                
                # İlgili kolona ait tüm değerleri topla
                col_values = []
                for r in data_rows:
                    val = r[col_idx] if col_idx < len(r) else None
                    col_values.append(val)
                    estimated_size_bytes += sys.getsizeof(val)

                non_null_values = [v for v in col_values if v is not None and str(v).strip() != ""]
                non_null_count = len(non_null_values)
                null_count = len(col_values) - non_null_count
                null_ratio = null_count / len(col_values) if len(col_values) > 0 else 0.0
                
                # Benzersiz Değerler
                unique_values = list(set(non_null_values))
                unique_count = len(unique_values)
                
                # Örnek İlk 5 Değer
                sample_values = non_null_values[:5]

                # Veri Tipi Tahmini (Inferred Type)
                inferred_type = self._infer_column_type(non_null_values)
                
                # Şüpheli Bulgular: Karışık Veri Tipi (Mixed Type Column)
                if inferred_type == "mixed":
                    suspicious_findings.append(
                        f"Mixed type column detected in sheet '{name}' at column '{col_name}'"
                    )

                if non_null_count == 0:
                    completely_empty_columns.append(col_name)

                columns_profile.append(
                    ColumnProfile(
                        name=col_name,
                        index=col_idx + 1,
                        inferred_type=inferred_type,
                        null_ratio=null_ratio,
                        non_null_count=non_null_count,
                        null_count=null_count,
                        unique_count=unique_count,
                        sample_values=sample_values
                    )
                )

            # 8. Satır Bazlı Boşluk ve Çoğulluk Analizi
            # Tamamen boş satırları bul (Header ve metadata hariç)
            for row_idx, r in enumerate(data_rows):
                if all(v is None or str(v).strip() == "" for v in r):
                    completely_empty_rows.append(header_idx + row_idx + 2)

            # Duplicate Satır Sayısı
            row_tuples = []
            for r in data_rows:
                row_tuples.append(tuple(str(v) if v is not None else "" for v in r))
            
            unique_rows = set(row_tuples)
            duplicate_rows_count = len(row_tuples) - len(unique_rows) if len(row_tuples) > 0 else 0

            dataset_summary = DatasetSummary(
                total_rows=total_rows,
                total_columns=total_columns,
                estimated_size_bytes=estimated_size_bytes,
                completely_empty_columns=completely_empty_columns,
                completely_empty_rows=completely_empty_rows,
                duplicate_rows_count=duplicate_rows_count
            )

            sheets_profile.append(
                SheetProfile(
                    name=name,
                    total_rows=total_rows,
                    total_columns=total_columns,
                    header=header,
                    used_range=used_range,
                    sheet_role=sheet_role,
                    header_row_index=header_row_index,
                    metadata_rows=metadata_rows,
                    data_start_row_index=data_start_row_index,
                    columns=columns_profile,
                    dataset_summary=dataset_summary
                )
            )

        if is_workbook_empty:
            suspicious_findings.append("Empty workbook: No sheet contains data rows.")

        # 9. Workbook Raporunu Oluştur
        profile = WorkbookProfile(
            file_name=file_path.name,
            file_path=str(file_path.resolve()),
            file_size_bytes=file_size_bytes,
            created_at=created_at,
            modified_at=modified_at,
            total_sheets=total_sheets,
            sheet_names=sheet_names,
            sheets=sheets_profile,
            suspicious_findings=suspicious_findings
        )

        # 10. JSON Olarak Çıktı Dizinine Kaydet
        self.ensure_profiles_directory()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file_path = self.profiles_dir / f"profile_{timestamp}.json"
        
        try:
            profile_dict = asdict(profile)
            with open(output_file_path, "w", encoding="utf-8") as f:
                json.dump(profile_dict, f, cls=ProfileJsonEncoder, ensure_ascii=False, indent=2)
            logger.info(f"Profil raporu başarıyla kaydedildi: {output_file_path}")
        except Exception as e:
            logger.error(f"Profil JSON dosyası yazılırken hata oluştu: {e}")

        wb_data.close()
        wb_formulas.close()

        return profile

    def _infer_column_type(self, non_null_values: List[Any]) -> str:
        """
        Neden: Kolondaki boş olmayan hücrelerin değerlerine bakarak
        en uygun veri tipini tahmin etmek.
        """
        if not non_null_values:
            return "empty"

        types = set()
        for val in non_null_values:
            if isinstance(val, bool):
                types.add("boolean")
            elif isinstance(val, int):
                types.add("integer")
            elif isinstance(val, float):
                if val.is_integer():
                    types.add("integer")
                else:
                    types.add("float")
            elif isinstance(val, (datetime, date)):
                types.add("datetime")
            elif isinstance(val, str):
                val_str = val.strip()
                if val_str.lower() in ["true", "false"]:
                    types.add("boolean")
                    continue
                try:
                    int(val_str)
                    types.add("integer")
                    continue
                except ValueError:
                    pass
                try:
                    float(val_str)
                    types.add("float")
                    continue
                except ValueError:
                    pass
                if re.match(r"^\d{4}-\d{2}-\d{2}", val_str):
                    types.add("datetime")
                    continue
                types.add("text")
            else:
                types.add("text")

        if len(types) > 1:
            if types == {"integer", "float"}:
                return "float"
            return "mixed"
        return list(types)[0]
