import datetime
import time
from pathlib import Path
from typing import Optional, Dict, Any, List

import pandas as pd

from app.infrastructure.browser.playwright_client import PlaywrightClient
from app.extractors.isolar.extractor import IsolarExtractor
from app.sources.gaosb.extractor import (
    GaosbExtractor,
    GaosbCaptchaRequiredError,
    CAPTCHA_FLAG_PATH,
)
from app.settlement.engine import SettlementEngine
from app.settlement.models import HourlySettlement
from app.database.settlement_repository import RECON_METRICS
from app.notifications.notification_service import NotificationService
from app.core.logger import setup_logger
from app.core.config import settings

logger = setup_logger("MonthlySettlementJob")

# Neden: Task Scheduler cwd'si System32 olabilir — göreli yol işi sessizce öldürür
# (2026-07-22 DailySettlement olayı). Çıktı yolları proje köküne sabitlenir.
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

AY_ADLARI = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
             "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]


class MonthlySettlementJob:
    """
    Neden: DailySettlementJob'un aylık versiyonu — bir takvim ayının tamamı için
    iSolar Curve (mode="month", saatlik seri) ve GAOSB verilerini çekip saatlik
    mahsuplaşma hesaplar, aylık Excel raporu üretir ve e-posta ile bildirir.
    """

    @staticmethod
    def _load_gaosb_month(file_path: Path, target_month: str) -> pd.DataFrame:
        """
        Neden: SettlementEngine.load_gaosb() günlük akış için tasarlanmıştır ve
        veriyi dosyadaki İLK GÜNE filtreler; aylık akışta bu, 744 saatlik veriyi
        24 satıra indirir. Burada aynı okuma/normalizasyon yapılır ancak hedef
        AYA (YYYY-MM) filtrelenir. Kolonlara pozisyonla erişilir (0=Tarih, 5=Endeks değeri).
        """
        import numpy as np

        try:
            df = pd.read_excel(file_path, engine='openpyxl')
        except Exception:
            df = pd.read_excel(file_path, engine='xlrd')

        if df.empty:
            return pd.DataFrame(columns=['timestamp', 'consumption_kwh'])

        date_col = df.columns[0]
        val_col = df.columns[5]
        df = df.dropna(subset=[date_col, val_col])

        # Neden: Excel seri tarih formatını (varsa) standart datetime nesnesine dönüştürmek
        dates = []
        for val in df[date_col]:
            if isinstance(val, (int, float, np.integer, np.floating)) and not isinstance(val, bool):
                dates.append(datetime.datetime(1899, 12, 30) + datetime.timedelta(days=val))
            else:
                dates.append(pd.to_datetime(val))
        df['parsed_date'] = pd.to_datetime(dates)

        df['consumption_kwh'] = pd.to_numeric(df[val_col], errors='coerce').fillna(0.0)
        df['timestamp'] = df['parsed_date'].dt.strftime("%Y-%m-%d %H:00:00")

        result = df[['timestamp', 'consumption_kwh']].groupby('timestamp', as_index=False).sum()
        result = result[result['timestamp'].str.startswith(target_month)].reset_index(drop=True)
        return result

    @staticmethod
    def _calculate_monthly(df_prod: pd.DataFrame, df_cons: pd.DataFrame) -> List[HourlySettlement]:
        """
        Neden: SettlementEngine.calculate() ile aynı saatlik mahsup matematiği,
        ancak aylık (ay filtreli) üretim/tüketim DataFrame'leri üzerinden çalışır.
        """
        merged = pd.merge(df_prod, df_cons, on='timestamp').sort_values('timestamp')

        settlements: List[HourlySettlement] = []
        for _, row in merged.iterrows():
            prod = float(row['production_kwh'])
            cons = float(row['consumption_kwh'])
            settled = min(prod, cons)
            settlements.append(HourlySettlement(
                timestamp=str(row['timestamp']),
                production_kwh=prod,
                consumption_kwh=cons,
                settled_kwh=settled,
                grid_export_kwh=max(0.0, prod - cons),
                grid_import_kwh=max(0.0, cons - prod),
            ))
        return settlements

    # Neden: ADR-0003 Faz 1 tolerans eşiği. Fark, İKİ eşiği birden aşarsa anlamlı
    # sayılır — tek başına oransal eşik gece 0'a yakın değerlerde (grid_export = 0)
    # gürültüyü %100 fark gösterirdi, tek başına mutlak eşik ise ~200.000 kWh'lik
    # günlük değerlerde anlamsız kalırdı. Eşik yalnızca UYARIYI belirler; ham
    # db/scrape değerleri her koşuda saklandığı için eşik geriye dönük olarak
    # yeniden değerlendirilebilir (veri yeniden toplanmaz).
    RECON_TOLERANCE_ABS_KWH = 1.0
    RECON_TOLERANCE_REL = 0.001  # %0,1

    @staticmethod
    def _compare_month_with_db(
        settlements: List[HourlySettlement],
        db_snapshot: Dict[str, Dict[str, float]],
    ) -> List[Dict[str, Any]]:
        """
        Neden: ADR-0003 Faz 1 (Karşılaştır-ve-Uyar). Aylık iş DB'ye YAZMADAN ÖNCE,
        yeni hesapladığı değerlerle DB'de hâlihazırda duran (günlük işin yazdığı)
        değerleri karşılaştırır. Bu faz hiçbir davranışı değiştirmez — yalnızca
        "iki yol aynı sayıyı veriyor mu" sorusuna veri biriktirir.

        Saf fonksiyon: DB'ye dokunmaz, log yazmaz, istisna fırlatmaz. Her iki
        taraftaki günlerin BİRLEŞİMİ üzerinden her gün × her metrik için bir kayıt
        döner (eşleşenler dahil — payda kaybolmasın).

        within_tolerance=False iki durumdan biridir: metrik toleransı aştı VEYA
        kapsam uyuşmazlığı var (db_hours != scrape_hours). Kapsam uyuşmazlığı
        eşikten bağımsız olarak her zaman işaretlenir; metrik değeri tesadüfen
        tolerans içinde kalabilir ama eksik/kısmi gün en değerli sinyaldir.
        """
        by_day: Dict[str, List[HourlySettlement]] = {}
        for s in settlements:
            by_day.setdefault(str(s.timestamp)[:10], []).append(s)

        abs_tol = MonthlySettlementJob.RECON_TOLERANCE_ABS_KWH
        rel_tol = MonthlySettlementJob.RECON_TOLERANCE_REL

        results: List[Dict[str, Any]] = []
        for day in sorted(set(by_day) | set(db_snapshot)):
            day_settlements = by_day.get(day, [])
            db_day = db_snapshot.get(day, {})

            scrape_hours = len(day_settlements)
            db_hours = int(db_day.get("hours", 0))
            coverage_mismatch = db_hours != scrape_hours

            for metric in RECON_METRICS:
                scrape_value = sum(float(getattr(s, metric) or 0.0) for s in day_settlements)
                db_value = float(db_day.get(metric, 0.0))
                diff = scrape_value - db_value
                diff_pct = (diff / db_value * 100.0) if db_value else None

                significant = abs(diff) > abs_tol and abs(diff) > rel_tol * abs(db_value)
                results.append({
                    "date": day,
                    "metric": metric,
                    "db_value": db_value,
                    "scrape_value": scrape_value,
                    "diff": diff,
                    "diff_pct": diff_pct,
                    "within_tolerance": not (significant or coverage_mismatch),
                    "db_hours": db_hours,
                    "scrape_hours": scrape_hours,
                })
        return results

    @staticmethod
    def _log_reconciliation(target_month: str, comparisons: List[Dict[str, Any]]) -> None:
        """
        Neden: Karşılaştırma sonucu logdan da okunabilmeli (tablo sorgulamadan).
        Fark yoksa tek satır INFO; fark varsa gün/metrik bazında WARNING.
        """
        days = {c["date"] for c in comparisons}
        problems = [c for c in comparisons if not c["within_tolerance"]]
        if not problems:
            logger.info(
                "Karşılaştırma (ADR-0003 Faz 1): %s ayı için %d gün, tümü tolerans içinde.",
                target_month, len(days),
            )
            return

        problem_days = sorted({c["date"] for c in problems})
        logger.warning(
            "Karşılaştırma (ADR-0003 Faz 1): %s ayı için %d günün %d'inde fark var. "
            "Aylık iş DB'yi yine de üzerine yazacak (Faz 1 gözlem amaçlıdır).",
            target_month, len(days), len(problem_days),
        )
        for c in problems:
            if c["db_hours"] != c["scrape_hours"]:
                logger.warning(
                    "  %s / %s — KAPSAM: DB'de %d saat, yeni çekimde %d saat "
                    "(DB=%.2f, yeni=%.2f)",
                    c["date"], c["metric"], c["db_hours"], c["scrape_hours"],
                    c["db_value"], c["scrape_value"],
                )
            else:
                pct = "—" if c["diff_pct"] is None else f"{c['diff_pct']:+.3f}%"
                logger.warning(
                    "  %s / %s — DB=%.2f, yeni=%.2f, fark=%+.2f (%s)",
                    c["date"], c["metric"], c["db_value"], c["scrape_value"],
                    c["diff"], pct,
                )

    @staticmethod
    def _reconcile_best_effort(repo, run_id: str, target_month: str, year: int,
                               month: int, settlements: List[HourlySettlement]) -> None:
        """
        Neden: ADR-0003 Faz 1 karşılaştırmasını yürütür ve **hiçbir koşulda istisna
        fırlatmaz**. Bu garanti kritik: çağıran taraftaki upsert'ler bu çağrıdan
        SONRA geliyor ve gözlem amaçlı bir katmanın yazma akışını atlatması kabul
        edilemez. Hata yalnızca stack trace ile loglanır (sessiz hata yok).
        """
        try:
            db_snapshot = repo.get_hourly_month_snapshot(year, month)
            comparisons = MonthlySettlementJob._compare_month_with_db(settlements, db_snapshot)
            MonthlySettlementJob._log_reconciliation(target_month, comparisons)
            repo.save_reconciliation(run_id, target_month, comparisons)
        except Exception as recon_err:
            logger.error(
                "Karşılaştırma katmanı başarısız (ADR-0003 Faz 1, gözlem amaçlı; "
                "yazma ve rapor akışı etkilenmedi): %s", recon_err, exc_info=True,
            )

    @staticmethod
    def _five_metrics(settlements: List[HourlySettlement]) -> Dict[str, float]:
        """Neden: Rapor sayfalarında tekrarlanan beş metrik toplamını tek yerde hesaplamak."""
        return {
            "Üretim (kWh)": sum(s.production_kwh for s in settlements),
            "Tüketim (kWh)": sum(s.consumption_kwh for s in settlements),
            "Mahsup (kWh)": sum(s.settled_kwh for s in settlements),
            "Şebekeden Çekiş (kWh)": sum(s.grid_import_kwh for s in settlements),
            "Fazla Satış (kWh)": sum(s.grid_export_kwh for s in settlements),
        }

    # Neden: TL alanı yoksa 0 yazmak "sıfır fatura kesildi" anlamına gelirdi.
    # Eksik veri ile sıfır değer birbirinden ayrılmalı (ADR-0002 §6).
    BILLING_PENDING_TEXT = "Bekleniyor"

    @staticmethod
    def _fmt_try(value) -> str:
        """Neden: Türkçe para biçimi (binlik nokta, ondalık virgül); yoksa 'Bekleniyor'."""
        if value is None:
            return MonthlySettlementJob.BILLING_PENDING_TEXT
        return f"{float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    @staticmethod
    def _fmt_rate(value) -> str:
        """Neden: Birim fiyat 6 haneli gösterilir (2,909687); yoksa tire."""
        if value is None:
            return "—"
        return f"{float(value):,.6f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def _append_billing_section(self, ws1, month_dt, prev_month_dt, ay_str, prev_ay_str, style_header):
        """
        Neden: Ay Özeti sayfasına "FATURALAMA (TL, KDV HARİÇ)" bölümünü eklemek.
        Best-effort: billing kaydı okunamazsa bölüm atlanır, rapor üretimi sürer
        (Sprint A'daki job entegrasyonuyla aynı prensip).

        Değişim (%) yalnızca her iki ay da hesaplanmışsa yazılır; biri "Bekleniyor"
        ise yüzde uydurulmaz, tire konur.
        """
        try:
            from app.billing import BillingService

            service = BillingService()
            cur = service.get_monthly(month_dt.year, month_dt.month)
            prev = service.get_monthly(prev_month_dt.year, prev_month_dt.month)
        except Exception as e:
            logger.error(f"Faturalama bölümü Excel'e yazılamadı (rapora devam ediliyor): {e}")
            return

        if cur is None:
            logger.warning(
                "%04d-%02d için faturalama kaydı yok; Excel'e faturalama bölümü eklenmedi.",
                month_dt.year, month_dt.month,
            )
            return

        ws1.append([])
        ws1.append(["FATURALAMA (TL, KDV HARİÇ)", ay_str, f"Önceki Ay ({prev_ay_str})", "DEĞİŞİM (%)"])
        # Neden: Satır numarası başlık YAZILDIKTAN SONRA okunur. append([]) hücre
        # yazmadığı için max_row'u artırmaz (yalnızca _current_row artar); önceden
        # "max_row + 1" ile hesaplamak stili bir satır yukarıya, yani boş ara satıra
        # uyguluyordu ve gerçek başlık çıplak kalıyordu.
        style_header(ws1, ws1.max_row, 4)

        def _row(label: str, cur_val, prev_val):
            if cur_val is not None and prev_val is not None and float(prev_val) != 0:
                degisim = round((float(cur_val) - float(prev_val)) / float(prev_val) * 100, 1)
            else:
                degisim = None
            ws1.append([
                label,
                self._fmt_try(cur_val),
                self._fmt_try(prev_val) if prev is not None else "-",
                degisim if degisim is not None else "-",
            ])

        _row("Fazla Satış Faturası", cur.excess_sale_invoice_try,
             prev.excess_sale_invoice_try if prev else None)
        _row("OSB Kesintisi", cur.osb_deduction_try,
             prev.osb_deduction_try if prev else None)

        # Neden: Teyit raporunun asıl amacı — tutar hangi TL/kWh ile çıktı?
        ws1.append([
            "Kullanılan Katsayılar (Fazla Satış / OSB)",
            f"{self._fmt_rate(cur.excess_sale_rate_try)} / {self._fmt_rate(cur.osb_unit_price_try)}",
            (f"{self._fmt_rate(prev.excess_sale_rate_try)} / {self._fmt_rate(prev.osb_unit_price_try)}"
             if prev else "-"),
            "-",
        ])
        durum = "Kilitli" if cur.is_locked else "OSB birim fiyatı bekleniyor"
        ws1.append(["Durum", durum, ("Kilitli" if prev and prev.is_locked else "-"), "-"])
        logger.info("Excel faturalama bölümü yazıldı (%s, durum=%s).", ay_str, cur.status)

    def _write_billing_summary_sheet(self, wb, month_dt, ay_str, style_header):
        """
        Neden: OSB faturasıyla karşılaştırma yapan kişinin eline aldığı çıktı.
        Sheet 1'deki FATURALAMA bölümü "geçen aya göre değişim" görünümüdür;
        bu sayfa ise "bu ayın faturası" görünümü — kalem × (kWh, TL) şeklinde.
        2. sıraya konur ki kWh kırılımlarını geçip aranmasın.

        Tüm değerler monthly_billing'deki mevcut alanlardan TÜRETİLİR; yeni
        hesaplama mantığı yoktur. Hesaplanmamış tutar "Bekleniyor" yazar, 0 değil.

        KAPSAM DIŞI: Santral tipi (arazi/çatı) ayrımı yapılmaz — sistemde böyle
        bir alan yok, tek "Toplam Üretim" kalemi kullanılır.
        """
        try:
            from app.billing import BillingService

            cur = BillingService().get_monthly(month_dt.year, month_dt.month)
        except Exception as e:
            logger.error(f"Faturalama Özeti sayfası yazılamadı (rapora devam ediliyor): {e}")
            return

        if cur is None:
            logger.warning(
                "%04d-%02d için faturalama kaydı yok; Faturalama Özeti sayfası eklenmedi.",
                month_dt.year, month_dt.month,
            )
            return

        ws = wb.create_sheet("Faturalama Özeti", 1)
        ws.append([f"{ay_str} — FATURALAMA ÖZETİ (KDV HARİÇ)", "", ""])
        # Neden: Bu satır gerçek bir SAYFA BAŞLIĞI (yanındaki iki hücre boş), sütun
        # etiketi değil — birleştirilip ortalanabilir.
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=3)
        style_header(ws, ws.max_row, 3)
        ws.append(["KALEM", "kWh", "TL"])
        style_header(ws, ws.max_row, 3)

        prod = cur.production_kwh
        excess = cur.excess_sale_kwh
        # Neden: OSB'ye kalan, üretimin fazla satış dışında kalan kısmıdır.
        # Snapshot yoksa (tutarsız veri) uydurulmaz, tire konur.
        kalan_kwh = (prod - excess) if (prod is not None and excess is not None) else None

        def _kwh(v):
            return "-" if v is None else float(v)

        def _tl(v):
            return self.BILLING_PENDING_TEXT if v is None else self._fmt_try(v)

        toplam = None
        if cur.excess_sale_invoice_try is not None and cur.osb_deduction_try is not None:
            toplam = cur.excess_sale_invoice_try + cur.osb_deduction_try

        ws.append(["Toplam Üretim", _kwh(prod), "—"])
        ws.append(["Fazla Satış (Enerjisa'ya)", _kwh(excess), _tl(cur.excess_sale_invoice_try)])
        ws.append(["OSB'ye Kalan (Üretim − Fazla Satış)", _kwh(kalan_kwh), _tl(cur.osb_deduction_try)])
        ws.append(["TOPLAM (Enerjisa + OSB Kesintisi)", "—", _tl(toplam)])
        style_header(ws, ws.max_row, 3)

        ws.append([])
        # Neden: Raporu okuyan kişi "bu kesinti ne zaman tahsil edilecek" sorusunu
        # sormaya devam ediyordu; cevap raporun kendisinde olmalı. YENİ HESAPLAMA YOK —
        # tutar yukarıdaki osb_deduction_try'ın aynısı. Ay adı BillingService.next_month
        # ile bulunuyor; ay aritmetiği için ikinci bir yol açılmıyor (Aralık→Ocak yıl
        # dönümü orada testle sabitlenmiş).
        try:
            from app.billing import BillingService as _BS

            sonraki_yil, sonraki_ay = _BS.next_month(month_dt.year, month_dt.month)
            sonraki_str = f"{AY_ADLARI[sonraki_ay - 1]} {sonraki_yil}"
            ws.append([
                f"Bu Ayki OSB Kesintisi, {sonraki_str} Faturasından Düşülecektir",
                "—",
                _tl(cur.osb_deduction_try),
            ])
        except Exception as e:
            # Neden: Açıklayıcı bir satır raporu düşürmemeli (best-effort).
            logger.warning("OSB kesinti açıklama satırı yazılamadı: %s", e)

        ws.append([
            "Kullanılan Katsayılar (Fazla Satış / OSB)",
            "—",
            f"{self._fmt_rate(cur.excess_sale_rate_try)} / {self._fmt_rate(cur.osb_unit_price_try)}",
        ])
        ws.append([
            "Durum", "—",
            "Kilitli" if cur.is_locked else "OSB birim fiyatı bekleniyor",
        ])

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 24
        for row in ws.iter_rows(min_row=3, min_col=2, max_col=3):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '#,##0.0'
        logger.info("Faturalama Özeti sayfası yazıldı (%s, durum=%s).", ay_str, cur.status)

    def _write_monthly_report(
        self,
        settlements: List[HourlySettlement],
        output_path: Path,
        month_dt: datetime.datetime,
        prev_totals: Optional[Dict[str, float]],
    ) -> Path:
        """
        Neden: Aylık raporu 4 sayfalı üretmek:
        1) Ay Özeti (+ önceki ay karşılaştırması), 2) Haftalık Kırılım,
        3) Günlük Kırılım, 4) Saatlik Detay (tüm ay, 744 satır).
        Günlük rapordaki SettlementReportWriter formatı yerine aylık format kullanılır.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        font_header = Font(bold=True)
        fill_gray = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

        def _style_header(ws, row_idx: int, n_cols: int):
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.font = font_header
                cell.fill = fill_gray
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        metric_keys = ["Üretim (kWh)", "Tüketim (kWh)", "Mahsup (kWh)",
                       "Şebekeden Çekiş (kWh)", "Fazla Satış (kWh)"]
        prev_key_map = {
            "Üretim (kWh)": "production_kwh",
            "Tüketim (kWh)": "consumption_kwh",
            "Mahsup (kWh)": "settled_kwh",
            "Şebekeden Çekiş (kWh)": "grid_import_kwh",
            "Fazla Satış (kWh)": "grid_export_kwh",
        }

        ay_str = f"{AY_ADLARI[month_dt.month - 1]} {month_dt.year}"
        prev_month_dt = (month_dt.replace(day=1) - datetime.timedelta(days=1))
        prev_ay_str = f"{AY_ADLARI[prev_month_dt.month - 1]} {prev_month_dt.year}"

        wb = openpyxl.Workbook()

        # ---- Sheet 1: Ay Özeti ----
        ws1 = wb.active
        ws1.title = "Ay Özeti"
        totals = self._five_metrics(settlements)
        # Neden: Sayfa başlığı — Faturalama Özeti sayfasıyla simetrik görünüm için.
        # Mevcut 1. satır GERÇEK SÜTUN BAŞLIKLARI taşıyor (METRİK | ay | önceki ay |
        # değişim); onu birleştirmek etiketleri silerdi. Bu yüzden üstüne ayrı bir
        # başlık satırı eklenip A:D birleştiriliyor, tablo bir satır aşağı kayıyor.
        ws1.append([f"{ay_str} — AY ÖZETİ"])
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        _style_header(ws1, 1, 4)
        ws1.append(["METRİK", f"{ay_str}", f"Önceki Ay ({prev_ay_str})", "DEĞİŞİM (%)"])
        _style_header(ws1, ws1.max_row, 4)
        for key in metric_keys:
            cur = round(totals[key], 1)
            if prev_totals:
                prev = round(prev_totals[prev_key_map[key]], 1)
                degisim = round((cur - prev) / prev * 100, 1) if prev else None
            else:
                prev, degisim = None, None
            ws1.append([f"Toplam {key}", cur, prev if prev is not None else "-",
                        degisim if degisim is not None else "-"])

        # Neden: Kullanıcının Excel'de elle eklediği satır — üretimin fazla satış dışında
        # kalan kısmı, yani OSB'ye giden miktar. YENİ HESAPLAMA DEĞİL: yukarıdaki iki
        # satırın farkı. Faturalama Özeti sayfasındaki "OSB'ye Kalan" ile aynı sayıdır;
        # etikette bunu belirtmek iki sayfadaki aynı değerin farklı şeyler sanılmasını
        # önlüyor. Kaynak settlement toplamları — faturalama kaydı olmasa da çalışır.
        kalan_cur = round(totals["Üretim (kWh)"] - totals["Fazla Satış (kWh)"], 1)
        if prev_totals:
            kalan_prev = round(prev_totals["production_kwh"] - prev_totals["grid_export_kwh"], 1)
            kalan_degisim = (round((kalan_cur - kalan_prev) / kalan_prev * 100, 1)
                             if kalan_prev else None)
        else:
            kalan_prev, kalan_degisim = None, None
        # Neden: Etikette eşitlik AÇIKÇA yazılıyor — bu değer tanım gereği "Toplam
        # Mahsup" ile aynıdır (motor: fazla satış = üretim − mahsup). Aynı tabloda iki
        # özdeş sayı farklı adlarla dursaydı "neden aynı bunlar?" sorusu doğardı;
        # etiket hem OSB bağlantısını hem eşitliği söylüyor.
        ws1.append([
            "Üretim − Fazla Satış (= Mahsup, OSB'ye Kalan)",
            kalan_cur,
            kalan_prev if kalan_prev is not None else "-",
            kalan_degisim if kalan_degisim is not None else "-",
        ])

        # ---- Faturalama bölümü (ADR-0002) ----
        # Neden: Bu rapor OSB faturasının teyidi için kullanılıyor; tutarların yanında
        # HANGİ katsayıyla hesaplandığı da raporun kendisinde görünmeli. Değer yoksa
        # hücreye "Bekleniyor" yazılır — boş bırakılmaz (sessiz hata yok kuralı).
        self._append_billing_section(ws1, month_dt, prev_month_dt, ay_str, prev_ay_str, _style_header)

        ws1.column_dimensions["A"].width = 30
        for col in ("B", "C", "D"):
            ws1.column_dimensions[col].width = 22

        # Neden: Bu sayfadaki sayılar binlik ayraçsız görünüyordu (7612731.2), Faturalama
        # Özeti'nde ise biçimlendiriliyordu — aynı raporun iki sayfası farklı okunuyordu.
        # Yalnızca sayısal hücrelere uygulanır; "-", "Bekleniyor" gibi metinler ve
        # başlık satırları etkilenmez. Yüzde sütunu (D) da sayı olduğu için aynı biçimi
        # alır; tek ondalık gösterim yüzde için de doğru okuma.
        for row in ws1.iter_rows(min_row=2, min_col=2, max_col=4):
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = '#,##0.0'

        # ---- Sheet 2: Faturalama Özeti (ADR-0002) ----
        # Neden: 2. sıraya konur — OSB faturasıyla karşılaştırma yapan kişi
        # kWh kırılımlarını geçip aramasın.
        self._write_billing_summary_sheet(wb, month_dt, ay_str, _style_header)

        # Saatlik kayıtları güne göre grupla (Sheet 2 ve 3 için ortak)
        by_day: Dict[str, List[HourlySettlement]] = {}
        for s in settlements:
            by_day.setdefault(str(s.timestamp)[:10], []).append(s)
        days_sorted = sorted(by_day.keys())

        # ---- Sheet 2: Haftalık Kırılım (gün 1-7, 8-14, 15-21, 22-28, 29-son) ----
        ws2 = wb.create_sheet("Haftalık Kırılım")
        ws2.append(["HAFTA", "TARİH ARALIĞI"] + [k.upper() for k in metric_keys])
        _style_header(ws2, 1, 7)
        for week_idx in range(5):
            start_day = week_idx * 7 + 1
            week_days = [d for d in days_sorted if start_day <= int(d[8:10]) <= start_day + 6]
            if not week_days:
                continue
            week_settlements = [s for d in week_days for s in by_day[d]]
            wt = self._five_metrics(week_settlements)
            ws2.append(
                [f"Hafta {week_idx + 1}", f"{week_days[0]} - {week_days[-1]}"]
                + [round(wt[k], 1) for k in metric_keys]
            )
        for col in ("A", "B", "C", "D", "E", "F", "G"):
            ws2.column_dimensions[col].width = 24

        # ---- Sheet 3: Günlük Kırılım ----
        ws3 = wb.create_sheet("Günlük Kırılım")
        ws3.append(["TARİH"] + [k.upper() for k in metric_keys])
        _style_header(ws3, 1, 6)
        for d in days_sorted:
            dt_metrics = self._five_metrics(by_day[d])
            ws3.append([d] + [round(dt_metrics[k], 1) for k in metric_keys])
        for col in ("A", "B", "C", "D", "E", "F"):
            ws3.column_dimensions[col].width = 22

        # ---- Sheet 4: Saatlik Detay (tüm ay) ----
        ws4 = wb.create_sheet("Saatlik Detay")
        ws4.append(["TARİH", "SAAT ARALIĞI"] + [k.upper() for k in metric_keys])
        _style_header(ws4, 1, 7)
        for s in settlements:
            ts = str(s.timestamp)
            hour = int(ts[11:13])
            ws4.append([
                ts[:10],
                f"{hour:02d}:00-{(hour + 1) % 24:02d}:00",
                round(s.production_kwh, 1),
                round(s.consumption_kwh, 1),
                round(s.settled_kwh, 1),
                round(s.grid_import_kwh, 1),
                round(s.grid_export_kwh, 1),
            ])
        for col in ("A", "B", "C", "D", "E", "F", "G"):
            ws4.column_dimensions[col].width = 20

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        wb.close()
        logger.info(f"4 sayfalı aylık rapor kaydedildi: {output_path}")
        return output_path

    def run(
        self,
        target_month: Optional[str] = None,
        isolar_file: Optional[Path] = None,
        gaosb_file: Optional[Path] = None,
    ) -> Dict[str, Any]:
        """
        target_month: "YYYY-MM" formatında; None ise geçen ay.
        isolar_file / gaosb_file: verilirse indirme atlanır, mevcut dosya kullanılır
        (test ve yeniden koşum senaryoları için).
        """
        start_time = datetime.datetime.now()

        # 1. target_month hesapla (None -> geçen ay)
        if not target_month:
            first_of_month = datetime.date.today().replace(day=1)
            prev_month = first_of_month - datetime.timedelta(days=1)
            target_month = prev_month.strftime("%Y-%m")

        logger.info(f"Monthly Settlement Job BAŞLADI. Hedef Ay: {target_month}")

        try:
            month_dt = datetime.datetime.strptime(target_month, "%Y-%m")
        except ValueError as e:
            logger.error(f"Geçersiz ay formatı (YYYY-MM olmalı): {target_month}")
            return {
                "status": "FAILED",
                "month": target_month,
                "report_path": None,
                "settlement_count": 0,
                "error": f"Geçersiz ay formatı: {e}",
            }

        # 2. Tarih aralığı: ayın ilk günü -> sonraki ayın ilk günü (GAOSB açık aralık)
        date_from = month_dt.strftime("%Y-%m-01")
        if month_dt.month == 12:
            next_month = month_dt.replace(year=month_dt.year + 1, month=1)
        else:
            next_month = month_dt.replace(month=month_dt.month + 1)
        date_to = next_month.strftime("%Y-%m-01")

        output_dir = PROJECT_ROOT / "outputs" / "reports" / target_month
        output_dir.mkdir(parents=True, exist_ok=True)

        run_id = f"job-settlement-monthly-{target_month}-{int(time.time())}"
        headless = settings.headless

        isolar_path: Optional[Path] = Path(isolar_file) if isolar_file else None
        gaosb_path: Optional[Path] = Path(gaosb_file) if gaosb_file else None
        rapor_path: Optional[Path] = None
        settlements = []
        settlement_count = 0
        prev_totals: Optional[Dict[str, float]] = None
        error_msg: Optional[str] = None

        # 3. iSolar Curve indir — mode="month" (Best-effort)
        if isolar_path:
            logger.info(f"1. Aşama ATLANDI: mevcut iSolar dosyası kullanılıyor: {isolar_path}")
        else:
            try:
                logger.info(f"1. Aşama: iSolar Curve (month) indirme başlatılıyor (Ay: {target_month})...")
                with PlaywrightClient(headless=headless) as client:
                    page = client.create_page()
                    extractor = IsolarExtractor(page, run_id=run_id)
                    extractor.login_and_verify()
                    extractor.navigate_to_curve_page()
                    isolar_path = extractor.download_hourly_curve_report(
                        date_str=target_month, mode="month"
                    )
                logger.info(f"1. Aşama BAŞARILI. İndirilen dosya: {isolar_path}")
            except Exception as e:
                err_txt = f"iSolar Curve (month) indirme aşaması başarısız: {e}"
                logger.error(err_txt)
                error_msg = err_txt

        # 4. GAOSB indir (Best-effort)
        if gaosb_path:
            logger.info(f"2. Aşama ATLANDI: mevcut GAOSB dosyası kullanılıyor: {gaosb_path}")
        else:
            try:
                logger.info(f"2. Aşama: GAOSB raporu indirme başlatılıyor ({date_from} -> {date_to})...")
                extractor = GaosbExtractor()
                gaosb_path = extractor.download_report(
                    output_dir=output_dir,
                    date_from=date_from,
                    date_to=date_to,
                    headless=headless,
                )
                logger.info(f"2. Aşama BAŞARILI. İndirilen dosya: {gaosb_path}")
            except GaosbCaptchaRequiredError:
                # Neden: Captcha manuel doğrulama ister; job duraklatılır, yönetici
                # e-posta ile bilgilendirilir ve dashboard doğrulama akışı devreye girer.
                logger.warning("GAOSB captcha doğrulaması gerekiyor; aylık job duraklatılıyor.")
                try:
                    import json as _json
                    flag_info = {}
                    if CAPTCHA_FLAG_PATH.exists():
                        try:
                            flag_info = _json.loads(CAPTCHA_FLAG_PATH.read_text(encoding="utf-8-sig"))
                        except Exception:
                            flag_info = {}
                    flag_info.update({"job_type": "monthly", "target": target_month})
                    CAPTCHA_FLAG_PATH.write_text(_json.dumps(flag_info), encoding="utf-8")
                except Exception as flag_err:
                    logger.error(f"Captcha flag güncellenemedi (best-effort): {flag_err}")

                try:
                    notifier = NotificationService()
                    notifier.notify_pipeline(
                        run_id=run_id,
                        exit_code=2,
                        duration_ms=int((datetime.datetime.now() - start_time).total_seconds() * 1000),
                        stage_summary=(
                            f"{target_month} ayına ait aylık mahsuplaşma için GAOSB "
                            f"güvenlik doğrulaması gerekiyor.\n\n"
                            f"Lütfen dashboard'a girin ve "
                            f"'GAOSB Doğrulamasını Tamamla' butonuna tıklayın.\n"
                            f"Doğrulama sonrası rapor otomatik yeniden hazırlanacak."
                        ),
                        event_type="CAPTCHA_REQUIRED",
                        force=True,
                        # Neden: Arıza bildirimi — teknik ekibe gider (SMTP_TO_SYSTEM).
                        # "default" profili SMTP_TO'ya, yani tek kişiye düşüyordu.
                        email_profile="system"
                    )
                except Exception as mail_err:
                    logger.error(f"Captcha bildirimi gönderilemedi (best-effort): {mail_err}")

                return {
                    "status": "CAPTCHA_REQUIRED",
                    "month": target_month,
                    "report_path": None,
                    "settlement_count": 0,
                    "error": "GAOSB captcha doğrulaması gerekiyor",
                }
            except Exception as e:
                err_txt = f"GAOSB raporu indirme aşaması başarısız: {e}"
                logger.error(err_txt)
                error_msg = f"{error_msg} | {err_txt}" if error_msg else err_txt

        # 5. Settlement hesapla ve Excel rapor üret (Best-effort)
        try:
            if not isolar_path or not gaosb_path:
                raise ValueError(
                    "iSolar veya GAOSB dosya yollarından en az biri eksik olduğu için mahsuplaşma hesaplanamaz."
                )

            logger.info("3. Aşama: Aylık mahsuplaşma hesabı başlatılıyor...")
            engine = SettlementEngine()
            isolar_df = engine.load_isolar_curve(isolar_path)
            gaosb_df = self._load_gaosb_month(gaosb_path, target_month)
            settlements = self._calculate_monthly(isolar_df, gaosb_df)
            settlement_count = len(settlements)
            logger.info(f"Mahsuplaşma hesabı tamamlandı. Kayıt sayısı: {settlement_count}")

            # 3b. Önceki ay karşılaştırması için DB'den oku, sonra bu ayın
            # sonuçlarını yaz (Best-effort: DB hatası rapor üretimini engellemez).
            try:
                from app.database.settlement_repository import SettlementRepository
                repo = SettlementRepository()

                prev_month_dt = month_dt.replace(day=1) - datetime.timedelta(days=1)
                prev_totals = repo.get_monthly(prev_month_dt.year, prev_month_dt.month)

                # 3b-i. ADR-0003 Faz 1 — Karşılaştır-ve-Uyar (yazmadan ÖNCE, asla fırlatmaz).
                self._reconcile_best_effort(
                    repo, run_id, target_month, month_dt.year, month_dt.month, settlements
                )

                repo.upsert_hourly(settlements)  # tüm ay saatlik

                # Neden: Dashboard'daki günlük grafikler ve "Mahsuplaşmalarım"
                # settlement_daily'den beslenir; yalnızca hourly+monthly yazmak
                # aylık job ile doldurulan ayları dashboard'da boş bırakıyordu.
                # Saatlik kayıtlar güne gruplanıp her gün için daily upsert edilir.
                by_day: dict = {}
                for s in settlements:
                    day_key = str(s.timestamp)[:10]  # "YYYY-MM-DD HH:MM:SS" -> "YYYY-MM-DD"
                    by_day.setdefault(day_key, []).append(s)
                for day_key in sorted(by_day):
                    repo.upsert_daily(day_key, by_day[day_key])
                logger.info(f"settlement_daily yazıldı: {len(by_day)} gün")

                repo.upsert_monthly(month_dt.year, month_dt.month, settlements)
                logger.info("Mahsuplaşma sonuçları veritabanına yazıldı (hourly + daily + monthly).")
            except Exception as db_err:
                logger.error(f"Mahsuplaşma DB yazımı başarısız (rapor üretimine devam ediliyor): {db_err}")

            # 3c. Faturalama hesabı (ADR-0002) — BEST-EFFORT.
            # Neden: TL hesabı kWh raporunu asla engellemez. Katsayı tanımsızsa veya
            # OSB birim fiyatı henüz girilmemişse ay PENDING_RATE'te kalır; rapor
            # yine üretilir. Sprint A'da sonuç yalnızca DB'ye yazılır — Excel,
            # e-posta ve dashboard gösterimi Sprint B/C'de gelecek.
            try:
                from app.billing import BillingService

                billing = BillingService().compute(
                    year=month_dt.year,
                    month=month_dt.month,
                    production_kwh=sum(s.production_kwh for s in settlements),
                    excess_sale_kwh=sum(s.grid_export_kwh for s in settlements),
                )
                logger.info(
                    "Faturalama hesabı yazıldı: %04d-%02d durum=%s fatura=%s TL kesinti=%s TL",
                    month_dt.year, month_dt.month, billing.status,
                    billing.excess_sale_invoice_try, billing.osb_deduction_try,
                )
            except Exception as billing_err:
                logger.error(
                    f"Faturalama hesabı başarısız (rapor üretimine devam ediliyor): {billing_err}"
                )

            logger.info("4. Aşama: Excel raporu (4 sayfa) yazılıyor...")
            rapor_path = output_dir / f"mahsup_{month_dt.strftime('%Y%m')}_aylik.xlsx"
            self._write_monthly_report(settlements, rapor_path, month_dt, prev_totals)
            logger.info(f"4. Aşama BAŞARILI. Rapor üretildi: {rapor_path}")
        except Exception as e:
            err_txt = f"Mahsuplaşma veya rapor yazma aşaması başarısız: {e}"
            logger.error(err_txt)
            error_msg = f"{error_msg} | {err_txt}" if error_msg else err_txt

        # 6. E-Posta bildirimi gönder (Best-effort)
        try:
            logger.info("5. Aşama: E-Posta bildirimi gönderiliyor...")
            notifier = NotificationService()

            if rapor_path and rapor_path.exists():
                # Neden: E-posta yöneticiye gider; teknik detay yerine Türkçe ay adı
                # ve mahsup istatistikleri gösterilir (DailySettlementJob ile aynı kalıp).
                ay_str = f"{AY_ADLARI[month_dt.month - 1]} {month_dt.year}"

                def _fmt_kwh(value: float) -> str:
                    # Neden: Türkçe sayı biçimi (binlik ayracı nokta, ondalık virgül).
                    return f"{value:,.1f}".replace(",", "X").replace(".", ",").replace("X", ".")

                toplam_uretim = sum(s.production_kwh for s in settlements)
                toplam_tuketim = sum(s.consumption_kwh for s in settlements)
                toplam_mahsup = sum(s.settled_kwh for s in settlements)
                toplam_cekis = sum(s.grid_import_kwh for s in settlements)
                toplam_satis = sum(s.grid_export_kwh for s in settlements)

                stage_summary = (
                    f"{ay_str} ayına ait aylık mahsuplaşma raporu otomatik olarak hazırlanmıştır.\n\n"
                    f"Toplam Üretim: {_fmt_kwh(toplam_uretim)} kWh\n"
                    f"Toplam Tüketim: {_fmt_kwh(toplam_tuketim)} kWh\n"
                    f"Toplam Mahsup: {_fmt_kwh(toplam_mahsup)} kWh\n"
                    f"Şebekeden Çekiş: {_fmt_kwh(toplam_cekis)} kWh\n"
                    f"Fazla Satış: {_fmt_kwh(toplam_satis)} kWh"
                )

                # Neden: Faturalama tutarları yöneticinin ilk baktığı yer olan e-posta
                # özetinde de görünmeli. Best-effort: okunamazsa kWh özeti değişmez.
                try:
                    from app.billing import BillingService

                    billing = BillingService().get_monthly(month_dt.year, month_dt.month)
                    if billing:
                        # Neden: "Bekleniyor TL" saçma okunur; birim yalnızca gerçek
                        # tutara eklenir.
                        def _tl(value) -> str:
                            if value is None:
                                return self.BILLING_PENDING_TEXT
                            return f"{self._fmt_try(value)} TL"

                        stage_summary += (
                            f"\n\nFaturalama (KDV hariç):\n"
                            f"Fazla Satış Faturası: {_tl(billing.excess_sale_invoice_try)}\n"
                            f"OSB Kesintisi: {_tl(billing.osb_deduction_try)}"
                        )
                        if not billing.is_locked:
                            # Neden: "Bekleniyor" yazıp bırakmak yönetici için eyleme
                            # dönük değil; ne yapması gerektiği söylenir.
                            stage_summary += (
                                "\n\nOSB birim fiyatı henüz girilmedi. Dashboard'dan girip "
                                "raporu yeniden üretebilirsiniz."
                            )
                except Exception as billing_err:
                    logger.error(
                        f"E-posta özetine faturalama eklenemedi (bildirime devam ediliyor): {billing_err}"
                    )

                # Neden: Önceki ay DB'de kayıtlıysa yönetici özetine karşılaştırma eklenir.
                if prev_totals:
                    prev_month_dt = month_dt.replace(day=1) - datetime.timedelta(days=1)
                    prev_ay_str = f"{AY_ADLARI[prev_month_dt.month - 1]} {prev_month_dt.year}"
                    prev_uretim = prev_totals["production_kwh"]
                    degisim = ((toplam_uretim - prev_uretim) / prev_uretim * 100) if prev_uretim else 0.0
                    yon = "artış" if degisim >= 0 else "azalış"
                    stage_summary += (
                        f"\n\nÖnceki ay ({prev_ay_str}) üretimi {_fmt_kwh(prev_uretim)} kWh idi; "
                        f"%{abs(degisim):.1f} {yon} gerçekleşti."
                    )
                notifier.notify_pipeline(
                    run_id=run_id,
                    exit_code=0,
                    duration_ms=int((datetime.datetime.now() - start_time).total_seconds() * 1000),
                    stage_summary=stage_summary,
                    event_type="SUCCESS",
                    attachment_path=str(rapor_path.absolute()),
                    force=True,
                    email_profile="monthly"
                )
            else:
                stage_summary = (
                    f"Monthly Settlement Job BAŞARISIZ oldu.\n"
                    f"Hedef Ay: {target_month}\n"
                    f"iSolar Raporu: {'İndirildi' if isolar_path else 'BAŞARISIZ'}\n"
                    f"GAOSB Raporu: {'İndirildi' if gaosb_path else 'BAŞARISIZ'}\n"
                    f"Hata Detayları: {error_msg}"
                )
                notifier.notify_pipeline(
                    run_id=run_id,
                    exit_code=1,
                    duration_ms=int((datetime.datetime.now() - start_time).total_seconds() * 1000),
                    stage_summary=stage_summary,
                    # Neden: Arıza bildirimi rapor alıcılarına DEĞİL teknik ekibe gider
                    # (SMTP_TO_SYSTEM) — bkz. daily_settlement_job.py'deki aynı karar.
                    email_profile="system",
                )
            logger.info("5. Aşama BAŞARILI. Bildirim tamamlandı.")
        except Exception as e:
            logger.error(f"E-Posta bildirimi gönderilirken hata oluştu: {e}")

        # 7. Sonuç döndür
        is_success = (rapor_path is not None and rapor_path.exists())
        logger.info(f"Monthly Settlement Job TAMAMLANDI. Durum: {'SUCCESS' if is_success else 'FAILED'}")

        return {
            "status": "SUCCESS" if is_success else "FAILED",
            "month": target_month,
            "report_path": str(rapor_path) if rapor_path else None,
            "settlement_count": settlement_count,
            "error": error_msg if not is_success else None,
        }
