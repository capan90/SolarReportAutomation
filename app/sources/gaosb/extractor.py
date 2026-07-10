import os
import sys
import time
from pathlib import Path
from typing import Optional, List, Dict, Any

from app.sources.interface import ISourceExtractor
from app.sources.models import SourceMetadata
from app.sources.exceptions import SourceAuthenticationError
from app.sources.context import get_source_context
from app.infrastructure.browser.playwright_client import PlaywrightClient
from app.core.logger import setup_logger

logger = setup_logger("GaosbExtractor")


class GaosbCaptchaRequiredError(Exception):
    """
    Neden: BotGuard captcha manuel doğrulama gerektirdiğinde, interaktif olmayan
    çalıştırmalarda (dashboard/scheduler) süreci bloklamak yerine bu istisna
    fırlatılır. Dashboard, flag dosyası üzerinden kullanıcıya doğrulama akışı sunar.
    """
    pass


PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent

# Neden: Captcha bekleyen koşuların işareti; dashboard bu dosyayı okuyup kullanıcıya
# sarı uyarı bandı gösterir. Proje köküne sabitlenir (scheduler'ın cwd'sinden bağımsız).
CAPTCHA_FLAG_PATH = PROJECT_ROOT / "config" / "gaosb_captcha_flag.txt"

# Neden: BotGuard clearance + oturum çerezlerinin çalıştırmalar arasında kalıcı olması için
# gerçek bir Chromium profili (persistent context) kullanılır. Yol proje köküne sabitlenir
# ki scheduler hangi dizinden çalışırsa çalışsın aynı profil bulunsun.
USER_DATA_DIR = PROJECT_ROOT / "config" / "gaosb_browser_profile"

# Neden: Otomasyon parmak izini azaltmak için gerçekçi bir masaüstü Chrome User-Agent'ı.
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)


class GaosbExtractor(ISourceExtractor):
    """
    Neden: Gaziantep Organize Sanayi Bölgesi (GAOSB) sayaç sorgulama portalından
    Playwright aracılığıyla kimlik doğrulayıp sayaç sorgu raporunu indiren adaptör sınıfı.

    Portal BotGuard (botguard.net) anti-bot katmanı arkasında hCaptcha ile korunuyor.
    Bu yüzden kalıcı (persistent) bir Chromium profili kullanılır: captcha bir kez manuel
    çözülür, BotGuard clearance çerezi profile kaydedilir ve sonraki headless çalıştırmalar
    captcha görmeden geçer.
    """
    source_id = "gaosb"
    display_name = "GAOSB Sayaç Sorgu"
    capabilities = ["HOURLY_CONSUMPTION", "DAILY_METER_INDEX"]

    @property
    def metadata(self) -> SourceMetadata:
        """
        Neden: ISourceExtractor arayüzü gereksinimlerini karşılamak amacıyla
        kaynak hakkında değişmez meta verileri döner.
        """
        return SourceMetadata(
            source_name="gaosb",
            vendor="Gaziantep OSB",
            version="1.0.0",
            supported_report_types=["hourly_consumption", "daily_meter_index"],
            authentication_type="credentials",
            capabilities={
                "supports_excel_export": True,
                "supports_api": False,
                "supports_scheduler": True,
                "supports_health_check": True,
                "HOURLY_CONSUMPTION": True,
                "DAILY_METER_INDEX": True
            }
        )

    # ------------------------------------------------------------------
    # Persistent context yardımcıları
    # ------------------------------------------------------------------
    def _launch_persistent(self, headless: bool):
        """
        Neden: Kalıcı Chromium profiliyle bir persistent context başlatır. Böylece BotGuard
        clearance çerezi ve oturum bilgileri diske yazılır ve sonraki çalıştırmalarda yeniden
        kullanılır. (pw, context) döner; çağıran taraf kapatmaktan sorumludur.
        """
        from playwright.sync_api import sync_playwright

        USER_DATA_DIR.mkdir(parents=True, exist_ok=True)
        pw = sync_playwright().start()
        context = pw.chromium.launch_persistent_context(
            user_data_dir=str(USER_DATA_DIR),
            headless=headless,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-setuid-sandbox",
            ],
            user_agent=USER_AGENT,
            viewport={"width": 1280, "height": 800},
            accept_downloads=True,
        )
        logger.info("Persistent context başlatıldı (headless=%s, profil=%s)", headless, USER_DATA_DIR)
        return pw, context

    def _is_captcha_page(self, page) -> bool:
        """
        Neden: BotGuard/hCaptcha güvenlik doğrulama sayfasını içerik imzalarından tespit eder.
        Normal DevExpress login sayfası bu imzaları içermez, bu yüzden yanlış pozitif olmaz.
        """
        try:
            content = (page.content() or "").lower()
        except Exception:
            return False
        markers = [
            "hcaptcha",
            "botguard",
            "security check",
            "ben bir insanım",
            "we're so sorry",
        ]
        return any(m in content for m in markers)

    def _stdin_interactive(self) -> bool:
        """Neden: GAOSB_INTERACTIVE ortam değişkenine göre interaktif (terminal) modda mı çalışıyoruz?"""
        return os.environ.get("GAOSB_INTERACTIVE", "true").lower() == "true"

    def _notify_captcha_required(self) -> None:
        """
        Neden: BotGuard captcha manuel müdahale gerektirdiğinde kullanıcıyı bilgilendirir.
        Terminal bildirimi her zaman; e-posta yalnızca GAOSB_ALERT_* ortam değişkenleri
        yapılandırılmışsa (best-effort, hata durumunda sessiz) gönderilir.
        """
        logger.warning("BotGuard/captcha manuel müdahale gerektiriyor — kullanıcı bilgilendiriliyor.")

        to_addr = os.environ.get("GAOSB_ALERT_EMAIL_TO")
        smtp_host = os.environ.get("GAOSB_ALERT_SMTP_HOST")
        if not (to_addr and smtp_host):
            logger.info(
                "E-posta uyarısı yapılandırılmamış (GAOSB_ALERT_EMAIL_TO / GAOSB_ALERT_SMTP_HOST yok); "
                "sadece terminal bildirimi yapılıyor."
            )
            return

        try:
            import smtplib
            from email.message import EmailMessage

            msg = EmailMessage()
            msg["Subject"] = "GAOSB otomasyon: manuel captcha gerekli"
            msg["From"] = os.environ.get("GAOSB_ALERT_EMAIL_FROM", to_addr)
            msg["To"] = to_addr
            msg.set_content(
                "GAOSB portalı BotGuard captcha doğrulaması gösterdi. "
                "Lütfen sunucuda açılan tarayıcıda captcha'yı çözüp GAOSB'ye giriş yapın, "
                "ardından terminalde Enter'a basın."
            )
            port = int(os.environ.get("GAOSB_ALERT_SMTP_PORT", "587"))
            user = os.environ.get("GAOSB_ALERT_SMTP_USER")
            pwd = os.environ.get("GAOSB_ALERT_SMTP_PASS")
            with smtplib.SMTP(smtp_host, port, timeout=20) as smtp:
                smtp.starttls()
                if user and pwd:
                    smtp.login(user, pwd)
                smtp.send_message(msg)
            logger.info("Captcha uyarı e-postası gönderildi: %s", to_addr)
        except Exception as ex:
            logger.warning("Captcha uyarı e-postası gönderilemedi: %s", ex)

    def _write_captcha_flag(self) -> None:
        """
        Neden: Dashboard'un captcha durumunu görebilmesi için bekleyen doğrulamayı
        diske işaretler. Job katmanı bu dosyaya job_type/target bilgisi ekler.
        """
        import json
        from datetime import datetime

        CAPTCHA_FLAG_PATH.parent.mkdir(parents=True, exist_ok=True)
        CAPTCHA_FLAG_PATH.write_text(
            json.dumps({
                "detected_at": datetime.now().isoformat(),
                "status": "pending",
            }),
            encoding="utf-8",
        )
        logger.info("Captcha bekleme işareti yazıldı: %s", CAPTCHA_FLAG_PATH)

    def _wait_for_manual_captcha(self) -> None:
        """
        Neden: Kullanıcının açılan tarayıcıda captcha'yı çözüp giriş yapmasını bekler.
        Interaktif olmayan (dashboard/scheduler) çalıştırmada bloklamak yerine flag
        dosyası yazılır ve GaosbCaptchaRequiredError fırlatılır; dashboard kullanıcıyı
        doğrulamaya yönlendirir.
        """
        if self._stdin_interactive():
            # Terminal: kullanıcı çözene kadar bekle (mevcut davranış).
            banner = (
                "\n" + "=" * 60 + "\n"
                "CAPTCHA TESPİT EDİLDİ\n"
                "Açılan tarayıcıda captcha'yı çözün ve GAOSB'ye giriş yapın.\n"
                "Tamamladıktan sonra buraya dönüp Enter'a basın...\n"
                + "=" * 60
            )
            print(banner)
            try:
                input()
            except EOFError as e:
                logger.error("stdin kapalı; manuel captcha beklenemedi.")
                self._write_captcha_flag()
                raise GaosbCaptchaRequiredError(
                    "GAOSB güvenlik doğrulaması gerekiyor. "
                    "Dashboard'dan doğrulamayı tamamlayın."
                ) from e
            logger.info("Kullanıcı captcha'yı geçti, devam ediliyor.")
            return

        # Dashboard/scheduler: flag yaz ve özel hata fırlat.
        logger.warning(
            "Manuel captcha gerekiyor ancak terminal interaktif değil; "
            "flag yazılıp GaosbCaptchaRequiredError fırlatılıyor."
        )
        self._write_captcha_flag()
        raise GaosbCaptchaRequiredError(
            "GAOSB güvenlik doğrulaması gerekiyor. "
            "Dashboard'dan doğrulamayı tamamlayın."
        )

    def _perform_login(self, page, username: str, password: str) -> None:
        """
        Neden: DevExpress tabanlı ASP.NET WebForms login akışını yürütür ve mainpage.aspx'e
        yönlenmeyi bekler. BotGuard clearance mevcutsa (profilde) headless olarak da çalışır.
        """
        try:
            logger.info("Kullanıcı kimlik bilgileri dolduruluyor...")
            page.locator("#ctl00_ContentPlaceHolder1_eUserName_I").first.fill(username)
            page.locator("#ctl00_ContentPlaceHolder1_ePassword_I").first.fill(password)

            # EULA onay kutusu (görünürse)
            eula_cb = page.locator("#ctl00_ContentPlaceHolder1_chkReadcontract_S").first
            try:
                if eula_cb.is_visible(timeout=2000):
                    eula_cb.click()
                    logger.info("EULA sözleşme onay kutusu işaretlendi.")
            except Exception:
                pass

            # Neden: DevExpress butonu görünür <div class="dxb"> ile sarılı, arkasındaki
            # <input type="submit"> pointer events almıyor; görünür div'e tıklamak gerekir.
            login_btn_div = page.locator("#ctl00_ContentPlaceHolder1_LoginButton_CD").first
            try:
                if login_btn_div.is_visible(timeout=3000):
                    login_btn_div.click()
                    logger.info("Login butonu (DevExpress div) tıklandı.")
                else:
                    logger.info("Div görünür değil, JavaScript submit deneniyor...")
                    page.evaluate("document.querySelector('form').submit()")
            except Exception as e:
                logger.error(f"Login butonu tıklanamadı: {e}")
                raise SourceAuthenticationError("gaosb") from e

            page.wait_for_url("**/mainpage.aspx", timeout=30000)
            logger.info("Giriş başarılı, mainpage.aspx sayfasına ulaşıldı.")
        except SourceAuthenticationError:
            raise
        except Exception as e:
            logger.error(f"Kullanıcı girişi veya yönlendirme başarısız oldu: {e}")
            raise SourceAuthenticationError("gaosb") from e

    def _ensure_authenticated(self, username: str, password: str):
        """
        Neden: Persistent context'i (HER ZAMAN headless=False) başlatır, sayfa durumunu
        (mainpage / login / captcha) sınıflandırır ve gereğini yapar; mainpage.aspx'e ulaşmış
        (pw, context, page, captcha_encountered) döner.

        Neden headless=False sabit: BotGuard clearance çerezi headless modda üretilmiyor/geçmiyor.
        Uygulama sunucu değil masaüstünde çalıştığından görünür tarayıcı sorun değil
        (scheduler'da ~30 sn açılıp kapanır). Bu yüzden GAOSB_HEADLESS yok sayılır.

        Akış:
          1. headless=False ile başlat, ana sayfaya git
          2. Captcha sayfası ise → bildir + kullanıcı çözene kadar bekle
          3. page.url'ye göre sınıflandır:
               - "mainpage" içeriyorsa → giriş zaten yapılmış, login atla
               - hâlâ captcha sayfası ise → hata (çözülmemiş)
               - aksi halde (login sayfası) → _perform_login()

        Dönüş: (pw, context, page, captcha_encountered)
        Hata durumunda açılan context/pw kapatılır ve istisna yeniden fırlatılır.
        """
        gaosb_url = os.environ.get("GAOSB_URL", "https://elk.gaosb.org/")
        # Neden: Her zaman görünür tarayıcı — BotGuard clearance yalnızca headed modda geçerli.
        pw, context = self._launch_persistent(headless=False)
        try:
            page = context.new_page()
            captcha_encountered = False

            try:
                logger.info("GAOSB adresine gidiliyor: %s", gaosb_url)
                page.goto(gaosb_url, wait_until="domcontentloaded", timeout=30000)
            except Exception as e:
                logger.error("GAOSB portal ana sayfasına ulaşılamadı: %s", e)
                raise SourceAuthenticationError("gaosb") from e

            # Adım 2: BotGuard captcha sayfası ise kullanıcı çözene kadar bekle.
            # (Zaten headless=False olduğundan tarayıcı görünür; yeniden başlatmaya gerek yok.)
            if self._is_captcha_page(page):
                captcha_encountered = True
                logger.warning("BotGuard captcha tespit edildi.")
                self._notify_captcha_required()
                self._wait_for_manual_captcha()

            # Adım 3: Captcha sonrası (veya doğrudan) sayfa durumunu sınıflandır.
            current_url = (page.url or "").lower()
            if "mainpage" in current_url:
                # Durum A: Portal doğrudan mainpage.aspx'e yönlendirmiş → giriş yapılmış.
                logger.info("mainpage.aspx tespit edildi, giriş zaten yapılmış — login atlanıyor.")
            elif self._is_captcha_page(page):
                # Durum B: Kullanıcı captcha'yı geçmemiş / hâlâ challenge sayfasında.
                logger.error("Captcha hâlâ çözülmemiş görünüyor; giriş yapılamıyor.")
                raise SourceAuthenticationError("gaosb")
            else:
                # Durum C: Login sayfasındayız → otomatik login yap.
                logger.info("Login sayfası tespit edildi, otomatik login yapılıyor...")
                self._perform_login(page, username, password)

            return pw, context, page, captcha_encountered
        except Exception:
            # Neden: Hata halinde açık kaynakları sızdırmadan kapat.
            try:
                context.close()
            except Exception:
                pass
            try:
                pw.stop()
            except Exception:
                pass
            raise

    def check_session(self, headless: Optional[bool] = None) -> Dict[str, Any]:
        """
        Neden: Rapor indirmeden yalnızca oturum/persistent-context durumunu doğrular.
        Test ve health-check için hafif bir yol sağlar (export tetiklemez).

        Not: headless parametresi geriye dönük uyumluluk için tutulur ancak YOK SAYILIR;
        BotGuard clearance yalnızca görünür (headless=False) modda geçerli olduğundan her
        zaman headed çalışılır.

        Dönüş: logged_in, captcha_encountered, final_url, headless, profile_dir, profile_saved
        """
        username = os.environ.get("GAOSB_USERNAME")
        password = os.environ.get("GAOSB_PASSWORD")
        if not username or not password:
            raise ValueError("GAOSB kimlik bilgileri eksik (GAOSB_USERNAME / GAOSB_PASSWORD).")

        pw = None
        context = None
        page = None
        try:
            pw, context, page, captcha_encountered = self._ensure_authenticated(
                username=username, password=password
            )
            final_url = page.url
            logged_in = "mainpage" in (final_url or "").lower()
            profile_saved = USER_DATA_DIR.exists() and any(USER_DATA_DIR.iterdir())
            return {
                "logged_in": logged_in,
                "captcha_encountered": captcha_encountered,
                "final_url": final_url,
                "headless": False,
                "profile_dir": str(USER_DATA_DIR),
                "profile_saved": profile_saved,
            }
        finally:
            try:
                if context is not None:
                    context.close()
            except Exception:
                pass
            try:
                if pw is not None:
                    pw.stop()
            except Exception:
                pass

    # ------------------------------------------------------------------
    # Format dönüştürme yardımcıları
    # ------------------------------------------------------------------
    @staticmethod
    def _is_ole2_file(path: Path) -> bool:
        """
        Neden: Portal, eski OLE2/BIFF (.xls) formatındaki dosyayı .xlsx uzantısıyla
        gönderiyor; uzantıya güvenilemez, dosya imzasından (magic bytes) tespit edilir.
        """
        try:
            with open(path, "rb") as f:
                return f.read(8) == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
        except Exception:
            return False

    def _convert_to_xlsx(self, ole2_path: Path) -> Path:
        """
        Neden: OLE2 (.xls/BIFF) içerikli dosya .xlsx uzantısıyla geldiğinde Excel ve
        openpyxl açamıyor. xlrd ile okunup gerçek OOXML (.xlsx) olarak yeniden yazılır;
        orijinal OLE2 dosya kaybolmasın diye yanına .xls uzantısıyla taşınır.
        Dönüş: gerçek .xlsx dosyasının yolu (orijinal ad korunur).
        """
        import xlrd
        import openpyxl

        ole2_path = Path(ole2_path)
        book = xlrd.open_workbook(str(ole2_path))

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sheet in book.sheets():
            # Neden: xlsx sayfa adı en fazla 31 karakter olabilir.
            ws = wb.create_sheet(title=(sheet.name or "Sheet1")[:31])
            for r in range(sheet.nrows):
                for c in range(sheet.ncols):
                    cell = sheet.cell(r, c)
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        # Neden: BIFF tarihleri float seri numarasıdır; datetime'a çevrilmezse
                        # xlsx'te anlamsız sayı olarak görünür.
                        try:
                            value = xlrd.xldate.xldate_as_datetime(cell.value, book.datemode)
                        except Exception:
                            pass
                    elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                        value = bool(cell.value)
                    elif cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR):
                        value = None
                    if value is not None and value != "":
                        ws.cell(row=r + 1, column=c + 1, value=value)

        # Orijinal OLE2 dosyayı .xls olarak sakla, gerçek xlsx'i aynı adla yaz.
        xls_backup = ole2_path.with_suffix(".xls")
        if xls_backup.exists():
            xls_backup.unlink()
        ole2_path.rename(xls_backup)

        wb.save(str(ole2_path))
        wb.close()
        logger.info(
            "OLE2 (.xls) içerik gerçek xlsx'e dönüştürüldü: %s (orijinal: %s)",
            ole2_path, xls_backup.name,
        )
        return ole2_path

    # ------------------------------------------------------------------
    # Ana rapor indirme akışı
    # ------------------------------------------------------------------
    def download_report(self, output_dir: Path, date_from: Optional[str] = None, date_to: Optional[str] = None, **kwargs) -> Path:
        """
        Neden: Belirtilen tarih aralığı için GAOSB portalından sayaç sorgu excel raporunu indirir.
        Persistent context + BotGuard captcha fallback ile kimlik doğrular.
        """
        # Parametre kontrolü ve kwargs'tan fallback okuma
        date_from = date_from or kwargs.get("date_from")
        date_to = date_to or kwargs.get("date_to")
        if not date_from or not date_to:
            logger.error("date_from veya date_to parametreleri eksik.")
            raise ValueError("date_from ve date_to parametreleri zorunludur.")

        # Yapılandırma ortam değişkenlerinden okunur
        username = os.environ.get("GAOSB_USERNAME")
        password = os.environ.get("GAOSB_PASSWORD")

        if not username or not password:
            logger.error("GAOSB_USERNAME veya GAOSB_PASSWORD ortam değişkenleri eksik.")
            raise ValueError("GAOSB kimlik bilgileri eksik (GAOSB_USERNAME / GAOSB_PASSWORD).")

        # Neden: headless modu artık çözümlenmiyor — GAOSB_HEADLESS ve headless kwargs YOK SAYILIR.
        # BotGuard clearance yalnızca görünür tarayıcıda geçerli olduğundan her zaman headed çalışılır.

        # Tarih formatlama yardımcı fonksiyonları (Türkçe portal için DD.MM.YYYY)
        def format_date_tr(date_str: str) -> str:
            if not date_str:
                return ""
            if "-" in date_str:
                try:
                    parts = date_str.split("-")
                    if len(parts) == 3 and len(parts[0]) == 4:  # YYYY-MM-DD
                        return f"{parts[2]}.{parts[1]}.{parts[0]}"
                except Exception:
                    pass
            return date_str

        def get_yyyymmdd(date_str: str) -> str:
            clean = date_str.replace(".", "").replace("-", "")
            if len(clean) == 8:
                if "." in date_str:  # DD.MM.YYYY
                    return clean[4:8] + clean[2:4] + clean[0:2]
                return clean  # YYYYMMDD
            return clean

        formatted_date_from = format_date_tr(date_from)
        formatted_date_to = format_date_tr(date_to)

        # Arama ve doldurma için selector listeleri
        from_selectors = [
            "#ctl00_ContentPlaceHolder1_Date1_I",
            "input[id*='Date1_I']",
            "input[id*='dtFrom']",
            "input[id*='DateFrom']",
            "input[id*='StartDate']",
            "input[name*='dtFrom']",
            "input[name*='DateFrom']",
        ]

        to_selectors = [
            "#ctl00_ContentPlaceHolder1_Date2_I",
            "input[id*='Date2_I']",
            "input[id*='dtTo']",
            "input[id*='DateTo']",
            "input[id*='EndDate']",
            "input[name*='dtTo']",
            "input[name*='DateTo']",
        ]

        query_patterns = [
            "input[value*='Sorgula']",
            "input[value*='Ara']",
            "input[value*='Listele']",
            "input[value*='Query']",
            "input[value*='Search']",
            "input[value*='Filter']",
            "[id*='btnQuery']",
            "[id*='btnSearch']",
            "[id*='btnFilter']",
            "[id*='btnSorgula']",
            ".dxbButton:has-text('Sorgula')",
            ".dxbButton:has-text('Ara')",
        ]

        export_selectors = [
            "#ctl00_ContentPlaceHolder1_ButtonExportExcel_CD",
            "[id*='ButtonExportExcel']",
            "[id*='Export']",
            "[id*='Excel']",
        ]

        logger.info("GAOSB veri toplama akışı başlatılıyor...")

        # Çıktı ve ekran görüntüsü klasörlerini hazırla
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        screenshot_dir = output_dir / "screenshots"
        screenshot_dir.mkdir(parents=True, exist_ok=True)

        pw = None
        context = None
        page = None

        def take_screenshot(name: str):
            try:
                if page is not None:
                    page.screenshot(path=str(screenshot_dir / f"{name}.png"))
                    logger.info(f"Ekran görüntüsü kaydedildi: {name}.png")
            except Exception as ex:
                logger.warning(f"Ekran görüntüsü alınamadı ({name}): {ex}")

        try:
            # Adım 1-2: Persistent context (headed) + BotGuard captcha bekleme + login
            pw, context, page, captcha_encountered = self._ensure_authenticated(
                username=username, password=password
            )
            take_screenshot("02_mainpage")

            # Adım 3: Sayaç Sorgu Sayfasına Geç
            try:
                logger.info("Sayaç sorgu sayfasına geçiş yapılıyor...")
                page.evaluate("__doPostBack('ctl00$mnuQuery', '')")
                page.wait_for_load_state("domcontentloaded", timeout=20000)
                time.sleep(2)
                take_screenshot("03_query_page")
            except Exception as e:
                logger.error(f"Sayaç sorgu sayfasına geçiş yapılamadı: {e}")
                raise

            # Adım 4: VIEWSTATE Parse (Değerler post'a eklenir ama logda gösterilmez)
            try:
                logger.info("VIEWSTATE değerleri çözümleniyor...")
                viewstate = page.locator("#__VIEWSTATE").get_attribute("value") or ""
                viewstate_gen = page.locator("#__VIEWSTATEGENERATOR").get_attribute("value") or ""
                event_val = page.locator("#__EVENTVALIDATION").get_attribute("value") or ""

                # Güvenlik ve log temizliği için sadece uzunluk bilgileri yazılır
                logger.info(f"__VIEWSTATE okundu (Uzunluk: {len(viewstate)})")
                logger.info(f"__VIEWSTATEGENERATOR okundu (Uzunluk: {len(viewstate_gen)})")
                logger.info(f"__EVENTVALIDATION okundu (Uzunluk: {len(event_val)})")
            except Exception as e:
                logger.error(f"VIEWSTATE/EventValidation bilgileri okunamadı: {e}")
                raise

            # Adım 5: Tarih Doldurma ve Sorgulama
            try:
                logger.info("Sayfa üzerindeki text input alanları taranıyor...")
                text_inputs = page.query_selector_all("input[type='text'], input:not([type])")
                for idx, inp in enumerate(text_inputs):
                    inp_id = inp.get_attribute("id") or ""
                    inp_name = inp.get_attribute("name") or ""
                    inp_placeholder = inp.get_attribute("placeholder") or ""
                    logger.info(f"  - Input [{idx}]: id='{inp_id}', name='{inp_name}', placeholder='{inp_placeholder}'")

                # Başlangıç ve Bitiş tarih input alanlarını tespit et
                from_el = None
                for sel in from_selectors:
                    try:
                        el = page.locator(sel).first
                        if el.is_visible(timeout=1000):
                            from_el = el
                            logger.info(f"date_from için eşleşen selector bulundu: {sel}")
                            break
                    except Exception:
                        continue

                to_el = None
                for sel in to_selectors:
                    try:
                        el = page.locator(sel).first
                        if el.is_visible(timeout=1000):
                            to_el = el
                            logger.info(f"date_to için eşleşen selector bulundu: {sel}")
                            break
                    except Exception:
                        continue

                # Eğer tarih alanları bulunamazsa hata fırlatılır (tahmin edilmez)
                if not from_el or not to_el:
                    logger.error("Başlangıç veya bitiş tarihi giriş alanları sayfada tespit edilemedi.")
                    raise ValueError("Tarih input selector'ları bulunamadı. Lütfen sayfa yapısını kontrol edin.")

                logger.info(f"Tarih alanları dolduruluyor: {formatted_date_from} - {formatted_date_to}")

                # Neden: DevExpress DateEdit bileşeni fill() ile değer almayabilir.
                # JavaScript ile value set edip change event tetiklemek daha güvenilir.
                def fill_dx_date(el, value: str):
                    el.click()
                    time.sleep(0.3)
                    # Önce triple-click ve Control+A ile tümünü seç, sonra yaz
                    el.click(click_count=3)
                    time.sleep(0.2)
                    page.keyboard.press("Control+A")
                    time.sleep(0.2)
                    el.type(value, delay=80)
                    time.sleep(0.3)
                    page.keyboard.press("Tab")
                    time.sleep(0.3)

                fill_dx_date(from_el, formatted_date_from)
                logger.info(f"  date_from girildi: {formatted_date_from}")

                fill_dx_date(to_el, formatted_date_to)
                logger.info(f"  date_to girildi: {formatted_date_to}")

                # Neden: Tarihlerin input'a yerleştiğini doğrulamak için mevcut değeri logla
                from_val = from_el.input_value()
                to_val = to_el.input_value()
                logger.info(f"  Doğrulama — Date1 değeri: '{from_val}', Date2 değeri: '{to_val}'")
                time.sleep(1)
                take_screenshot("04_dates_filled")

                # Neden: Portal varsayılan endeks kodu P.01.1.9 olmayabilir.
                # Mahsuplaşma için Aktif enerji çekiş LoadProfile seçilmeli.
                logger.info("Endeks kodu seçiliyor: P.01.1.9...")
                try:
                    indexer = page.locator("#ctl00_ContentPlaceHolder1_cIndexer_I").first
                    if indexer.is_visible(timeout=3000):
                        # Mevcut değeri logla
                        current_val = indexer.input_value()
                        logger.info(f"Mevcut endeks: {current_val}")

                        # Neden: DevExpress ComboBox'ta input'a tıklamak listeyi açmaz;
                        # açılır ok butonuna (B-1) tıklamak, olmazsa Alt+Down göndermek gerekir.
                        dropdown_opened = False
                        for btn_sel in ["#ctl00_ContentPlaceHolder1_cIndexer_B-1",
                                        "[id*='cIndexer_B-1']",
                                        "[id*='cIndexer_B']"]:
                            try:
                                btn = page.locator(btn_sel).first
                                if btn.is_visible(timeout=1000):
                                    btn.click()
                                    dropdown_opened = True
                                    logger.info(f"Dropdown butonu tıklandı: {btn_sel}")
                                    break
                            except Exception:
                                continue
                        if not dropdown_opened:
                            indexer.click()
                            time.sleep(0.3)
                            page.keyboard.press("Alt+ArrowDown")
                            logger.info("Dropdown Alt+Down ile açılmaya çalışıldı.")
                        time.sleep(0.8)

                        # P.01.1.9 içeren option'ı bul ve tıkla
                        # Neden: DevExpress liste öğeleri td.dxeListBoxItem olarak render edilir.
                        option_selector = (
                            "[id*='cIndexer_DDD'] td[class*='dxeListBoxItem'], "
                            "td.dxeListBoxItem, .dxeListBoxItemRow, li[class*='dxe']"
                        )
                        page.wait_for_selector(option_selector, timeout=5000)
                        options = page.locator(option_selector).all()

                        selected = False
                        for opt in options:
                            try:
                                text = (opt.inner_text() or "").strip()
                            except Exception:
                                continue
                            if not text:
                                continue
                            logger.info(f"  Option: {text}")
                            if "P.01" in text or "01.1.9" in text or "loadprofile" in text.lower():
                                opt.click()
                                selected = True
                                logger.info(f"Endeks seçildi: {text}")
                                break

                        if not selected:
                            logger.warning("P.01.1.9 option bulunamadı, varsayılan kullanılıyor")

                        # Neden: Seçim DevExpress callback tetikleyebilir; kısa bekleme ile
                        # değerin input'a yerleşmesini garanti altına al.
                        time.sleep(1)
                        try:
                            logger.info(f"Seçim sonrası endeks değeri: {indexer.input_value()}")
                        except Exception:
                            pass
                        take_screenshot("04b_indexer_selected")
                except Exception as e:
                    logger.warning(f"Endeks seçimi atlandı: {e}")

                # Neden: Önce sayfadaki tüm butonları logla, sorgu butonunun
                # gerçek ID/name değerini tespit etmek için.
                all_buttons = page.evaluate("""() => {
                    return [...document.querySelectorAll(
                        'input[type="button"], input[type="submit"], .dxbButton, .dxb'
                    )].map(el => ({
                        tag: el.tagName,
                        id: el.id || '',
                        name: el.name || '',
                        value: el.value || el.innerText || '',
                        visible: el.offsetParent !== null
                    }));
                }""")
                for b in all_buttons:
                    logger.info(f"  Buton: id='{b.get('id')}' name='{b.get('name')}' value='{b.get('value')}' visible={b.get('visible')}")

                # Neden: DevExpress sorgu butonu PostBack ile çalışıyor,
                # görünür div'e tıklamak gerekiyor.
                query_clicked = False

                # Strateji 1: BtnSubmit veya bilinen ID'lere göre tıkla
                known_query_ids = [
                    "#ctl00_ContentPlaceHolder1_BtnSubmit_CD",
                    "#ctl00_ContentPlaceHolder1_btnQuery_CD",
                    "#ctl00_ContentPlaceHolder1_btnSorgula_CD",
                    "#ctl00_ContentPlaceHolder1_Button1_CD",
                ]
                for qid in known_query_ids:
                    try:
                        el = page.locator(qid).first
                        if el.is_visible(timeout=1000):
                            el.click()
                            logger.info(f"Sorgu butonu tıklandı: {qid}")
                            query_clicked = True
                            break
                    except Exception:
                        continue

                # Strateji 2: .dxb div içinde Sorgula/Listele/Ara metni ara
                if not query_clicked:
                    for text in ["Sorgula", "Listele", "Ara", "Query", "Getir"]:
                        try:
                            el = page.locator(f".dxb:has-text('{text}')").first
                            if el.is_visible(timeout=1000):
                                el.click()
                                logger.info(f"Sorgu butonu metin ile tıklandı: {text}")
                                query_clicked = True
                                break
                        except Exception:
                            continue

                # Strateji 3: JavaScript PostBack ile dene
                if not query_clicked:
                    logger.info("JavaScript PostBack ile sorgu deneniyor...")
                    try:
                        page.evaluate("__doPostBack('ctl00$ContentPlaceHolder1$BtnSubmit', '')")
                        query_clicked = True
                        logger.info("PostBack ile sorgu gönderildi.")
                    except Exception as e:
                        logger.warning(f"PostBack denemesi başarısız: {e}")

                if not query_clicked:
                    logger.error("Sorgu butonu hiçbir yöntemle bulunamadı.")
                    raise ValueError("Sorgu butonu bulunamadı.")

                take_screenshot("05_query_clicked")

                # Neden: PostBack sonrası sayfa yenilenir, grid yüklenmesini bekle.
                logger.info("Grid yüklenmesi bekleniyor (.dxgvDataRow veya .dxgvEmptyDataRow)...")
                try:
                    page.wait_for_selector(
                        ".dxgvDataRow, tr.dxgvDataRow_Default, .dxgvEmptyDataRow, tr.dxgvEmptyDataRow_Default",
                        timeout=20000
                    )
                    logger.info("Veri gridi başarıyla yüklendi.")
                except Exception as ex:
                    logger.warning(f"Grid yükleme beklentisi timeout oldu, yine de export adımına geçiliyor: {ex}")

                time.sleep(1.5)
                take_screenshot("06_grid_loaded")
            except Exception as e:
                logger.error(f"Tarih doldurma veya sorgulama adımı başarısız oldu: {e}")
                raise

            # Adım 6: Excel Export ve Rapor Kaydetme
            try:
                logger.info("Export butonu aranıyor...")
                export_btn = None
                for sel in export_selectors:
                    try:
                        btn = page.locator(sel).first
                        if btn.is_visible(timeout=1000):
                            export_btn = btn
                            logger.info(f"Export butonu eşleşen selector: {sel}")
                            break
                    except Exception:
                        continue

                if not export_btn:
                    logger.error("Excel export butonu bulunamadı.")
                    raise ValueError("Export butonu bulunamadı.")

                logger.info("İndirme işlemi bekleniyor...")
                with page.expect_download(timeout=45000) as download_info:
                    export_btn.click()
                download = download_info.value

                # Çıktı klasörünü hazırla
                from_yyyymmdd = get_yyyymmdd(formatted_date_from)
                to_yyyymmdd = get_yyyymmdd(formatted_date_to)
                dest_filename = f"gaosb_{from_yyyymmdd}_{to_yyyymmdd}.xlsx"
                dest_path = output_dir / dest_filename

                download.save_as(str(dest_path))
                logger.info(f"Excel başarıyla indirildi ve kaydedildi: {dest_path}")
                take_screenshot("07_export_completed")

                # Neden: Portal OLE2 (.xls) içeriği .xlsx uzantısıyla gönderiyor;
                # Excel/openpyxl açabilsin diye gerçek xlsx'e dönüştürülür.
                if self._is_ole2_file(dest_path):
                    logger.info("OLE2 (.xls) formatı tespit edildi, gerçek xlsx'e dönüştürülüyor...")
                    dest_path = self._convert_to_xlsx(dest_path)

                return dest_path
            except Exception as e:
                logger.error(f"Excel export veya indirme adımı başarısız oldu: {e}")
                raise
        except Exception as e:
            # Tarayıcı kapanmadan önce hata ekran görüntüsünü kaydet
            take_screenshot("error_screenshot")
            logger.error(f"GAOSB extraction işlemi sırasında genel hata oluştu: {e}")
            raise
        finally:
            # Neden: Persistent context ve Playwright süreçlerini temiz kapat (orphan engelle).
            try:
                if context is not None:
                    context.close()
            except Exception:
                pass
            try:
                if pw is not None:
                    pw.stop()
            except Exception:
                pass
