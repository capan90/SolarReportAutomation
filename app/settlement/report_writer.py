import datetime
from pathlib import Path
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.settlement.models import HourlySettlement

class SettlementReportWriter:
    """
    Neden: Mahsuplaşma sonuçlarını Excel dosyasına standart şablon ve biçimlendirme kurallarına göre yazar.
    """

    @staticmethod
    def _parse_timestamp(timestamp: str) -> tuple:
        """
        Neden: 'YYYY-MM-DD HH:00:00' timestamp'ını rapor formatına
        (DD.MM.YYYY, 'HH:00-HH+1:00') dönüştürmek — iki sayfada da kullanılır.
        """
        try:
            date_part, time_part = timestamp.split(' ')
            saat = int(time_part.split(':')[0])
        except Exception:
            normalized = timestamp.replace("T", " ")
            date_part, time_part = normalized.split(' ')
            saat = int(time_part.split(':')[0])

        try:
            dt_obj = datetime.datetime.strptime(date_part, "%Y-%m-%d")
            tarih_str = dt_obj.strftime("%d.%m.%Y")
        except Exception:
            tarih_str = date_part

        saat_araligi = f"{saat:02d}:00-{(saat + 1):02d}:00"
        if saat == 23:
            saat_araligi = "23:00-24:00"
        return tarih_str, saat_araligi

    def write(self, settlements: List[HourlySettlement], output_path: Path, isolar_df=None, summary_title=None) -> Path:
        """
        Neden: HourlySettlement listesini alır ve belirtilen output_path konumunda
        şekillendirilmiş bir Excel raporu üretir.

        isolar_df: load_isolar_curve()'den gelen DataFrame (opsiyonel).
        Verilirse "GES Kırılımı" ikinci sayfası eklenir.
        """
        if not summary_title:
            if settlements:
                unique_dates = set()
                for s in settlements:
                    tarih_str, _ = self._parse_timestamp(s.timestamp)
                    unique_dates.add(tarih_str)
                
                if len(unique_dates) == 1:
                    summary_title = f"{next(iter(unique_dates))} Günlük Toplamları"
                else:
                    first_ts = settlements[0].timestamp.replace("T", " ")
                    date_part = first_ts.split(' ')[0]
                    try:
                        year, month, _ = map(int, date_part.split('-'))
                        AYLAR = {
                            1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan",
                            5: "Mayıs", 6: "Haziran", 7: "Temmuz", 8: "Ağustos",
                            9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"
                        }
                        summary_title = f"{AYLAR.get(month, '')} {year} Aylık Toplamları"
                    except Exception:
                        summary_title = "AYLIK TOPLAMLAR"
            else:
                summary_title = "AYLIK TOPLAMLAR"
        # Neden: Yeni bir openpyxl çalışma kitabı oluşturup aktif sayfayı seçeriz.
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mahsuplaşma Raporu"

        # Neden: Grid çizgilerinin görünür olmasını garanti altına alırız.
        ws.views.sheetView[0].showGridLines = True

        # Neden: Stil tanımlarını (font, arka plan dolgusu ve kenarlıklar) hazırlarız.
        font_header = Font(name="Calibri", size=11, bold=True)
        font_data = Font(name="Calibri", size=11, bold=False)
        font_bold = Font(name="Calibri", size=11, bold=True)

        fill_gray = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        double_bottom_border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='double', color='000000')
        )

        # Neden: Ana veri tablosu başlıklarını ekler ve stillerini uygularız.
        headers = [
            "TARİH",
            "SAAT ARALIĞI",
            "TÜKETİM (KWH)",
            "ÜRETİM (KWH)",
            "MAHSUP EDİLEN (KWH)",
            "ŞEBEKEDEN ÇEKİŞ (KWH)",
            "FAZLA SATIŞ (KWH)"
        ]

        for col_idx, h_text in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h_text)
            cell.font = font_header
            cell.fill = fill_gray
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # Neden: Saatlik mahsuplaşma verilerini satır satır yazarız.
        current_row = 2
        for s in settlements:
            # Neden: Timestamp alanından tarih ve saat aralığını ayrıştırırız.
            tarih_str, saat_araligi = self._parse_timestamp(s.timestamp)

            row_data = [
                tarih_str,
                saat_araligi,
                s.consumption_kwh,
                s.production_kwh,
                s.settled_kwh,
                s.grid_import_kwh,
                s.grid_export_kwh
            ]

            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = font_data
                cell.border = thin_border

                if col_idx in [1, 2]:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '0.00'

            current_row += 1

        # Neden: Tablonun en sonuna TOPLAM satırını ekler ve Excel SUM formüllerini yazarız.
        total_row_idx = current_row
        ws.cell(row=total_row_idx, column=1, value="TOPLAM").font = font_bold
        ws.cell(row=total_row_idx, column=1).border = double_bottom_border
        ws.cell(row=total_row_idx, column=2).border = double_bottom_border

        for col_idx in range(3, 8):
            col_letter = get_column_letter(col_idx)
            formula = f"=SUM({col_letter}2:{col_letter}{total_row_idx - 1})"
            cell = ws.cell(row=total_row_idx, column=col_idx, value=formula)
            cell.font = font_bold
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal="right")
            cell.border = double_bottom_border

        # Neden: Sağ tarafta yer alan özet bloğu (I-J sütunları) oluştururuz.
        # Dinamik Özet Başlığı
        ws.merge_cells("I2:J2")
        title_cell = ws["I2"]
        title_cell.value = summary_title
        title_cell.font = font_header
        title_cell.fill = fill_gray
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.border = thin_border
        ws["J2"].border = thin_border

        summary_labels = [
            ("Toplam Tüketim", "C"),
            ("Toplam Üretim", "D"),
            ("Toplam Mahsup", "E"),
            ("Toplam Şebekeden Çekiş", "F"),
            ("Toplam Fazla Satış", "G")
        ]

        for idx, (label, src_col) in enumerate(summary_labels, start=3):
            # Etiket hücresi (I sütunu)
            lbl_cell = ws.cell(row=idx, column=9, value=label)
            lbl_cell.font = font_data
            lbl_cell.border = thin_border

            # Değer hücresi (J sütunu) -> Formül ile TOPLAM satırına bağlanır.
            val_cell = ws.cell(row=idx, column=10, value=f"={src_col}{total_row_idx}")
            val_cell.font = font_bold
            val_cell.number_format = '0.00'
            val_cell.alignment = Alignment(horizontal="right")
            val_cell.border = thin_border

        # Neden: Hücre içeriklerinin tam sığması için sütun genişliklerini dinamik olarak ayarlarız.
        # Kolon genişliği ayarı veri sığmama (###) hatasını önler.
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            # Yalnızca A-J aralığındaki sütunları boyutlandır
            if col[0].column <= 10:
                max_len = 0
                for cell in col:
                    val = str(cell.value or '')
                    if cell.number_format == '0.00' and isinstance(cell.value, (int, float)):
                        val = f"{cell.value:.2f}"
                    max_len = max(max_len, len(val))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Başlık satır yüksekliğini ayarla
        ws.row_dimensions[1].height = 24

        # Neden: iSolar DataFrame verilmişse santral bazlı üretim kırılımını
        # ikinci sayfaya yazarız — kullanıcı hangi GES'in ne ürettiğini görmek ister.
        if isolar_df is not None and len(isolar_df) > 0:
            ges_cols = sorted(
                [c for c in isolar_df.columns if c.startswith("ges_") and c.endswith("_kwh")],
                key=lambda c: int(c.split('_')[1])
            )
            if ges_cols:
                ws2 = wb.create_sheet(title="GES Kırılımı")
                ws2.views.sheetView[0].showGridLines = True

                # Başlıklar: TARİH | SAAT ARALIĞI | GES 2 | ... | TOPLAM ÜRETİM
                ges_headers = [f"GES {c.split('_')[1]}" for c in ges_cols]
                headers2 = ["TARİH", "SAAT ARALIĞI"] + ges_headers + ["TOPLAM ÜRETİM"]

                for col_idx, h_text in enumerate(headers2, start=1):
                    cell = ws2.cell(row=1, column=col_idx, value=h_text)
                    cell.font = font_header
                    cell.fill = fill_gray
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin_border

                # Saatlik delta üretim satırları
                current_row = 2
                for _, row in isolar_df.iterrows():
                    tarih_str, saat_araligi = self._parse_timestamp(str(row['timestamp']))
                    row_data = [tarih_str, saat_araligi]
                    row_data += [float(row[c]) for c in ges_cols]
                    row_data.append(float(row['production_kwh']))

                    for col_idx, val in enumerate(row_data, start=1):
                        cell = ws2.cell(row=current_row, column=col_idx, value=val)
                        cell.font = font_data
                        cell.border = thin_border
                        if col_idx in [1, 2]:
                            cell.alignment = Alignment(horizontal="center")
                        else:
                            cell.alignment = Alignment(horizontal="right")
                            cell.number_format = '0.00'
                    current_row += 1

                # TOPLAM satırı (her sayısal sütun için SUM formülü)
                total_row_idx = current_row
                ws2.cell(row=total_row_idx, column=1, value="TOPLAM").font = font_bold
                ws2.cell(row=total_row_idx, column=1).border = double_bottom_border
                ws2.cell(row=total_row_idx, column=2).border = double_bottom_border

                for col_idx in range(3, len(headers2) + 1):
                    col_letter = get_column_letter(col_idx)
                    formula = f"=SUM({col_letter}2:{col_letter}{total_row_idx - 1})"
                    cell = ws2.cell(row=total_row_idx, column=col_idx, value=formula)
                    cell.font = font_bold
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal="right")
                    cell.border = double_bottom_border

                # Sütun genişlikleri (birinci sayfayla aynı yaklaşım)
                for col in ws2.columns:
                    col_letter = get_column_letter(col[0].column)
                    if col[0].column <= len(headers2):
                        max_len = 0
                        for cell in col:
                            val = str(cell.value or '')
                            if cell.number_format == '0.00' and isinstance(cell.value, (int, float)):
                                val = f"{cell.value:.2f}"
                            max_len = max(max_len, len(val))
                        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)

                ws2.row_dimensions[1].height = 24

        # Neden: Oluşturulan çalışma kitabını diskte hedef konuma kaydederiz.
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

        return output_path
