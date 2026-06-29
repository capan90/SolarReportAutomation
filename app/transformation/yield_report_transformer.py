import re
import sys
from pathlib import Path
from datetime import datetime, date
from typing import List, Dict, Any, Tuple
import openpyxl

from app.core.config import settings
from app.core.logger import setup_logger
from app.core.exceptions import IsolarError
from app.profiling.profile_models import WorkbookProfile
from app.canonical.mapping_registry import MappingRegistry
from app.transformation.transform_models import (
    TransformResult,
    TransformedRecord,
    TransformIssue
)
from app.transformation import transform_rules

logger = setup_logger("YieldReportTransformer")

class YieldReportTransformer:
    """
    Neden: Şema doğrulaması başarılı (SUCCESS) olmuş Excel verilerini,
    Canonical Mapping tanımlarına göre parse edip veritabanına yüklenmeye hazır
    canonical record listesine dönüştürmek ve çıktıyı JSON olarak kaydetmek.
    """
    def __init__(self):
        self.transformed_dir = settings.download_directory.parent / "transformed"
        self.mapping_registry = MappingRegistry()

    def ensure_transformed_directory(self) -> None:
        """
        Neden: Dönüştürülmüş verilerin kaydedileceği dizinin varlığından emin olmak.
        """
        if not self.transformed_dir.exists():
            logger.info(f"Dönüştürülmüş çıktı dizini oluşturuluyor: {self.transformed_dir}")
            self.transformed_dir.mkdir(parents=True, exist_ok=True)

    def transform(self, file_path: Path, profile: WorkbookProfile, validation_status: str) -> TransformResult:
        """
        Neden: Excel dosyasını ve profil verilerini alıp şema eşlemesine göre
        canonical veri dönüşümü sağlamak.
        """
        logger.info(f"Dönüşüm süreci başlatılıyor. Dosya: {file_path}")
        
        # 1. Validation SUCCESS Guard Clause
        if validation_status.upper() != "SUCCESS":
            logger.warning(f"Doğrulama durumu başarılı değil ({validation_status}). Dönüşüm iptal ediliyor.")
            return TransformResult(
                status="FAILED",
                source_file=file_path.name,
                mapping_key="isolar_yield_report_v1",
                generated_at=datetime.now().isoformat(),
                total_rows=0,
                transformed_rows=0,
                failed_rows=0,
                records=[],
                issues=[
                    TransformIssue(
                        row=None,
                        column=None,
                        field=None,
                        rule="validation_guard",
                        message=f"Transformation aborted because schema validation status is '{validation_status}'. Only 'SUCCESS' status is allowed.",
                        raw_value=None
                    )
                ]
            )

        # 2. Mapping Şablonunu Getir
        mapping = self.mapping_registry.get_mapping("isolar_yield_report_v1")
        if not mapping:
            raise IsolarError("isolar_yield_report_v1 mapping şablonu bulunamadı.")

        # 3. Excel Dosyasını Yükle
        try:
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
        except Exception as e:
            raise IsolarError(f"Excel dosyası yüklenirken hata oluştu: {e}")

        # 4. Data Rolündeki Sayfayı Bul ve İndeksleri Al
        data_sheet_profile = next((s for s in profile.sheets if s.sheet_role == "data"), None)
        if not data_sheet_profile:
            wb.close()
            raise IsolarError("Profil verilerinde 'data' rolüne sahip sayfa bulunamadı.")

        ws = wb[data_sheet_profile.name]
        header_row_index = data_sheet_profile.header_row_index
        data_start_row_index = data_sheet_profile.data_start_row_index

        # 5. Rapor Tarihini Metadata Satırından Çıkar (Çiftleme Önleme ve Zaman Boyutu İçin)
        report_date = None
        if data_sheet_profile.metadata_rows:
            # Metadata satırının ilk hücresini tara (örn: Yield report_Day_20260629)
            first_meta_val = str(data_sheet_profile.metadata_rows[0][0])
            date_match = re.search(r"Day_(\d{8})", first_meta_val)
            if date_match:
                date_str = date_match.group(1)
                try:
                    report_date = date(int(date_str[:4]), int(date_str[4:6]), int(date_str[6:8]))
                    logger.info(f"Rapor tarihi meta veriden okundu: {report_date}")
                except ValueError:
                    pass
                    
        if not report_date:
            # Dosya adından yedek tarih çıkar (örn: raw_isolar_20260629_...)
            fn_match = re.search(r"raw_isolar__?(\d{8})", profile.file_name)
            if fn_match:
                date_str = fn_match.group(1)
                report_date = date(int(date_str[:4]), int(date_str[4:6]), int(date_str[6:8]))
                logger.info(f"Rapor tarihi dosya adından okundu: {report_date}")
            else:
                report_date = date.today()
                logger.warning(f"Rapor tarihi tespit edilemedi. Bugünkü tarih kullanılıyor: {report_date}")

        # 6. Kolon İndekslerini Eşleştir
        # Excel başlık satırındaki hücre değerlerini oku
        excel_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[header_row_index]]
        
        # Mapping alanı -> Excel 0-based kolon indeksi haritası
        col_mapping_indices: Dict[str, int] = {}
        for m_field in mapping.mappings:
            col_idx = -1
            # Canonical eşleşme veya alias arama
            for idx, h_val in enumerate(excel_headers):
                if h_val.lower() == m_field.source_column.lower():
                    col_idx = idx
                    break
                if any(alias.lower() == h_val.lower() for alias in m_field.source_aliases):
                    col_idx = idx
                    break
                    
            if col_idx != -1:
                col_mapping_indices[m_field.canonical_field] = col_idx

        # 7. Satır Bazlı Dönüşüm İşlemi
        transformed_records: List[TransformedRecord] = []
        issues: List[TransformIssue] = []
        
        rows = list(ws.iter_rows(values_only=True))
        data_rows_to_process = rows[data_start_row_index - 1:]
        total_rows = len(data_rows_to_process)
        transformed_rows_count = 0
        failed_rows_count = 0

        for r_offset, row in enumerate(data_rows_to_process):
            source_row_num = data_start_row_index + r_offset
            
            # Satırın tamamen boş olup olmadığını denetle
            if all(v is None or str(v).strip() == "" for v in row):
                total_rows -= 1
                continue

            row_has_error = False
            
            # Varlıklar için veri hazneleri
            plant_data: Dict[str, Any] = {}
            gen_data: Dict[str, Any] = {
                "date": report_date  # CDM gereği tarih her zaman enjekte edilir
            }

            for m_field in mapping.mappings:
                canonical_field = m_field.canonical_field
                entity = m_field.entity
                rule_name = m_field.transform_rule
                
                # Excel'deki kolon indeksini al
                col_idx = col_mapping_indices.get(canonical_field)
                if col_idx is None or col_idx >= len(row):
                    if m_field.required:
                        row_has_error = True
                        issues.append(
                            TransformIssue(
                                row=source_row_num,
                                column=m_field.source_column,
                                field=canonical_field,
                                rule="missing_column_on_row",
                                message=f"Zorunlu alan '{canonical_field}' için Excel kolonu satırda bulunamadı.",
                                raw_value=None
                            )
                        )
                    continue

                raw_val = row[col_idx]

                # Kuralı işlet
                try:
                    rule_func = getattr(transform_rules, rule_name, None)
                    if not rule_func:
                        raise AttributeError(f"Dönüşüm kuralı bulunamadı: '{rule_name}'")
                    
                    if rule_name == "parse_revenue_currency":
                        # Tuple dönen özel kural (revenue_today, revenue_currency)
                        val, currency = rule_func(raw_val)
                        gen_data["revenue_today"] = val
                        gen_data["revenue_currency"] = currency
                    else:
                        transformed_val = rule_func(raw_val)
                        
                        # Varlık tipine göre ilgili hazneye ekle
                        if entity == "solar_plant":
                            plant_data[canonical_field] = transformed_val
                        elif entity == "daily_generation":
                            gen_data[canonical_field] = transformed_val
                            
                except Exception as e:
                    row_has_error = True
                    issues.append(
                        TransformIssue(
                            row=source_row_num,
                            column=m_field.source_column,
                            field=canonical_field,
                            rule=rule_name,
                            message=f"Dönüşüm hatası: {str(e)}",
                            raw_value=raw_val
                        )
                    )

            if row_has_error:
                failed_rows_count += 1
            else:
                transformed_rows_count += 1
                # Entity kayıtlarını oluştur
                if plant_data:
                    transformed_records.append(
                        TransformedRecord(
                            entity="solar_plant",
                            data=plant_data,
                            source_row_number=source_row_num
                        )
                    )
                if gen_data:
                    transformed_records.append(
                        TransformedRecord(
                            entity="daily_generation",
                            data=gen_data,
                            source_row_number=source_row_num
                        )
                    )

        wb.close()

        status = "FAILED" if failed_rows_count > 0 and transformed_rows_count == 0 else "SUCCESS"

        # 8. Sonucu Paketle
        result = TransformResult(
            status=status,
            source_file=file_path.name,
            mapping_key=mapping.key,
            generated_at=datetime.now().isoformat(),
            total_rows=total_rows,
            transformed_rows=transformed_rows_count,
            failed_rows=failed_rows_count,
            records=transformed_records,
            issues=issues
        )

        # 9. Sonuç JSON Dosyasını Kaydet
        self.ensure_transformed_directory()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        transformed_file_path = self.transformed_dir / f"transformed_{timestamp}.json"
        
        try:
            with open(transformed_file_path, "w", encoding="utf-8") as f:
                f.write(result.to_json())
            logger.info(f"Dönüştürülmüş veri raporu kaydedildi: {transformed_file_path}")
        except Exception as e:
            logger.error(f"Dönüşüm JSON raporu yazılırken hata: {e}")

        return result
