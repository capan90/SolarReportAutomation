from pathlib import Path
from typing import Callable, List, Optional, Pattern, Tuple

import openpyxl

from app.core.logger import setup_logger
from app.canonical.mapping_registry import MappingRegistry

logger = setup_logger("TurkishExcelWriter")

# Neden: Dinamik başlıklar (ör. Curve raporundaki "<santral>/Plant daily yield(kWh)")
# sabit sözlükle çevrilemez; regex deseni + üretici fonksiyon çifti kullanılır.
PatternRule = Tuple[Pattern, Callable]


class TurkishExcelWriter:
    """
    Neden: Portaldan indirilen İngilizce başlıklı Excel raporlarının, kullanıcıya
    gösterilecek Türkçe başlıklı bir kopyasını üretmek. Ham (raw archive) dosyaya
    asla dokunulmaz; kaynak dosya okunur, başlıkları çevrilmiş yeni bir kopya yazılır.
    Çeviri kaynağı canonical mapping'deki display_name_tr alanlarıdır (tek doğruluk kaynağı);
    dinamik başlıklar için pattern_rules ile regex tabanlı çeviri desteklenir.
    """

    def __init__(
        self,
        mapping_key: str = "isolar_yield_report_v1",
        pattern_rules: Optional[List[PatternRule]] = None,
    ):
        self.mapping_key = mapping_key
        self.pattern_rules = pattern_rules or []
        self.mapping_registry = MappingRegistry()

    def _build_header_translations(self) -> dict:
        """
        Neden: Mapping'deki source_column + alias'ların tamamını Türkçe görünen
        ada eşleyen sözlüğü kurmak (büyük/küçük harf duyarsız arama için lower key).
        """
        mapping = self.mapping_registry.get_mapping(self.mapping_key)
        if not mapping:
            raise ValueError(f"Mapping şablonu bulunamadı: {self.mapping_key}")

        translations = {}
        for m_field in mapping.mappings:
            if not m_field.display_name_tr:
                continue
            keys = [m_field.source_column] + list(m_field.source_aliases)
            for key in keys:
                translations[key.strip().lower()] = m_field.display_name_tr
        return translations

    def _translate_header(self, value: str, translations: dict) -> Optional[str]:
        """
        Neden: Önce birebir sözlük eşleşmesi denenir; bulunamazsa dinamik
        başlıklar için sıralı regex desenleri uygulanır. Çevrilemeyen başlık
        için None döner (hücre olduğu gibi bırakılır).
        """
        exact = translations.get(value.strip().lower())
        if exact:
            return exact
        for pattern, make_name in self.pattern_rules:
            m = pattern.match(value.strip())
            if m:
                return make_name(m)
        return None

    def save_turkish_copy(self, source_path: Path, output_path: Optional[Path] = None) -> Path:
        """
        Neden: Kaynak Excel'in başlık satırındaki bilinen İngilizce kolon adlarını
        Türkçe'ye çevirip '<ad>_TR.xlsx' kopyası olarak kaydetmek. output_path
        kaynakla aynı verilirse dosya yerinde çevrilir (manuel test çıktıları için).
        Bilinmeyen başlıklar olduğu gibi bırakılır; veri hücrelerine dokunulmaz.
        """
        source_path = Path(source_path)
        if output_path is None:
            output_path = source_path.with_name(f"{source_path.stem}_TR{source_path.suffix}")

        translations = self._build_header_translations()

        wb = openpyxl.load_workbook(str(source_path))
        translated_total = 0
        for ws in wb.worksheets:
            # Neden: Başlık satırı her zaman 1. satır olmayabilir (üstte metadata
            # satırı olabilir); ilk 5 satırda en çok bilinen başlığın geçtiği satırı seç.
            best_row, best_hits = None, 0
            for row_idx in range(1, min(6, ws.max_row + 1)):
                hits = sum(
                    1 for cell in ws[row_idx]
                    if cell.value and self._translate_header(str(cell.value), translations)
                )
                if hits > best_hits:
                    best_row, best_hits = row_idx, hits

            if not best_row:
                continue

            for cell in ws[best_row]:
                if cell.value is None:
                    continue
                tr_name = self._translate_header(str(cell.value), translations)
                if tr_name:
                    cell.value = tr_name
                    translated_total += 1

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        wb.close()
        logger.info(f"Türkçe kopya kaydedildi ({translated_total} başlık çevrildi): {output_path}")
        return output_path
